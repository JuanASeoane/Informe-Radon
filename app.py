# -*- coding: utf-8 -*-
"""
APP DE RADON - VERSION STREAMLIT
=================================
Migración de la app de escritorio (Tkinter + OpenCV) a Streamlit.

Conserva TODA la lógica original:
  - Gestión de Centros y Detectores en SQLite
  - Marcado del punto del detector sobre el plano
  - Generación de informe PDF (idéntica, con logo y cabecera)
  - Ajustes (técnico/empresa)

Cambios respecto a la versión Windows:
  - La cámara ya no usa cv2.VideoCapture (ventana propia de escritorio):
    usa st.camera_input, que abre la cámara nativa del navegador
    (funciona igual en PC como en el móvil Android).
  - El punto sobre el plano se marca con un clic usando el componente
    streamlit-image-coordinates.
  - Al generar el PDF aparece un botón "Enviar por WhatsApp" que usa la
    Web Share API de Android para abrir el diálogo nativo de compartir
    con el PDF ya adjunto.
  - Las fotos se guardan en Supabase Storage para persistencia.
"""

import os
import io
import re
import json
import html
import base64
import sqlite3
import zipfile
from datetime import datetime, date, timedelta

import streamlit as st
import streamlit.components.v1 as components
from PIL import Image, ImageDraw, ImageOps

# ============================================================
# SUPABASE
# ============================================================
try:
    from st_supabase_connection import SupabaseConnection
    from supabase import create_client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

try:
    from zoneinfo import ZoneInfo
    _ZONA_ESPANA = ZoneInfo("Europe/Madrid")
except Exception:
    # Si por lo que sea no está disponible la base de datos de zonas
    # horarias (raro, pero podría pasar en algún sistema muy pelado
    # sin el paquete "tzdata" instalado), se sigue funcionando con la
    # hora del propio servidor en vez de romper la app entera.
    _ZONA_ESPANA = None


def _ahora_espana() -> datetime:
    """La hora de "ahora" tal como la vería alguien en España (con el
    cambio de horario de verano/invierno correcto), en vez de la hora
    del servidor donde corre la app -que normalmente va en UTC, y por
    tanto puede ir 1 o 2 horas por detrás de la hora real de España-.
    Se usa siempre que se captura una fecha/hora "ahora mismo" pensada
    para que la vea o firme una persona (colocación, retirada, fecha
    del informe...)."""
    ahora = datetime.now(_ZONA_ESPANA) if _ZONA_ESPANA else datetime.now()
    return ahora.replace(tzinfo=None)




try:
    from streamlit_image_coordinates import streamlit_image_coordinates
    IMG_COORD_DISPONIBLE = True
except ImportError:
    IMG_COORD_DISPONIBLE = False

# Favicon personalizado (favicon.png debe estar en la misma carpeta que
# este script). Si por lo que sea no se encuentra, se usa el emoji de
# radiactividad como respaldo para que la app no falle al arrancar.


_carpeta_script = os.path.dirname(os.path.abspath(__file__))
_ruta_favicon = os.path.join(_carpeta_script, "favicon.png")
_icono_pagina = _ruta_favicon if os.path.exists(_ruta_favicon) else "☢️"

st.set_page_config(page_title="Detectores Rn", page_icon=_icono_pagina, layout="wide")

# Imagen de fondo de la app (fondo_app.jpg debe estar en la misma
# carpeta que este script). Se codifica en base64 para poder ponerla
# como fondo por CSS sin depender de servir el archivo como URL. Si no
# se encuentra, la app sigue funcionando sin fondo (solo el color
# oscuro de siempre).
_ruta_fondo = os.path.join(_carpeta_script, "fondo_app.jpg")
_fondo_css = ""
if os.path.exists(_ruta_fondo):
    with open(_ruta_fondo, "rb") as _f_fondo:
        _fondo_b64 = base64.b64encode(_f_fondo.read()).decode("utf-8")
    _fondo_css = f"""
    [data-testid="stAppViewContainer"] {{
        background-image:
            linear-gradient(rgba(8, 12, 18, 0.48), rgba(8, 12, 18, 0.55)),
            url("data:image/jpeg;base64,{_fondo_b64}");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        background-repeat: no-repeat;
    }}
    [data-testid="stHeader"] {{
        background-color: rgba(0, 0, 0, 0) !important;
    }}
    """

# Ajustes visuales generales: texto en negrita, títulos más pequeños,
# texto pequeño (captions/etiquetas) más grande, botones destacados en
# naranja, y fondo gris oscuro en todos los paneles/tarjetas.
st.markdown(
    """
    <style>
    """ + _fondo_css + """
    html, body, [class*="css"] {
        font-weight: 700 !important;
    }

    /* Títulos más pequeños (pantalla de inicio y el resto de pantallas) */
    h1 { font-size: 1.55rem !important; }
    h2 { font-size: 1.25rem !important; }
    h3 { font-size: 1.1rem !important; }

    /* Texto pequeño (captions, ayudas, pies de foto...) más grande y legible */
    [data-testid="stCaptionContainer"], .stCaption, small {
        font-size: 1rem !important;
    }
    label, .stMarkdown p {
        font-size: 1.02rem !important;
    }

    /* El texto de "Selecciona..." (placeholder) de los desplegables:
       forzado a un gris oscuro siempre, sin depender del tema, porque
       estos campos suelen tener fondo claro (blanco/gris/rosa) y el
       texto blanco del tema oscuro se volvía invisible encima. */
    div[data-baseweb="select"] [class*="placeholder"] {
        color: #4a4a4a !important;
    }

    /* Texto e etiquetas en blanco SIEMPRE, sin depender de que el archivo
       .streamlit/config.toml (tema oscuro) esté presente. Si ese archivo
       falta (p.ej. por no haberlo subido a GitHub, al ser una carpeta
       oculta que empieza por un punto), Streamlit usa su tema claro
       por defecto y estos textos se verían en negro sobre fondo
       oscuro. Estas reglas hacen que se vean bien en cualquier caso. */
    label, label p, label span,
    .stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown span,
    [data-testid="stCaptionContainer"], .stCaption,
    [data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
    [data-testid="stFileUploaderDropzoneInstructions"] span,
    [data-testid="stFileUploaderDropzoneInstructions"] small,
    h1, h2, h3, h4, h5, h6 {
        color: #f5f5f5 !important;
    }
    /* El fondo general de la app y el texto normal, por si tampoco
       está el tema oscuro aplicado (fondo oscuro por defecto). */
    [data-testid="stAppViewContainer"], [data-testid="stApp"], .stApp {
        color: #f5f5f5 !important;
    }
    /* La celda que se abre para editar un valor en una tabla
       editable (st.data_editor) NO es un input normal: es un
       elemento de "glide-data-grid" (clase gdg-input) insertado en
       un "portal" aparte. Sin esta regla heredaría el texto en
       blanco de más arriba y, sobre su fondo claro, no se vería lo
       que se escribe. */
    .gdg-input {
        background-color: #ffffff !important;
        color: #000000 !important;
    }

    /* Botones "primary" (acción principal: Abrir, Guardar...) en verde */
    button[kind="primary"], button[kind="primaryFormSubmit"] {
        background-color: #F5A623 !important;
        border-color: #F5A623 !important;
        color: #000000 !important;
    }
    button[kind="primary"]:hover, button[kind="primaryFormSubmit"]:hover {
        background-color: #d68f10 !important;
        border-color: #d68f10 !important;
        color: #000000 !important;
    }

    /* Botones "secondary" y "tertiary" (acciones neutras: Nuevo centro,
       Importar centro, Nuevo detector, Añadir plano, Generar
       documentos...) con estilo neutro de contorno, no relleno de
       color, para que solo la acción principal (amarilla) destaque. */
    button[kind="secondary"], button[kind="secondaryFormSubmit"],
    button[kind="tertiary"], button[kind="tertiaryFormSubmit"],
    .stDownloadButton button {
        background-color: #262626 !important;
        border: 1px solid #666666 !important;
        color: #ffffff !important;
    }
    button[kind="secondary"]:hover, button[kind="secondaryFormSubmit"]:hover,
    button[kind="tertiary"]:hover, button[kind="tertiaryFormSubmit"]:hover,
    .stDownloadButton button:hover {
        background-color: #333333 !important;
        border-color: #999999 !important;
        color: #ffffff !important;
    }

    /* Botón "Eliminar" (marcado con un div justo antes, ver código
       Python): con borde rojo y texto rojo, sin relleno, para que se
       vea claramente como una acción de peligro/secundaria. Se ancla
       al nivel exacto "stElementContainer" (sin ">", el marcador queda
       varios niveles por debajo) para no afectar por error a otros
       botones vecinos en la misma fila de columnas (p.ej. "Abrir"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] button {
        background-color: transparent !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
        width: auto !important;
        min-width: 0 !important;
        display: inline-flex !important;
        opacity: 0.9;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] button:hover {
        background-color: rgba(245, 166, 35, 0.15) !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
        opacity: 1;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-eliminar) + div[data-testid="stElementContainer"] {
        display: flex !important;
        justify-content: flex-start !important;
    }

    /* Acordeones de la pantalla del informe final: fondo rosa si le
       falta algo, gris si está completo (igual que el resto de
       campos de la app). Entre el marcador y el stExpander real hay
       un "stLayoutWrapper" de por medio. */
    div[data-testid="stElementContainer"]:has(div.marcador-acordeon-rosa) + div[data-testid="stLayoutWrapper"] div[data-testid="stExpander"] summary {
        background-color: #FBE1E6 !important;
        border-radius: 8px !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-acordeon-gris) + div[data-testid="stLayoutWrapper"] div[data-testid="stExpander"] summary {
        background-color: #C9C9C9 !important;
        border-radius: 8px !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-acordeon-rosa) + div[data-testid="stLayoutWrapper"] div[data-testid="stExpander"] summary p,
    div[data-testid="stElementContainer"]:has(div.marcador-acordeon-gris) + div[data-testid="stLayoutWrapper"] div[data-testid="stExpander"] summary p {
        color: #222 !important;
        font-weight: 700 !important;
    }

    /* Botón "Generar documentos": lo más pequeño posible, alineado a
       la izquierda (ya va en una columna estrecha). Fondo negro con
       borde y letra amarilla, igual que el resto de acciones
       principales de la app. Se ancla al nivel exacto
       "stElementContainer" (sin restringir a hijo directo, el marcador
       queda varios niveles por debajo) para no afectar a otros
       botones de columnas vecinas. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] button {
        font-size: 0.75rem !important;
        padding: 0.15rem 0.6rem !important;
        min-height: 1.7rem !important;
        height: 1.7rem !important;
        line-height: 1 !important;
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-generar) + div[data-testid="stElementContainer"] {
        display: flex !important;
        justify-content: flex-start !important;
    }

    /* Cabecera del bloque desplegable actualmente abierto: texto en
       amarillo (igual que el logo) para distinguirlo del resto. Se
       usa ":has(div...)" sin ">" (descendiente, no solo hijo directo)
       porque el marcador queda varios niveles por debajo del div que
       realmente tiene como hermano el botón. */
    div:has(div.marcador-bloque-activo) + div button {
        color: #F5A623 !important;
    }

    /* Botón "Importar centro": fondo oscuro con las letras en blanco
       (igual que el "+" de "Nuevo centro"), con borde amarillo. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-importar) + div[data-testid="stElementContainer"] button {
        background-color: #262626 !important;
        border: 1px solid #F5A623 !important;
        color: #ffffff !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-importar) + div[data-testid="stElementContainer"] button:hover {
        background-color: #333333 !important;
        border-color: #F5A623 !important;
        color: #ffffff !important;
    }

    /* Botón "Nuevo centro": mismo fondo oscuro de siempre, con borde
       amarillo (igual que "Importar centro"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-centro) + div[data-testid="stElementContainer"] button {
        border: 1px solid #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-centro) + div[data-testid="stElementContainer"] button:hover {
        border-color: #F5A623 !important;
    }

    /* Botones "Nuevo centro" / "Importar centro": cuando su formulario
       está abierto, el texto se pone en naranja (y vuelve a su color
       normal en cuanto se abre el otro, ya que solo uno lleva este
       marcador a la vez). Se restringe al nivel "stElementContainer"
       (con ">") porque, al estar en columnas, un selector más genérico
       también coincide con la COLUMNA vecina y pintaría el botón
       equivocado. */
    div[data-testid="stElementContainer"]:has(div.marcador-activo-naranja) + div[data-testid="stElementContainer"] button,
    div[data-testid="stElementContainer"]:has(div.marcador-activo-naranja) + div[data-testid="stElementContainer"] button:hover {
        color: #F5A623 !important;
    }

    /* Botones "Volver a Centros" / "Volver al centro": letra en azul
       claro. */
    /* Botones "Volver..." (Volver a Centros, Volver al centro,
       Volver): fondo azul claro y letra negra. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-azul-claro) + div[data-testid="stElementContainer"] button {
        background-color: #93C5FD !important;
        border: 1px solid #93C5FD !important;
        color: #000000 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-azul-claro) + div[data-testid="stElementContainer"] button:hover {
        background-color: #7CAEEB !important;
        border-color: #7CAEEB !important;
        color: #000000 !important;
    }

    /* Casilla de "Nº personas" en categorías profesionales: ancho
       máximo para que solo quepan 2 dígitos (en vez de ocupar todo el
       ancho disponible, sobre todo notorio en móvil). */
    div[data-testid="stElementContainer"]:has(div.marcador-num-personas) + div[data-testid="stElementContainer"] div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]) {
        max-width: 8rem !important;
    }

    /* Todos los botones dentro de "Imagen exterior" (Subir archivo,
       Activar/Tomar cámara, y el de capturar dentro del propio visor):
       fondo negro, borde y letra amarilla. Se restringe a la columna
       exacta ("stColumn") que contiene el marcador; un selector sin
       esa restricción coincide con TODA la página (cualquier ancestro
       que también contenga el marcador en su árbol) y pinta botones
       completamente ajenos. */
    div[data-testid="stColumn"]:has(div.marcador-imagen-exterior) button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stColumn"]:has(div.marcador-imagen-exterior) button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Dentro del formulario de importar: el botón de subir archivo y
       el de confirmar "Importar", en naranja sobre fondo negro. */
    div:has(div.marcador-uploader-importar) + div button,
    div:has(div.marcador-btn-confirmar-importar) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-uploader-importar) + div button:hover,
    div:has(div.marcador-btn-confirmar-importar) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-crear-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-crear-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-ajustes) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-ajustes) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-guardar-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Botón "➕ Añadir" (categorías profesionales): fondo negro, borde
       y letra amarilla. Está dentro de columnas, así que se restringe
       al nivel exacto "stElementContainer" para no afectar a otros
       botones cercanos (p.ej. "Eliminar seleccionadas"). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-anadir-categoria) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-anadir-categoria) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Planos del centro": botones "Añadir plano", "Subir archivo" (del
       selector de imagen) y "Guardar plano", todos en fondo negro con
       borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    /* "Añadir plano" iba con el texto muy pegado al borde: un poco
       más de espacio a los lados (afecta a los 3 botones de este
       mismo marcador: Añadir plano, Subir archivo y Guardar plano). */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-plano-amarillo) + div[data-testid="stElementContainer"] button {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }

    /* "Detectores colocados": Nuevo detector y Abrir detector, en
       fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
        padding-left: 1.1rem !important;
        padding-right: 1.1rem !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-nuevo-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-abrir-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-abrir-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Retirada de detectores": Capturar fecha y hora / Guardar, en
       fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-retirada-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-retirada-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Guardar detector" (al final de la pantalla del detector): fondo
       negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-guardar-detector) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-guardar-detector) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Títulos de cada apartado de anexo (I, II, III, IV): todos con
       el mismo tamaño y en amarillo. No vale hacerlo con un "style"
       en línea porque st.markdown sanea el HTML y le quita los
       "!important" a los estilos en línea; por eso se hace con el
       marcador + esta regla, igual que en el resto de la app. */
    div[data-testid="stElementContainer"]:has(div.marcador-titulo-anexo) + div[data-testid="stElementContainer"] p {
        color: #F5A623 !important;
        font-weight: 700 !important;
    }

    /* Casillas de "Categorías profesionales": el cuadrito de marcar
       apenas se veía (sin contraste con el fondo oscuro), así que se
       le pone un borde amarillo bien visible; y el texto de la
       categoría se pone al doble de grande para que se lea mejor. */
    div[data-testid="stElementContainer"]:has(div.marcador-checkbox-categoria) + div[data-testid="stElementContainer"] div[data-testid="stCheckbox"] label > div:not([data-testid]) {
        border: 2px solid #F5A623 !important;
        border-radius: 4px !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-checkbox-categoria) + div[data-testid="stElementContainer"] div[data-testid="stWidgetLabel"] p {
        font-size: 1.5em !important;
    }

    /* Tabla de resultados del informe (Resultado/Incertidumbre): en
       móvil, Streamlit apila cualquier st.columns() en cuanto el
       contenido no cabe (cada columna pasa a ocupar el 100% del
       ancho y salta de línea). Aquí se fuerza a que estas dos
       columnas concretas se mantengan siempre en la misma fila,
       repartiéndose el ancho a la mitad cada una, para que de verdad
       se vea como una tabla de dos columnas también en el móvil. */
    div[data-testid="stElementContainer"]:has(div.marcador-tabla-resultado-fila) + div[data-testid="stLayoutWrapper"] div[data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 0.6rem !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-tabla-resultado-fila) + div[data-testid="stLayoutWrapper"] div[data-testid="stColumn"] {
        width: 50% !important;
        min-width: 0 !important;
        flex: 1 1 0 !important;
    }

    div[data-testid="stElementContainer"]:has(div.marcador-btn-informe-completo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-informe-completo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Descargar PDF / Excel / fotos, y Seleccionar todas / Quitar
       selección: fondo negro con borde y letra amarilla. */
    div[data-testid="stElementContainer"]:has(div.marcador-btn-descarga-amarillo) + div[data-testid="stElementContainer"] button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div[data-testid="stElementContainer"]:has(div.marcador-btn-descarga-amarillo) + div[data-testid="stElementContainer"] button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* Título grande de la pantalla de inicio ("Detectores de Radón"),
       al doble de tamaño que un título normal. Usa una clase (en vez
       de estilo en línea) porque Streamlit elimina por completo el
       atributo style="" de cualquier HTML si contiene "!important". */
    p.titulo-home {
        color: #F5A623 !important;
        font-size: 1.55rem !important;
        font-weight: 800 !important;
        line-height: 1.2 !important;
        margin: 0.5rem 0 !important;
    }

    /* Título del nombre del centro (pantalla del centro), al doble de
       tamaño que un título normal. */
    p.titulo-centro {
        color: #F5A623 !important;
        font-size: 1.55rem !important;
        font-weight: 800 !important;
        line-height: 1.2 !important;
        margin: 0.5rem 0 !important;
    }

    /* Subtítulos de sección en amarillo (p.ej. "Centros registrados"),
       con el mismo truco de clase por la limitación de style="" con
       !important explicada arriba. */
    p.subtitulo-amarillo {
        color: #F5A623 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        margin: 0.8rem 0 0.3rem 0 !important;
    }

    /* Botón "Abrir centro": fondo negro con letras naranjas (en vez
       del amarillo/negro normal de los botones "primary"). */
    div:has(div.marcador-btn-abrir-centro) + div button {
        background-color: #000000 !important;
        border: 1px solid #F5A623 !important;
        color: #F5A623 !important;
    }
    div:has(div.marcador-btn-abrir-centro) + div button:hover {
        background-color: #1a1a1a !important;
        border-color: #F5A623 !important;
        color: #F5A623 !important;
    }

    /* "Ventanas" (tarjetas, expanders, formularios) con fondo gris oscuro */
    div[data-testid="stVerticalBlockBorderWrapper"],
    div[data-testid="stExpander"],
    div[data-testid="stForm"] {
        background-color: #262626 !important;
        border-radius: 10px;
    }

    /* Títulos de cada ventana/panel en naranja */
    h1, h2, h3 {
        color: #f5f5f5 !important;
    }

    /* Cabecera de los desplegables tipo acordeón (expanders) en blanco,
       para que combine con las casillas (p.ej. "➕ Nuevo centro") */
    div[data-testid="stExpander"] summary {
        background-color: #ffffff !important;
        border-radius: 8px !important;
    }

    /* Ocultar el texto "Press Enter to apply" que Streamlit muestra
       bajo las casillas de texto mientras se escribe */
    [data-testid="InputInstructions"] {
        display: none !important;
    }

    /* Casillas de texto y desplegables: SOLO el recuadro donde se
       escribe/selecciona queda en blanco con letra negra, un 25% más
       grande (1rem -> 1.25rem) y con esquinas rectas (sin redondear).
       La etiqueta de encima (el texto que describe el campo, con
       testid "stWidgetLabel") se excluye a propósito para que quede
       sobre el fondo gris de la ventana.
       Se usa "*" para pintar TODO lo de dentro del recuadro (incluida
       la parte del desplegable de Streamlit, que anida varios niveles
       de <div> internos) y luego se fuerza la etiqueta a transparente
       para que no quede afectada. */
    div[data-testid="stTextInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stTextInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stTextArea"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stTextArea"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stNumberInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stDateInput"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stDateInput"] > div:not([data-testid="stWidgetLabel"]) *,
    div[data-testid="stSelectbox"] > div:not([data-testid="stWidgetLabel"]),
    div[data-testid="stSelectbox"] > div:not([data-testid="stWidgetLabel"]) * {
        background-color: #ffffff !important;
        color: #000000 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stTextArea"] textarea,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stDateInput"] input,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] * {
        background-color: #ffffff !important;
        color: #000000 !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
    }
    /* Los campos "disabled" (de solo lectura, como la vista previa del
       tipo de zona en el informe) reciben en algunos móviles (Chrome/
       Safari en Android e iOS) un color de texto propio y más claro
       para los campos deshabilitados, que las reglas de arriba no
       llegan a pisar del todo (algunos navegadores usan la propiedad
       -webkit-text-fill-color específicamente para esto, que manda
       más que "color" a secas). Se fuerzan aquí las dos, y se quita
       cualquier atenuación, para que el texto se siga leyendo bien
       aunque el campo esté deshabilitado. */
    div[data-testid="stTextInput"] input:disabled,
    div[data-testid="stTextArea"] textarea:disabled,
    div[data-testid="stNumberInput"] input:disabled {
        background-color: #ffffff !important;
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
        opacity: 1 !important;
    }
    /* La etiqueta de encima NO se toca: se fuerza a volver al color de
       texto normal de la app, sobre el fondo gris de la ventana. */
    div[data-testid="stWidgetLabel"],
    div[data-testid="stWidgetLabel"] * {
        background-color: transparent !important;
    }
    /* Menú desplegable del selectbox al abrirlo: se renderiza en un
       "popover" flotante fuera del recuadro, así que se cubre con
       TODOS los posibles selectores que puede usar Streamlit/BaseWeb
       (testid, data-baseweb de popover/menu/menu-item, y roles ARIA
       listbox/option) y con "*" para llegar a todas las capas
       internas de texto/fondo, sea cual sea la estructura exacta. */
    div[data-baseweb="popover"],
    div[data-baseweb="popover"] *,
    div[data-baseweb="menu"],
    div[data-baseweb="menu"] *,
    li[data-baseweb="menu-item"],
    li[data-baseweb="menu-item"] *,
    ul[data-testid="stSelectboxVirtualDropdown"],
    ul[data-testid="stSelectboxVirtualDropdown"] *,
    [role="listbox"],
    [role="listbox"] *,
    [role="option"],
    [role="option"] * {
        background-color: #ffffff !important;
        color: #000000 !important;
        border-radius: 8px !important;
    }
    div[data-baseweb="popover"] li,
    li[data-baseweb="menu-item"],
    ul[data-testid="stSelectboxVirtualDropdown"] li,
    [role="listbox"] li,
    [role="option"] {
        font-size: 1.1rem !important;
    }
    /* Opción resaltada al pasar el ratón/dedo: un gris muy claro para
       distinguirla, siempre con texto negro */
    div[data-baseweb="popover"] li:hover,
    li[data-baseweb="menu-item"]:hover,
    ul[data-testid="stSelectboxVirtualDropdown"] li:hover,
    [role="listbox"] li:hover,
    [role="option"]:hover,
    div[data-baseweb="popover"] li[aria-selected="true"],
    li[data-baseweb="menu-item"][aria-selected="true"],
    [role="listbox"] li[aria-selected="true"],
    [role="option"][aria-selected="true"] {
        background-color: #eeeeee !important;
        color: #000000 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# BASE DE DATOS  (idéntica a la app de escritorio)
# ============================================================

DB_NAME = "radon_data.db"


def get_data_dir():
    """Carpeta donde se guardan la BD y las imágenes.

    Se usa una carpeta junto al propio script para que funcione igual
    en local, en un servidor o en Streamlit Community Cloud. Si prefieres
    guardar los datos en el perfil del usuario como hacía la app de
    Windows, cambia la línea siguiente por:
        data_dir = os.path.join(os.path.expanduser("~"), "RadonApp")
    """
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RadonApp_data")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def get_db_path():
    return os.path.join(get_data_dir(), DB_NAME)


# ============================================================
# CONEXIÓN A SUPABASE
# ============================================================

def get_supabase_connection():
    """Obtiene la conexión a Supabase usando st.connection con mejor manejo de errores."""
    if not SUPABASE_AVAILABLE:
        st.warning("⚠️ Supabase no está disponible. Instala las dependencias.")
        return None
    
    try:
        # Verificar que los secrets existen
        if "connections" not in st.secrets:
            st.error("❌ 'connections' no está en secrets")
            st.info("💡 Revisa que secrets.toml tenga la sección [connections.supabase]")
            return None
        
        if "supabase" not in st.secrets["connections"]:
            st.error("❌ 'supabase' no está en connections")
            st.info("💡 Las claves disponibles son: " + ", ".join(st.secrets["connections"].keys()))
            return None
        
        # Intentar conectar usando st.connection
        return st.connection("supabase", type=SupabaseConnection)
        
    except Exception as e:
        st.error(f"❌ Error conectando a Supabase: {e}")
        return None

def get_supabase_client():
    """Obtiene el cliente de Supabase para operaciones avanzadas."""
    if not SUPABASE_AVAILABLE:
        return None
    try:
        url = st.secrets["connections.supabase"]["SUPABASE_URL"]
        key = st.secrets["connections.supabase"]["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        st.error(f"Error creando cliente Supabase: {e}")
        return None

def verificar_conexion_supabase():
    """Verifica que la conexión a Supabase funciona con más detalles."""
    if not SUPABASE_AVAILABLE:
        st.warning("⚠️ Supabase no disponible. Instala las dependencias.")
        return False
    
    try:
        # Verificar secrets
        if "connections" not in st.secrets:
            st.error("❌ 'connections' no está en secrets")
            st.info("💡 Revisa que secrets.toml tenga la sección [connections.supabase]")
            return False
        
        if "supabase" not in st.secrets["connections"]:
            st.error("❌ 'supabase' no está en connections")
            st.info("💡 Las claves disponibles son: " + ", ".join(st.secrets["connections"].keys()))
            return False
        
        url = st.secrets["connections"]["supabase"]["SUPABASE_URL"]
        key = st.secrets["connections"]["supabase"]["SUPABASE_KEY"]
        
        # Validar URL
        if not url.startswith("https://"):
            st.error("❌ La URL debe comenzar con 'https://'")
            return False
        
        if not url.endswith(".supabase.co"):
            st.error("❌ La URL debe terminar con '.supabase.co'")
            return False
        
        if url.endswith("/"):
            st.error("❌ La URL no debe tener barra al final")
            st.info("💡 Elimina la '/' final de la URL en secrets.toml")
            return False
        
        if not key or len(key) < 50:
            st.error("❌ La clave parece incompleta o incorrecta")
            return False
        
        conn = get_supabase_connection()
        if conn is None:
            st.error("❌ No se pudo crear la conexión")
            return False
        
        # Intentar una consulta simple
        try:
            result = conn.table("centros").select("*").limit(1).execute()
            st.success(f"✅ Conectado a Supabase ({len(result.data)} centros)")
            return True
        except Exception as e:
            error_msg = str(e)
            if "PGRST125" in error_msg:
                st.error(f"❌ Error de URL: {error_msg}")
                st.info("💡 Verifica que la URL en secrets.toml sea exactamente la Project URL de Supabase")
            elif "permission denied" in error_msg.lower():
                st.error("❌ Error de permisos: Verifica las políticas RLS en Supabase")
            elif "does not exist" in error_msg.lower() or "relation" in error_msg.lower():
                st.error("❌ La tabla 'centros' no existe en Supabase. Créala primero.")
            elif "JWT" in error_msg:
                st.error("❌ Error de autenticación: La clave ANON puede ser incorrecta")
            else:
                st.error(f"❌ Error: {error_msg}")
            return False
            
    except Exception as e:
        st.error(f"❌ Error general: {e}")
        return False


def sync_to_supabase(operacion, tabla, datos, id_valor=None):
    """Sincroniza operaciones con Supabase.
    
    Args:
        operacion: 'insert', 'update', 'delete'
        tabla: nombre de la tabla en Supabase
        datos: dict con los datos (para insert/update)
        id_valor: valor del ID (para update/delete)
    """
    if not SUPABASE_AVAILABLE:
        return None
    
    conn = get_supabase_connection()
    if not conn:
        return None
    
    try:
        if operacion == 'insert':
            return conn.table(tabla).insert(datos).execute()
        elif operacion == 'update':
            return conn.table(tabla).update(datos).eq('id', id_valor).execute()
        elif operacion == 'delete':
            return conn.table(tabla).delete().eq('id', id_valor).execute()
    except Exception as e:
        # No mostramos error para no interrumpir el flujo
        # si Supabase no está disponible
        pass
    return None

# ============================================================
# ALMACENAMIENTO EN SUPABASE (BUCKETS)
# ============================================================

def subir_archivo_supabase(bucket, ruta_local, ruta_remota):
    """Sube un archivo a un bucket de Supabase Storage."""
    if not SUPABASE_AVAILABLE:
        return None
    
    conn = get_supabase_connection()
    if not conn:
        return None
    
    try:
        with open(ruta_local, 'rb') as f:
            return conn.storage.from_(bucket).upload(ruta_remota, f)
    except Exception as e:
        return None

def obtener_url_publica_supabase(bucket, ruta_remota):
    """Obtiene la URL pública de un archivo en Supabase Storage."""
    if not SUPABASE_AVAILABLE:
        return None
    
    conn = get_supabase_connection()
    if not conn:
        return None
    
    try:
        return conn.storage.from_(bucket).get_public_url(ruta_remota)
    except Exception as e:
        return None

def es_url_supabase(ruta):
    """Verifica si una ruta es una URL de Supabase Storage."""
    if not ruta:
        return False
    return ruta.startswith('http') and '.supabase.co' in ruta

def descargar_desde_supabase(url, destino_local):
    """Descarga un archivo desde Supabase Storage a una ruta local."""
    if not SUPABASE_AVAILABLE or not es_url_supabase(url):
        return False
    
    conn = get_supabase_connection()
    if not conn:
        return False
    
    try:
        # Extraer bucket y ruta de la URL
        # Ejemplo: https://ctytennxlftnhxdgcbtj.supabase.co/storage/v1/object/public/fotos/situacion/foto.jpg
        partes = url.split('/public/')
        if len(partes) > 1:
            bucket = partes[1].split('/')[0]
            ruta_remota = '/'.join(partes[1].split('/')[1:])
            data = conn.storage.from_(bucket).download(ruta_remota)
            with open(destino_local, 'wb') as f:
                f.write(data)
            return True
    except Exception as e:
        st.warning(f"⚠️ No se pudo descargar la imagen: {e}")
        return False
    return False

# ============================================================
# FUNCIONES DE MIGRACIÓN
# ============================================================

def migrar_centros_a_supabase():
    """Migra todos los centros de SQLite a Supabase."""
    if not SUPABASE_AVAILABLE:
        st.error("❌ Supabase no está disponible.")
        return False
    
    conn = get_supabase_connection()
    if not conn:
        return False
    
    try:
        # Obtener centros de SQLite
        conn_sqlite = sqlite3.connect(get_db_path())
        c = conn_sqlite.cursor()
        c.execute("SELECT id, nombre, zona, fecha_medicion, imagen_exterior_path, tecnico, direccion, tipo_centro FROM centros")
        centros = c.fetchall()
        conn_sqlite.close()
        
        if not centros:
            st.info("No hay centros para migrar.")
            return True
        
        migrados = 0
        for centro in centros:
            try:
                # Verificar si ya existe en Supabase
                existing = conn.table("centros").select("id").eq("id", centro[0]).execute()
                
                if not existing.data:
                    # Si la imagen exterior es una URL de Supabase, la dejamos como está
                    imagen_path = centro[4]
                    if imagen_path and not es_url_supabase(imagen_path) and os.path.exists(imagen_path):
                        # Subir la imagen a Supabase Storage
                        try:
                            with open(imagen_path, 'rb') as f:
                                nombre_img = os.path.basename(imagen_path)
                                ruta_remota = f"exterior/{centro[0]}_{nombre_img}"
                                conn.storage.from_("imagenes").upload(ruta_remota, f)
                                imagen_path = conn.storage.from_("imagenes").get_public_url(ruta_remota)
                        except Exception:
                            pass  # Si falla, dejar la ruta local
                    
                    conn.table("centros").insert({
                        "id": centro[0],
                        "nombre": centro[1],
                        "zona": centro[2],
                        "fecha_medicion": centro[3],
                        "imagen_exterior_path": imagen_path,
                        "tecnico": centro[5] if len(centro) > 5 else "",
                        "direccion": centro[6] if len(centro) > 6 else "",
                        "tipo_centro": centro[7] if len(centro) > 7 else ""
                    }).execute()
                    migrados += 1
            except Exception:
                pass
        
        if migrados > 0:
            st.success(f"✅ {migrados} centros migrados a Supabase")
        else:
            st.info("ℹ️ No se migraron centros nuevos (ya existían en Supabase)")
        return True
        
    except Exception as e:
        st.error(f"❌ Error en la migración: {e}")
        return False

def migrar_detectores_a_supabase():
    """Migra todos los detectores de SQLite a Supabase."""
    if not SUPABASE_AVAILABLE:
        st.error("❌ Supabase no está disponible.")
        return False
    
    conn = get_supabase_connection()
    if not conn:
        return False
    
    try:
        conn_sqlite = sqlite3.connect(get_db_path())
        c = conn_sqlite.cursor()
        c.execute("SELECT * FROM detectores")
        detectores = c.fetchall()
        conn_sqlite.close()
        
        if not detectores:
            st.info("No hay detectores para migrar.")
            return True
        
        migrados = 0
        for det in detectores:
            try:
                existing = conn.table("detectores").select("id").eq("id", det[0]).execute()
                
                if not existing.data:
                    # Procesar fotos
                    foto_sit = det[9] if len(det) > 9 else None
                    foto_det = det[10] if len(det) > 10 else None
                    plano_path = det[6] if len(det) > 6 else None
                    
                    # Subir fotos a Supabase Storage si son rutas locales
                    for campo, bucket, carpeta in [
                        (foto_sit, "fotos", "situacion"),
                        (foto_det, "fotos", "detectores"),
                        (plano_path, "planos", "planos")
                    ]:
                        if campo and not es_url_supabase(campo) and os.path.exists(campo):
                            try:
                                with open(campo, 'rb') as f:
                                    nombre_img = os.path.basename(campo)
                                    ruta_remota = f"{carpeta}/{det[0]}_{nombre_img}"
                                    conn.storage.from_(bucket).upload(ruta_remota, f)
                                    url = conn.storage.from_(bucket).get_public_url(ruta_remota)
                                    if campo == foto_sit:
                                        foto_sit = url
                                    elif campo == foto_det:
                                        foto_det = url
                                    elif campo == plano_path:
                                        plano_path = url
                            except Exception:
                                pass  # Si falla, dejar la ruta local
                    
                    conn.table("detectores").insert({
                        "id": det[0],
                        "centro_id": det[1],
                        "planta": det[2] if len(det) > 2 else "",
                        "sala": det[3] if len(det) > 3 else "",
                        "fecha": det[4] if len(det) > 4 else "",
                        "detector_codigo": det[5] if len(det) > 5 else "",
                        "plano_path": plano_path,
                        "punto_x": det[7] if len(det) > 7 else -1,
                        "punto_y": det[8] if len(det) > 8 else -1,
                        "foto_situacion_path": foto_sit,
                        "foto_detector_path": foto_det,
                        "fecha_creacion": det[11] if len(det) > 11 else "",
                        "codigo_sala": det[12] if len(det) > 12 else "",
                        "profesionales_sala": det[13] if len(det) > 13 else "",
                        "hora_colocacion": det[14] if len(det) > 14 else "",
                        "turno_trabajo": det[15] if len(det) > 15 else "",
                        "nivel": det[16] if len(det) > 16 else "",
                        "plano_centro_id": det[17] if len(det) > 17 else None,
                        "fecha_retirada_real": det[18] if len(det) > 18 else "",
                        "hora_retirada_real": det[19] if len(det) > 19 else "",
                        "resultado_bq_m3": det[20] if len(det) > 20 else None,
                        "incertidumbre": det[21] if len(det) > 21 else ""
                    }).execute()
                    migrados += 1
            except Exception:
                pass
        
        if migrados > 0:
            st.success(f"✅ {migrados} detectores migrados a Supabase")
        else:
            st.info("ℹ️ No se migraron detectores nuevos (ya existían en Supabase)")
        return True
        
    except Exception as e:
        st.error(f"❌ Error en la migración: {e}")
        return False

def migrar_datos_a_supabase():
    """Función que migra TODOS los datos a Supabase."""
    st.markdown("---")
    st.markdown("### ☁️ Migrar datos a Supabase")
    
    if not SUPABASE_AVAILABLE:
        st.error("⚠️ Supabase no está configurado. Instala las dependencias.")
        return
    
    if verificar_conexion_supabase():
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Migrar centros a Supabase", type="primary", use_container_width=True):
                with st.spinner("Migrando centros..."):
                    migrar_centros_a_supabase()
        with col2:
            if st.button("🔄 Migrar detectores a Supabase", type="primary", use_container_width=True):
                with st.spinner("Migrando detectores..."):
                    migrar_detectores_a_supabase()


def init_db():
    db_path = get_db_path()

    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS centros
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  nombre TEXT, zona TEXT, fecha_medicion TEXT, imagen_exterior_path TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS detectores
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER,
                  planta TEXT, sala TEXT, fecha TEXT, detector_codigo TEXT,
                  plano_path TEXT, punto_x REAL, punto_y REAL,
                  foto_situacion_path TEXT, foto_detector_path TEXT,
                  fecha_creacion TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (id INTEGER PRIMARY KEY CHECK (id=1), tecnico TEXT)''')
    c.execute("INSERT OR IGNORE INTO settings (id, tecnico) VALUES (1, '')")
    # Planos del centro (sin límite de cantidad), cada uno con su nombre
    # (p.ej. "Planta baja", "Planta 1"...). El punto rojo de cada
    # detector se guarda en el propio detector (punto_x/punto_y),
    # apuntando a uno de estos planos mediante plano_centro_id.
    c.execute('''CREATE TABLE IF NOT EXISTS planos_centro
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER, nombre TEXT, ruta TEXT, orden INTEGER)''')

    # Categorías profesionales expuestas en el centro (p.ej. "Enfermería"
    # -> 8 personas), con número de personas expuestas de cada una.
    c.execute('''CREATE TABLE IF NOT EXISTS categorias_centro
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  centro_id INTEGER, categoria TEXT, num_personas INTEGER)''')

    # --- Migraciones: añadir columnas nuevas sin perder los datos ya
    # guardados, por si la base de datos viene de una versión anterior. ---
    c.execute("PRAGMA table_info(centros)")
    columnas_centros = {fila[1] for fila in c.fetchall()}
    if "tecnico" not in columnas_centros:
        c.execute("ALTER TABLE centros ADD COLUMN tecnico TEXT DEFAULT ''")
    if "direccion" not in columnas_centros:
        c.execute("ALTER TABLE centros ADD COLUMN direccion TEXT DEFAULT ''")
    if "tipo_centro" not in columnas_centros:
        c.execute("ALTER TABLE centros ADD COLUMN tipo_centro TEXT DEFAULT ''")

    c.execute("PRAGMA table_info(detectores)")
    columnas_detectores = {fila[1] for fila in c.fetchall()}
    if "codigo_sala" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN codigo_sala TEXT DEFAULT ''")
    if "profesionales_sala" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN profesionales_sala TEXT DEFAULT ''")
    if "hora_colocacion" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN hora_colocacion TEXT DEFAULT ''")
    if "turno_trabajo" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN turno_trabajo TEXT DEFAULT ''")
    if "nivel" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN nivel TEXT DEFAULT ''")
    if "plano_centro_id" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN plano_centro_id INTEGER")
    if "fecha_retirada_real" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN fecha_retirada_real TEXT DEFAULT ''")
    if "hora_retirada_real" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN hora_retirada_real TEXT DEFAULT ''")

    c.execute("PRAGMA table_info(categorias_centro)")
    columnas_categorias = {fila[1] for fila in c.fetchall()}
    if "turno" not in columnas_categorias:
        c.execute("ALTER TABLE categorias_centro ADD COLUMN turno TEXT DEFAULT ''")

    if "resultado_bq_m3" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN resultado_bq_m3 REAL")
    if "incertidumbre" not in columnas_detectores:
        c.execute("ALTER TABLE detectores ADD COLUMN incertidumbre TEXT DEFAULT ''")

    c.execute("PRAGMA table_info(centros)")
    columnas_centros_extra = {fila[1] for fila in c.fetchall()}
    for columna_nueva in (
        "superficie_construida", "superficie_util", "num_plantas",
        "fecha_comunicacion_trab", "medio_comunicacion",
    ):
        if columna_nueva not in columnas_centros_extra:
            c.execute(f"ALTER TABLE centros ADD COLUMN {columna_nueva} TEXT DEFAULT ''")

    # Logotipo personalizado (dato global de la empresa, no de cada
    # centro): se guarda como archivo en la carpeta de datos, con la
    # ruta y el nombre original guardados en "settings".
    c.execute("PRAGMA table_info(settings)")
    columnas_settings = {fila[1] for fila in c.fetchall()}
    if "logo_path" not in columnas_settings:
        c.execute("ALTER TABLE settings ADD COLUMN logo_path TEXT DEFAULT ''")
    if "logo_nombre" not in columnas_settings:
        c.execute("ALTER TABLE settings ADD COLUMN logo_nombre TEXT DEFAULT ''")

    # Datos de la empresa (globales, no cambian de un centro a otro):
    # nombre de la empresa y CIF, con sus valores por defecto.
    c.execute("PRAGMA table_info(settings)")
    columnas_settings = {fila[1] for fila in c.fetchall()}
    if "empresa" not in columnas_settings:
        c.execute(
            "ALTER TABLE settings ADD COLUMN empresa TEXT DEFAULT 'Área Sanitaria da Coruña e Cee'"
        )
    if "cif" not in columnas_settings:
        c.execute("ALTER TABLE settings ADD COLUMN cif TEXT DEFAULT 'Q151569009B'")
    if "logo_laboratorio_path" not in columnas_settings:
        c.execute("ALTER TABLE settings ADD COLUMN logo_laboratorio_path TEXT")

    # Migración de planos antiguos: los detectores que ya tenían un
    # plano propio (plano_path, de antes de este cambio) y todavía no
    # están vinculados a un plano del centro (plano_centro_id vacío) se
    # migran automáticamente: se crea un plano del centro con esa
    # misma imagen (agrupando detectores que ya compartían el mismo
    # archivo) y se les asigna. Así no se pierde ningún plano ni punto
    # ya marcado.
    c.execute('''SELECT id, centro_id, plano_path FROM detectores
                 WHERE plano_centro_id IS NULL
                 AND plano_path IS NOT NULL AND plano_path != ''
                 ORDER BY centro_id, id''')
    pendientes = c.fetchall()
    if pendientes:
        cache_por_ruta = {}  # (centro_id, ruta) -> nuevo plano_centro_id
        contador_por_centro = {}
        for det_id, centro_id, ruta in pendientes:
            clave = (centro_id, ruta)
            if clave not in cache_por_ruta:
                contador_por_centro[centro_id] = contador_por_centro.get(centro_id, 0) + 1
                nombre_auto = f"Planta {contador_por_centro[centro_id]}"
                c.execute(
                    "INSERT INTO planos_centro (centro_id, nombre, ruta, orden) VALUES (?,?,?,?)",
                    (centro_id, nombre_auto, ruta, contador_por_centro[centro_id] - 1),
                )
                cache_por_ruta[clave] = c.lastrowid
            c.execute(
                "UPDATE detectores SET plano_centro_id=? WHERE id=?",
                (cache_por_ruta[clave], det_id),
            )

    conn.commit()
    conn.close()
   



def fetch_planos_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, nombre, ruta, orden FROM planos_centro WHERE centro_id=? ORDER BY orden, id",
               (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def get_plano_centro(plano_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, nombre, ruta, orden FROM planos_centro WHERE id=?", (plano_id,))
    row = c.fetchone()
    conn.close()
    return row


def insert_plano_centro(centro_id, nombre, ruta, orden):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("INSERT INTO planos_centro (centro_id, nombre, ruta, orden) VALUES (?,?,?,?)",
              (centro_id, nombre, ruta, orden))
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase (si está disponible)
    if SUPABASE_AVAILABLE:
        # Si la ruta es local, subir a Supabase Storage
        if ruta and not es_url_supabase(ruta) and os.path.exists(ruta):
            try:
                conn_sup = get_supabase_connection()
                if conn_sup:
                    nombre_archivo = os.path.basename(ruta)
                    ruta_remota = f"planos/{centro_id}_{rowid}_{nombre_archivo}"
                    with open(ruta, 'rb') as f:
                        conn_sup.storage.from_("planos").upload(ruta_remota, f)
                    ruta = conn_sup.storage.from_("planos").get_public_url(ruta_remota)
            except Exception:
                pass  # Si falla, dejar la ruta local
        
        sync_to_supabase("insert", "planos_centro", {
            "id": rowid,
            "centro_id": centro_id,
            "nombre": nombre,
            "ruta": ruta,
            "orden": orden
        })
    
    return rowid


def delete_plano_centro(plano_id):
    """Borra un plano del centro. Los detectores que lo tuvieran
    asignado se quedan sin plano ni punto marcado (ya no tendría
    sentido, al desaparecer la imagen sobre la que estaba)."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE detectores SET plano_centro_id=NULL, punto_x=-1, punto_y=-1 WHERE plano_centro_id=?",
              (plano_id,))
    c.execute("DELETE FROM planos_centro WHERE id=?", (plano_id,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("delete", "planos_centro", None, id_valor=plano_id)


def fetch_categorias_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, centro_id, categoria, num_personas, turno FROM categorias_centro WHERE centro_id=? ORDER BY id",
              (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def insert_categoria_centro(centro_id, categoria, num_personas, turno=""):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("INSERT INTO categorias_centro (centro_id, categoria, num_personas, turno) VALUES (?,?,?,?)",
              (centro_id, categoria, num_personas, turno))
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase (si está disponible)
    if SUPABASE_AVAILABLE:
        sync_to_supabase("insert", "categorias_centro", {
            "id": rowid,
            "centro_id": centro_id,
            "categoria": categoria,
            "num_personas": num_personas,
            "turno": turno
        })
    
    return rowid


def delete_categoria_centro(categoria_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM categorias_centro WHERE id=?", (categoria_id,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("delete", "categorias_centro", None, id_valor=categoria_id)


def crear_centro(nombre, zona="", tipo_centro=""):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute(
        "INSERT INTO centros (nombre, zona, fecha_medicion, imagen_exterior_path, direccion, tipo_centro) "
        "VALUES (?, ?, ?, NULL, '', ?)",
        (nombre, zona or "", _ahora_espana().strftime("%d/%m/%Y"), tipo_centro or "")
    )
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase (si está disponible)
    if SUPABASE_AVAILABLE:
        sync_to_supabase("insert", "centros", {
            "id": rowid,
            "nombre": nombre,
            "zona": zona,
            "fecha_medicion": _ahora_espana().strftime("%d/%m/%Y"),
            "imagen_exterior_path": None,
            "tecnico": "",
            "direccion": "",
            "tipo_centro": tipo_centro
        })
    
    return rowid


def fetch_centros():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, nombre, zona, fecha_medicion, imagen_exterior_path FROM centros ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows


def get_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT id, nombre, zona, fecha_medicion, imagen_exterior_path, tecnico, direccion FROM centros WHERE id=?", (centro_id,))
    row = c.fetchone()
    conn.close()
    return row


def get_tipo_centro(centro_id):
    """Función aparte (no incluida en get_centro, para no tener que
    tocar los muchos sitios que desempaquetan su tupla de 7 valores):
    devuelve solo el tipo de centro guardado (p.ej. "Consultorio"),
    usado para saber el prefijo correcto del código de sala."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT tipo_centro FROM centros WHERE id=?", (centro_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None


def update_centro(centro_id, nombre, zona, fecha, imagen_path, direccion=""):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE centros SET nombre=?, zona=?, fecha_medicion=?, imagen_exterior_path=?, direccion=? WHERE id=?",
              (nombre, zona, fecha, imagen_path, direccion, centro_id))
    conn.commit()
    conn.close()
    
    # Si la imagen es una URL de Supabase, no la modificamos
    if imagen_path and not es_url_supabase(imagen_path) and os.path.exists(imagen_path):
        # Intentar subir a Supabase
        try:
            conn_sup = get_supabase_connection()
            if conn_sup:
                nombre_img = os.path.basename(imagen_path)
                ruta_remota = f"exterior/{centro_id}_{nombre_img}"
                with open(imagen_path, 'rb') as f:
                    conn_sup.storage.from_("imagenes").upload(ruta_remota, f)
                imagen_path = conn_sup.storage.from_("imagenes").get_public_url(ruta_remota)
        except Exception:
            pass
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "centros", {
            "nombre": nombre,
            "zona": zona,
            "fecha_medicion": fecha,
            "imagen_exterior_path": imagen_path,
            "direccion": direccion
        }, id_valor=centro_id)


def set_tecnico_centro(centro_id, valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE centros SET tecnico=? WHERE id=?", (valor, centro_id))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "centros", {
            "tecnico": valor
        }, id_valor=centro_id)


def get_datos_informe_centro(centro_id):
    """Función aparte (mismo motivo que get_tipo_centro): datos que se
    piden en la pantalla del informe final y que se guardan para que
    también salgan en el Excel -superficie construida, superficie
    útil, nº de plantas, fecha y medio de comunicación a los
    trabajadores-, sin tocar la tupla de get_centro."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute(
        "SELECT superficie_construida, superficie_util, num_plantas, "
        "fecha_comunicacion_trab, medio_comunicacion FROM centros WHERE id=?",
        (centro_id,),
    )
    row = c.fetchone()
    conn.close()
    campos = ("superficie_construida", "superficie_util", "num_plantas",
              "fecha_comunicacion_trab", "medio_comunicacion")
    if not row:
        return {c: "" for c in campos}
    return dict(zip(campos, (v or "" for v in row)))


def set_datos_informe_centro(centro_id, **valores):
    """Guarda uno o varios de los campos de get_datos_informe_centro,
    p.ej. set_datos_informe_centro(cid, superficie_util="120")."""
    columnas_validas = {
        "superficie_construida", "superficie_util", "num_plantas",
        "fecha_comunicacion_trab", "medio_comunicacion",
    }
    valores = {k: v for k, v in valores.items() if k in columnas_validas}
    if not valores:
        return
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    asignaciones = ", ".join(f"{k}=?" for k in valores)
    c.execute(f"UPDATE centros SET {asignaciones} WHERE id=?", (*valores.values(), centro_id))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "centros", valores, id_valor=centro_id)


def delete_centro(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM detectores WHERE centro_id=?", (centro_id,))
    c.execute("DELETE FROM centros WHERE id=?", (centro_id,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("delete", "centros", None, id_valor=centro_id)
        # Nota: Los detectores se borran en cascada en Supabase si tienes foreign key con ON DELETE CASCADE


def insert_detector(data):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''INSERT INTO detectores
                 (centro_id, planta, sala, fecha, detector_codigo, plano_path,
                  punto_x, punto_y, foto_situacion_path, foto_detector_path, fecha_creacion,
                  codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel,
                  plano_centro_id, fecha_retirada_real, hora_retirada_real)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', data)
    rowid = c.lastrowid
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase (si está disponible)
    if SUPABASE_AVAILABLE:
        # Convertir data (tupla) a dict con nombres de columnas
        columnas = [
            "centro_id", "planta", "sala", "fecha", "detector_codigo", "plano_path",
            "punto_x", "punto_y", "foto_situacion_path", "foto_detector_path", "fecha_creacion",
            "codigo_sala", "profesionales_sala", "hora_colocacion", "turno_trabajo", "nivel",
            "plano_centro_id", "fecha_retirada_real", "hora_retirada_real"
        ]
        datos_dict = dict(zip(columnas, data))
        datos_dict["id"] = rowid
        
        # Procesar rutas de imágenes - si son locales, subir a Supabase
        conn_sup = get_supabase_connection()
        for campo in ["plano_path", "foto_situacion_path", "foto_detector_path"]:
            if campo in datos_dict and datos_dict[campo] and not es_url_supabase(datos_dict[campo]):
                ruta_local = datos_dict[campo]
                if os.path.exists(ruta_local):
                    try:
                        bucket = "planos" if campo == "plano_path" else "fotos"
                        carpeta = {
                            "plano_path": f"planos/{rowid}",
                            "foto_situacion_path": f"situacion/{rowid}",
                            "foto_detector_path": f"detectores/{rowid}"
                        }.get(campo, "generales")
                        nombre_img = os.path.basename(ruta_local)
                        ruta_remota = f"{carpeta}/{nombre_img}"
                        with open(ruta_local, 'rb') as f:
                            conn_sup.storage.from_(bucket).upload(ruta_remota, f)
                        datos_dict[campo] = conn_sup.storage.from_(bucket).get_public_url(ruta_remota)
                    except Exception:
                        pass  # Si falla, dejar la ruta local
        
        sync_to_supabase("insert", "detectores", datos_dict)
    
    return rowid


def update_detector(detector_id, data):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute('''UPDATE detectores SET
                 centro_id=?, planta=?, sala=?, fecha=?, detector_codigo=?, plano_path=?,
                 punto_x=?, punto_y=?, foto_situacion_path=?, foto_detector_path=?, fecha_creacion=?,
                 codigo_sala=?, profesionales_sala=?, hora_colocacion=?, turno_trabajo=?, nivel=?,
                 plano_centro_id=?, fecha_retirada_real=?, hora_retirada_real=?
                 WHERE id=?''', data + (detector_id,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        columnas = [
            "centro_id", "planta", "sala", "fecha", "detector_codigo", "plano_path",
            "punto_x", "punto_y", "foto_situacion_path", "foto_detector_path", "fecha_creacion",
            "codigo_sala", "profesionales_sala", "hora_colocacion", "turno_trabajo", "nivel",
            "plano_centro_id", "fecha_retirada_real", "hora_retirada_real"
        ]
        datos_dict = dict(zip(columnas, data))
        
        # Procesar rutas de imágenes - si son locales, subir a Supabase
        conn_sup = get_supabase_connection()
        for campo in ["plano_path", "foto_situacion_path", "foto_detector_path"]:
            if campo in datos_dict and datos_dict[campo] and not es_url_supabase(datos_dict[campo]):
                ruta_local = datos_dict[campo]
                if os.path.exists(ruta_local):
                    try:
                        bucket = "planos" if campo == "plano_path" else "fotos"
                        carpeta = {
                            "plano_path": f"planos/{detector_id}",
                            "foto_situacion_path": f"situacion/{detector_id}",
                            "foto_detector_path": f"detectores/{detector_id}"
                        }.get(campo, "generales")
                        nombre_img = os.path.basename(ruta_local)
                        ruta_remota = f"{carpeta}/{nombre_img}"
                        with open(ruta_local, 'rb') as f:
                            conn_sup.storage.from_(bucket).upload(ruta_remota, f)
                        datos_dict[campo] = conn_sup.storage.from_(bucket).get_public_url(ruta_remota)
                    except Exception:
                        pass  # Si falla, dejar la ruta local
        
        sync_to_supabase("update", "detectores", datos_dict, id_valor=detector_id)


def actualizar_retirada_detector(detector_id, fecha_retirada_real, hora_retirada_real):
    """Guarda solo la fecha/hora real de retirada de un detector, sin
    necesidad de tocar el resto de sus datos (se usa desde el bloque
    "Retirada de detectores")."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE detectores SET fecha_retirada_real=?, hora_retirada_real=? WHERE id=?",
              (fecha_retirada_real, hora_retirada_real, detector_id))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "detectores", {
            "fecha_retirada_real": fecha_retirada_real,
            "hora_retirada_real": hora_retirada_real
        }, id_valor=detector_id)


def actualizar_resultado_detector(detector_id, resultado, incertidumbre):
    """Guarda el resultado de la medición (Bq/m³) y la incertidumbre de
    un detector, tal como se hayan introducido en la pantalla del
    informe final, para que también salgan en el registro para
    laboratorio y en el Excel (antes solo se usaban para ese informe
    en concreto y se perdían al salir)."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE detectores SET resultado_bq_m3=?, incertidumbre=? WHERE id=?",
              (resultado, incertidumbre, detector_id))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "detectores", {
            "resultado_bq_m3": resultado,
            "incertidumbre": incertidumbre
        }, id_valor=detector_id)


def fetch_detectores(centro_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT * FROM detectores WHERE centro_id=? ORDER BY id ASC", (centro_id,))
    rows = c.fetchall()
    conn.close()
    return rows


def get_detector(detector_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT * FROM detectores WHERE id=?", (detector_id,))
    row = c.fetchone()
    conn.close()
    return row


def _generar_codigo_sala(cid, detector_id, nivel, zona, tipo_centro=None):
    """Genera el código de la sala automáticamente, con el formato
    PREFIJO/NIVEL/CORRELATIVO (p.ej. "CS/S-1/01"):
      - PREFIJO: "CS" si el área del centro es de atención primaria,
        PAC o atención primaria + PAC; "CO" si el TIPO de centro es
        "Consultorio" (aunque su área/zona diga "Atención Primaria");
        o "HO" en el resto de casos (hospitales, centros de
        especialidades, etc.). Se mira primero el tipo de centro
        guardado (más fiable) y, si no está disponible, se recurre al
        texto del área como respaldo.
      - NIVEL: según la opción elegida en "Nivel":
          "3 niveles bajo rasante (Sótano -3)"  -> "S-3"
          "2 niveles bajo rasante (Sótano -2)"  -> "S-2"
          "1 nivel bajo rasante (Sótano -1)"    -> "S-1"
          "Nivel de la rasante (Planta Baja)"   -> "PB"
          "1 nivel sobre rasante"               -> "01"
          "2 niveles sobre rasante"             -> "02"
          "3 niveles sobre rasante"             -> "03"
      - CORRELATIVO: orden (01, 02...) de este detector entre todos
        los del centro que compartan el mismo nivel, por orden de
        creación (id). Es estable mientras no se borren detectores
        anteriores del mismo nivel.
    """
    if tipo_centro == "Consultorio":
        prefijo = "CO"
    else:
        zona_normalizada = (zona or "").strip().lower()
        es_atencion_primaria = any(
            palabra in zona_normalizada
            for palabra in ("atención primaria", "atencion primaria", "pac")
        )
        if es_atencion_primaria:
            prefijo = "CS"
        elif "consultorio" in zona_normalizada:
            # Respaldo para centros creados antes de este cambio, cuyo
            # Área/Zona todavía dice literalmente "Consultorio" (ahora
            # los nuevos guardan "Atención Primaria" ahí, y se
            # distinguen por tipo_centro en su lugar).
            prefijo = "CO"
        else:
            prefijo = "HO"

    nivel_code = NIVEL_A_CODIGO.get(nivel, "PB")

    hermanos = [d for d in fetch_detectores(cid) if d[16] == nivel]  # d[16] = nivel
    ids_hermanos = [d[0] for d in hermanos]  # fetch_detectores ya viene ordenado por id ASC
    if detector_id and detector_id in ids_hermanos:
        correlativo = ids_hermanos.index(detector_id) + 1
    else:
        correlativo = len(ids_hermanos) + 1

    return f"{prefijo}/{nivel_code}/{correlativo:02d}"


def delete_detector(detector_id):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("DELETE FROM detectores WHERE id=?", (detector_id,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("delete", "detectores", None, id_valor=detector_id)


def get_tecnico():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT tecnico FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_tecnico(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET tecnico=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "settings", {"tecnico": valor}, id_valor=1)


def get_empresa():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT empresa FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_empresa(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET empresa=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "settings", {"empresa": valor}, id_valor=1)


def get_cif():
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT cif FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""


def set_cif(valor):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET cif=? WHERE id=1", (valor,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "settings", {"cif": valor}, id_valor=1)


def get_logo_laboratorio():
    """Ruta al logotipo del laboratorio para la ficha de registro. Si
    no se ha subido uno propio, usa el que trae la app por defecto."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT logo_laboratorio_path FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    ruta = row[0] if row else None
    # Verificar si es URL de Supabase o ruta local
    if ruta:
        if es_url_supabase(ruta):
            return ruta
        elif os.path.exists(ruta):
            return ruta
    ruta_default = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "logo_laboratorio_default.png"
    )
    return ruta_default if os.path.exists(ruta_default) else None


def set_logo_laboratorio(ruta):
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("UPDATE settings SET logo_laboratorio_path=? WHERE id=1", (ruta,))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        sync_to_supabase("update", "settings", {"logo_laboratorio_path": ruta}, id_valor=1)


def get_logo_informe():
    """Logotipo personalizado (bytes, nombre) para el informe final
    (.docx) y el Informe PDF de colocación de detectores -distinto del
    logo de la ficha de registro para laboratorio-. Si no se ha subido
    uno propio, se usa el de UPRL/SERGAS que trae la app por defecto."""
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    c.execute("SELECT logo_path, logo_nombre FROM settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    ruta, nombre = (row[0], row[1]) if row else (None, None)
    if ruta and os.path.exists(ruta):
        with open(ruta, "rb") as f:
            return f.read(), nombre or os.path.basename(ruta)
    from utils_informe.assets import logo_por_defecto
    logo_bytes = logo_por_defecto()
    return (logo_bytes, "logo_uprl.png") if logo_bytes else (None, None)


def set_logo_informe(nombre_archivo, contenido_bytes):
    """Guarda un logotipo personalizado (o lo quita, si contenido_bytes
    es None) para el informe final y el Informe PDF de colocación."""
    data_dir = get_data_dir()
    conn = sqlite3.connect(get_db_path())
    c = conn.cursor()
    if contenido_bytes is None:
        c.execute("UPDATE settings SET logo_path='', logo_nombre='' WHERE id=1")
    else:
        ext = os.path.splitext(nombre_archivo)[1] or ".png"
        ruta = os.path.join(data_dir, f"logo_informe_personalizado{ext}")
        with open(ruta, "wb") as f:
            f.write(contenido_bytes)
        c.execute("UPDATE settings SET logo_path=?, logo_nombre=? WHERE id=1", (ruta, nombre_archivo))
    conn.commit()
    conn.close()
    
    # Sincronizar con Supabase
    if SUPABASE_AVAILABLE:
        if contenido_bytes is None:
            sync_to_supabase("update", "settings", {"logo_path": "", "logo_nombre": ""}, id_valor=1)
        else:
            sync_to_supabase("update", "settings", {"logo_path": ruta, "logo_nombre": nombre_archivo}, id_valor=1)


def _mostrar_previsualizacion(nombre_archivo, contenido_bytes):
    """Vista previa en pequeño de un anexo: miniatura real tanto si
    es una imagen como si es un PDF (se renderiza la primera
    página con PyMuPDF, que no depende de ningún programa externo
    del sistema como poppler, así que funciona igual en Streamlit
    Cloud que en Termux)."""
    extension = nombre_archivo.rsplit(".", 1)[-1].lower() if "." in nombre_archivo else ""
    if extension in ("png", "jpg", "jpeg"):
        st.image(contenido_bytes, width=160)
    elif extension == "pdf":
        try:
            import pymupdf
            doc_pdf = pymupdf.open(stream=contenido_bytes, filetype="pdf")
            num_paginas = doc_pdf.page_count
            pagina = doc_pdf.load_page(0)
            miniatura = pagina.get_pixmap(matrix=pymupdf.Matrix(0.3, 0.3))
            st.image(miniatura.tobytes("png"), width=160)
            st.caption(f"📄 {nombre_archivo} ({num_paginas} página{'s' if num_paginas != 1 else ''})")
            doc_pdf.close()
        except Exception:
            try:
                from pypdf import PdfReader
                num_paginas = len(PdfReader(io.BytesIO(contenido_bytes)).pages)
                st.caption(f"📄 {nombre_archivo} ({num_paginas} página{'s' if num_paginas != 1 else ''})")
            except Exception:
                st.caption(f"📄 {nombre_archivo}")
    else:
        st.caption(f"📎 {nombre_archivo}")

def _widget_archivo_con_eliminar(clave, etiqueta_subida, tipos, valor_por_defecto=None):
    """Campo para subir un archivo (anexo o logotipo) que, mientras
    haya algo cargado —recién subido, guardado de una vez anterior,
    o el que trae la app por defecto—, no vuelve a mostrar el
    campo de subida: solo se ve la miniatura y un botón para
    eliminarlo. El campo de subida solo reaparece si se elimina lo
    que hubiera (y, tras eliminar un valor por defecto, ya no se
    recupera solo: hay que subir uno nuevo a mano). Devuelve
    (nombre, bytes) o None."""
    guardado_key = f"{clave}_guardado"
    eliminado_key = f"{clave}_eliminado"
    reset_key = f"{clave}_reset"

    actual = st.session_state.get(guardado_key)
    if actual is None and valor_por_defecto and not st.session_state.get(eliminado_key):
        actual = valor_por_defecto

    if actual:
        _mostrar_previsualizacion(*actual)
        st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
        if st.button("❌ Eliminar", key=f"{clave}_eliminar"):
            st.session_state.pop(guardado_key, None)
            st.session_state[eliminado_key] = True
            st.session_state[reset_key] = st.session_state.get(reset_key, 0) + 1
            st.rerun()
        return actual

    uploader_key = f"{clave}_{st.session_state.get(reset_key, 0)}"
    archivo = st.file_uploader(etiqueta_subida, type=tipos, key=uploader_key)
    if archivo is not None:
        nuevo = (archivo.name, archivo.getvalue())
        st.session_state[guardado_key] = nuevo
        st.session_state[eliminado_key] = False
        st.rerun()
    return None


def _acordeon_informe(numero, titulo, completo):
    """Apartado desplegable de la pantalla del informe final: rosa si
    le falta algo, gris si ya está completo (mismo criterio de color
    que el resto de campos de la app). "numero" puede ir vacío (para
    el último apartado, "Generación del informe", que no lleva
    número)."""
    clase = "marcador-acordeon-gris" if completo else "marcador-acordeon-rosa"
    st.markdown(f'<div class="{clase}"></div>', unsafe_allow_html=True)
    etiqueta = f"{numero}. {titulo}" if numero else titulo
    return st.expander(etiqueta, expanded=False)


# ============================================================
# UTILIDADES
# ============================================================

def _slug(texto):
    texto = (texto or "centro").lower()
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto or "centro"


def _limpiar_para_nombre_archivo(texto):
    """Deja el texto tal cual (con acentos, mayúsculas y espacios) para
    que el nombre del archivo siga siendo legible, sustituyendo solo
    los caracteres que un nombre de archivo no puede llevar (sobre
    todo "/" y "\\", que crearían sin querer una "subcarpeta")."""
    texto = (texto or "").strip()
    texto = re.sub(r'[\\/:*?"<>|]+', "-", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _nombre_foto_situacion(codigo_detector, centro, zona):
    partes = [p for p in (
        _limpiar_para_nombre_archivo(codigo_detector) or "detector",
        "situación",
        _limpiar_para_nombre_archivo(centro),
        _limpiar_para_nombre_archivo(zona),
    ) if p]
    return "-".join(partes)


def _nombre_foto_detector(codigo_detector, centro, zona):
    partes = [p for p in (
        _limpiar_para_nombre_archivo(codigo_detector) or "detector",
        _limpiar_para_nombre_archivo(centro),
        _limpiar_para_nombre_archivo(zona),
    ) if p]
    return "-".join(partes)


def _nombre_foto_plano(codigo_detector, centro, zona):
    partes = [p for p in (
        _limpiar_para_nombre_archivo(codigo_detector) or "detector",
        "PLANO",
        _limpiar_para_nombre_archivo(centro),
        _limpiar_para_nombre_archivo(zona),
    ) if p]
    return "-".join(partes)


def _nombre_documento(centro, tipo_documento, sufijo_extra=""):
    """Nombre de archivo para los documentos generados (Informe PDF,
    Excel, Registro para laboratorio, Informe final...), con el
    formato "CENTRO-TIPO-fecha-hora" (p.ej.
    "C.S. Carballo-INFORME-COLOCACIÓN-28082026-114500")."""
    centro_limpio = _limpiar_para_nombre_archivo(centro) or "Centro"
    marca = _ahora_espana().strftime("%d%m%Y-%H%M%S")
    nombre = f"{centro_limpio}-{tipo_documento}-{marca}"
    return nombre + sufijo_extra if sufijo_extra else nombre


def generar_plano_con_punto(plano_path, px, py, destino_path):
    """Copia la imagen de un plano dibujando encima, en rojo, el punto
    de un detector concreto (posición relativa 0-1). Se usa para poder
    descargar/enviar el plano YA con el punto marcado, en vez de la
    imagen del plano "en blanco". Devuelve True si se generó bien."""
    try:
        # Si es URL de Supabase, descargar primero
        if es_url_supabase(plano_path):
            if not descargar_desde_supabase(plano_path, destino_path):
                return False
            img_path = destino_path
        else:
            img_path = plano_path
        
        with Image.open(img_path) as im:
            im = ImageOps.exif_transpose(im)
            im = im.convert("RGB")
            if px is not None and py is not None and px >= 0 and py >= 0:
                draw = ImageDraw.Draw(im)
                w, h = im.size
                cx, cy = px * w, py * h
                r = max(6, int(min(w, h) * 0.015))
                draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                             fill=(220, 20, 20), outline=(120, 0, 0), width=2)
            im.save(destino_path, quality=90)
        return True
    except Exception:
        return False


def guardar_bytes_imagen(file_bytes, prefijo, ext=".jpg"):
    """Guarda bytes de una imagen y la sube a Supabase Storage.
    
    Devuelve la URL pública si se subió a Supabase, o la ruta local si falló.
    """
    import time
    
    data_dir = get_data_dir()
    timestamp = _ahora_espana().strftime('%Y%m%d_%H%M%S_%f')
    nombre = f"{prefijo}_{timestamp}{ext}"
    ruta_local = os.path.join(data_dir, nombre)
    
    # Guardar localmente primero (siempre)
    with open(ruta_local, "wb") as f:
        f.write(file_bytes)
    
    # Intentar subir a Supabase Storage
    if SUPABASE_AVAILABLE:
        try:
            conn = get_supabase_connection()
            if conn:
                # Elegir bucket según el prefijo
                if "plano" in prefijo.lower():
                    bucket = "planos"
                    carpeta = "planos"
                elif "foto_situacion" in prefijo.lower():
                    bucket = "fotos"
                    carpeta = "situacion"
                elif "foto_detector" in prefijo.lower():
                    bucket = "fotos"
                    carpeta = "detectores"
                elif "logo" in prefijo.lower():
                    bucket = "imagenes"
                    carpeta = "logos"
                else:
                    bucket = "imagenes"
                    carpeta = "generales"
                
                ruta_remota = f"{carpeta}/{nombre}"
                
                # Subir archivo a Supabase
                with open(ruta_local, "rb") as f:
                    conn.storage.from_(bucket).upload(ruta_remota, f)
                
                # Obtener URL pública
                url_publica = conn.storage.from_(bucket).get_public_url(ruta_remota)
                
                # Eliminar archivo local para no ocupar espacio
                try:
                    os.remove(ruta_local)
                except:
                    pass
                
                return url_publica
                
        except Exception as e:
            # Mostrar error detallado
            st.error(f"❌ Error subiendo a Supabase: {e}")
            st.info("💡 Verifica que los buckets existan y sean públicos")
    
    # Fallback: devolver ruta local
    return ruta_local


def extension_de(uploaded_file, por_defecto=".jpg"):
    if uploaded_file is None:
        return por_defecto
    nombre = getattr(uploaded_file, "name", "") or ""
    ext = os.path.splitext(nombre)[1]
    return ext if ext else por_defecto


# ============================================================
# LOGOTIPO (para cabecera del informe PDF) - idéntico al original
# ============================================================

LOGO_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAABkAAAADeCAMAAABSbjycAAAAwFBMVEX9/v4Ce8QChMrK1+q0yOMvhsl5qtVrmc/a5PBSlM6Wtdr+/v51pNRJi8mku9ytwt7AzeW80eckfsU7kM6lvuCpvd6avOCaweFeoNN9sduewN7///+lvuFgj8qux+QDe7+AntGww9/V3u3Z"
    "5PAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADdAvdLAAAAMHRSTlP+///+///+/v7///+9//////7/////vf//////i73/vf//vb29AAAAAAAAAAAAAABMCEqxAABi"
    "NklEQVR42u2diZriOrKgs9AutbxgmKzTfWef93/GUYQkW7JlY0ggyURxv9snC4wRYMev2D/+61+VKlWqVKlytfzXx7/+qVKlSpUqVa6Wf1WAVKlSpUqVCpAqVapUqVIBUqVKlSpVKkCqVKlSpUoFSJUqVapUqQ=="
    "A=="
)

# ============================================================
# GENERAR PDF (idéntico al original, sin dependencias de Tkinter)
# ============================================================

def _escapar_pdf(texto):
    """Escapa caracteres especiales de XML (&, <, >) antes de meter un
    texto en un Paragraph de ReportLab. Sin esto, cualquier campo que
    el usuario escriba con esos símbolos (p. ej. "Sala < 20m2" o "A&B")
    rompe el generador de PDF con un error de "unclosed tags"."""
    if texto is None:
        return ""
    texto = str(texto)
    return texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def generar_pdf(centro_id, output_path):
    try:
        import base64
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.platypus import Image as RLImage
        from reportlab.lib.utils import ImageReader
        from PIL import Image as PILImage, ImageDraw
        
        centro = get_centro(centro_id)
        if not centro:
            raise ValueError("Centro no encontrado")
        
        _, nombre, zona, fecha, img_ext, tecnico, direccion = centro
        detectores = fetch_detectores(centro_id)
        
        # ---- Logotipo (para cabecera de todas las paginas): el que se
        # haya configurado en Datos de la empresa, o si no el de
        # UPRL/SERGAS que trae la app por defecto ----
        logo_bytes, _ = get_logo_informe()
        if not logo_bytes:
            logo_bytes = base64.b64decode(LOGO_PNG_B64)
        with PILImage.open(io.BytesIO(logo_bytes)) as _logo_im:
            logo_w_px, logo_h_px = _logo_im.size
        logo_aspect = logo_w_px / logo_h_px
        
        def _dibujar_cabecera(canvas_obj, doc_):
            canvas_obj.saveState()
            header_h = 1.3 * cm
            header_w = header_h * logo_aspect
            page_w, page_h = doc_.pagesize
            x = (page_w - header_w) / 2
            y = page_h - 0.5 * cm - header_h
            img_reader = ImageReader(io.BytesIO(logo_bytes))
            canvas_obj.drawImage(img_reader, x, y, width=header_w, height=header_h,
                                  mask='auto', preserveAspectRatio=True)
            canvas_obj.restoreState()
        
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2.1*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        centrado = ParagraphStyle('Centrado', parent=styles['Normal'], alignment=TA_CENTER)
        nombre_style = ParagraphStyle('NombreCentro', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=20, leading=24,
                                       alignment=TA_CENTER, spaceAfter=4)
        zona_style = ParagraphStyle('ZonaCentro', parent=styles['Normal'],
                                     fontName='Helvetica', fontSize=20, leading=24,
                                     alignment=TA_CENTER, textColor=colors.HexColor('#444444'),
                                     spaceAfter=10)
        story = []
        
        story.append(Paragraph("Informe de colocación de detectores de Rn", styles["Title"]))
        story.append(Spacer(1, 0.6*cm))
        story.append(Paragraph(_escapar_pdf(nombre) or '-', nombre_style))
        if zona:
            story.append(Paragraph(_escapar_pdf(zona), zona_style))
        story.append(Spacer(1, 0.3*cm))
        
        if img_ext and os.path.exists(img_ext):
            try:
                with PILImage.open(img_ext) as im_ext:
                    w, h = im_ext.size
                r = min(14*cm/w, 10.5*cm/h)
                img_portada = RLImage(img_ext, width=w*r, height=h*r)
                img_portada.hAlign = 'CENTER'
                story.append(img_portada)
                story.append(Spacer(1, 0.5*cm))
            except Exception:
                pass
        
        empresa_pdf = get_empresa()
        cif_pdf = get_cif()
        if empresa_pdf:
            story.append(Paragraph(f"<b>Empresa:</b> {_escapar_pdf(empresa_pdf)}", centrado))
        if cif_pdf:
            story.append(Paragraph(f"<b>CIF:</b> {_escapar_pdf(cif_pdf)}", centrado))
        if tecnico:
            story.append(Paragraph(f"<b>Técnico:</b> {_escapar_pdf(tecnico)}", centrado))
        story.append(Paragraph(f"<b>Fecha:</b> {_escapar_pdf(fecha) or '-'}", centrado))
        story.append(Paragraph(f"<b>Detectores:</b> {len(detectores)}", centrado))
        story.append(Spacer(1, 0.5*cm))

        categorias_centro_pdf = fetch_categorias_centro(centro_id)
        if categorias_centro_pdf:
            story.append(Paragraph("<b>Categorías profesionales</b>", centrado))
            story.append(Spacer(1, 0.2*cm))
            filas_cat = [["Categoría profesional", "Nº personas expuestas"]]
            for _, _, categoria_pdf, num_personas_pdf, _ in categorias_centro_pdf:
                filas_cat.append([_escapar_pdf(categoria_pdf), str(num_personas_pdf)])
            tabla_cat = Table(filas_cat, colWidths=[9*cm, 6*cm])
            tabla_cat.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F5A623")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            story.append(tabla_cat)
            story.append(Spacer(1, 0.5*cm))
        
        for idx, d in enumerate(detectores, 1):
            (did, _, planta, sala, fecha_det, codigo, _plano_antiguo, px, py, foto_sit, foto_det, _,
             codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real,
             resultado_bq_m3, incertidumbre) = d
            plano = None
            nombre_plano_actual = None
            if plano_centro_id:
                plano_info = get_plano_centro(plano_centro_id)
                if plano_info:
                    nombre_plano_actual = plano_info[2]
                    plano = plano_info[3]
            story.append(PageBreak())
            
            titulo_partes = [codigo or "-", nombre or "-"]
            if zona:
                titulo_partes.append(zona)
            titulo_detector = f"Detector {idx}: " + " - ".join(titulo_partes)
            titulo_detector = _escapar_pdf(titulo_detector)
            estilo_titulo_detector = ParagraphStyle(
                'TituloDetector', parent=styles['Heading2'], fontSize=13,
                spaceBefore=0, spaceAfter=6,
            )
            story.append(Paragraph(titulo_detector, estilo_titulo_detector))
            
            fecha_retirada_optima_pdf = calcular_fecha_retirada_optima(fecha_det)

            # Las celdas se envuelven en Paragraph para que el texto largo
            # (p.ej. varios profesionales, o "Mañana + tarde + noche") se
            # reparta en varias líneas en vez de desbordar la celda; la
            # fila crece de alto automáticamente según haga falta.
            estilo_label_celda = ParagraphStyle(
                'CeldaLabel', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica-Bold',
            )
            estilo_valor_celda = ParagraphStyle(
                'CeldaValor', parent=styles['Normal'], fontSize=9, leading=11,
            )

            fecha_y_hora_colocacion = _escapar_pdf(fecha_det) or "-"
            if hora_colocacion:
                fecha_y_hora_colocacion += f"<br/>{_escapar_pdf(hora_colocacion)}"

            fecha_y_hora_retirada_real = ""
            if fecha_retirada_real:
                fecha_y_hora_retirada_real = _escapar_pdf(fecha_retirada_real)
                if hora_retirada_real:
                    fecha_y_hora_retirada_real += f"<br/>{_escapar_pdf(hora_retirada_real)}"

            filas_texto = [
                ["Codigo del detector", _escapar_pdf(codigo) or "-", "Fecha y hora de colocacion", fecha_y_hora_colocacion],
                ["Planta", _escapar_pdf(planta) or "-", "Nivel", _escapar_pdf(nivel) or "-"],
                ["Sala", _escapar_pdf(sala) or "-", "Codigo de la sala", _escapar_pdf(codigo_sala) or "-"],
                ["Profesionales en la sala", _escapar_pdf(profesionales_sala) or "-", "Turno de trabajo", _escapar_pdf(turno_trabajo) or "-"],
                ["Fecha optima retirada", _escapar_pdf(fecha_retirada_optima_pdf) or "-", "Fecha y hora real de retirada", fecha_y_hora_retirada_real],
                ["Resultado (Bq/m3)", _escapar_pdf(str(resultado_bq_m3)) if resultado_bq_m3 is not None else "-",
                 "Incertidumbre", _escapar_pdf(incertidumbre) or "-"],
            ]
            filas_pdf = []
            for fila_txt in filas_texto:
                fila_pdf = []
                for i, val in enumerate(fila_txt):
                    if not val:
                        fila_pdf.append("")
                    else:
                        estilo = estilo_label_celda if i % 2 == 0 else estilo_valor_celda
                        fila_pdf.append(Paragraph(val, estilo))
                filas_pdf.append(fila_pdf)

            tabla = Table(filas_pdf, colWidths=[3.3*cm, 4.7*cm, 3.3*cm, 4.7*cm])
            tabla.setStyle(TableStyle([
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
                ("BACKGROUND", (2,0), (2,-1), colors.whitesmoke),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
            ]))
            story.append(tabla)
            story.append(Spacer(1, 0.2*cm))

            estilo_subtitulo = ParagraphStyle(
                'SubtituloCompacto', parent=styles['Normal'], fontSize=10,
                fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2,
            )

            # Obtener el plano (puede ser URL o ruta local)
            plano_real = None
            if plano and os.path.exists(plano):
                plano_real = plano
            elif plano and es_url_supabase(plano):
                # Descargar plano temporalmente
                temp_plano = os.path.join(get_data_dir(), f"_tmp_plano_{did}.jpg")
                if descargar_desde_supabase(plano, temp_plano):
                    plano_real = temp_plano
            
            if plano_real and os.path.exists(plano_real):
                nombre_plano_esc = _escapar_pdf(nombre_plano_actual)
                titulo_ubicacion = f"Ubicacion en el plano ({nombre_plano_esc}):" if nombre_plano_esc else "Ubicacion en el plano:"
                story.append(Paragraph(titulo_ubicacion, estilo_subtitulo))
                try:
                    with PILImage.open(plano_real) as im_plano:
                        im_plano = im_plano.convert("RGB")
                        w, h = im_plano.size
                        if px is not None and py is not None and px >= 0 and py >= 0:
                            draw = ImageDraw.Draw(im_plano)
                            cx, cy = px * w, py * h
                            r = max(6, int(min(w, h) * 0.012))
                            draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                                         fill=(220, 20, 20), outline=(120, 0, 0), width=2)
                        tmp_plano_path = os.path.join(get_data_dir(), f"_tmp_plano_{did}.jpg")
                        im_plano.save(tmp_plano_path, quality=90)
                    r = min(16.5*cm/w, 8.5*cm/h)
                    story.append(RLImage(tmp_plano_path, width=w*r, height=h*r))
                    story.append(Spacer(1, 0.2*cm))
                except Exception:
                    pass
            
            # Fotos: pueden ser URLs de Supabase o rutas locales
            foto_sit_real = None
            foto_det_real = None
            
            if foto_sit:
                if os.path.exists(foto_sit):
                    foto_sit_real = foto_sit
                elif es_url_supabase(foto_sit):
                    temp_sit = os.path.join(get_data_dir(), f"_tmp_sit_{did}.jpg")
                    if descargar_desde_supabase(foto_sit, temp_sit):
                        foto_sit_real = temp_sit
            
            if foto_det:
                if os.path.exists(foto_det):
                    foto_det_real = foto_det
                elif es_url_supabase(foto_det):
                    temp_det = os.path.join(get_data_dir(), f"_tmp_det_{did}.jpg")
                    if descargar_desde_supabase(foto_det, temp_det):
                        foto_det_real = temp_det
            
            if foto_sit_real and os.path.exists(foto_sit_real) and foto_det_real and os.path.exists(foto_det_real):
                story.append(Paragraph("Fotos:", estilo_subtitulo))
                titulo_foto_style = ParagraphStyle(
                    'TituloFoto', parent=styles['Normal'], alignment=TA_CENTER, fontName="Helvetica-Bold",
                )
                try:
                    with PILImage.open(foto_sit_real) as im:
                        w, h = im.size
                    r = min(8.2*cm/w, 7.5*cm/h)
                    img1 = RLImage(foto_sit_real, width=w*r, height=h*r)
                except:
                    img1 = Paragraph("(no disponible)", styles["Normal"])
                try:
                    with PILImage.open(foto_det_real) as im:
                        w, h = im.size
                    r = min(8.2*cm/w, 7.5*cm/h)
                    img2 = RLImage(foto_det_real, width=w*r, height=h*r)
                except:
                    img2 = Paragraph("(no disponible)", styles["Normal"])
                cap1 = Paragraph("Situación del detector", titulo_foto_style)
                cap2 = Paragraph("Detector", titulo_foto_style)
                story.append(Table([[cap1, cap2], [img1, img2]], colWidths=[8.5*cm, 8.5*cm]))
        
        doc.build(story, onFirstPage=_dibujar_cabecera, onLaterPages=_dibujar_cabecera)
        return True
    except Exception as e:
        raise Exception(f"Error: {str(e)}")


# ============================================================
# GENERAR EXCEL (hoja de cálculo del centro, con las fotos de cada
# detector incrustadas cada una en su propia celda)
# ============================================================

# ============================================================
# CÁLCULO DE LA FECHA DE RETIRADA ÓPTIMA (90 días laborables desde la
# colocación, saltando fines de semana y festivos de Galicia)
# ============================================================

def _domingo_de_pascua(anio):
    """Algoritmo de Gauss/Meeus para calcular el Domingo de Pascua de
    un año dado (necesario para Jueves y Viernes Santo, festivos
    movibles que dependen de la Pascua cada año)."""
    a = anio % 19
    b = anio // 100
    c = anio % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(anio, mes, dia)


def _festivos_galicia(anio):
    """Festivos de ámbito estatal/autonómico aplicables en Galicia para
    un año dado (no incluye fiestas locales de cada ayuntamiento, que
    varían por municipio)."""
    pascua = _domingo_de_pascua(anio)
    jueves_santo = pascua - timedelta(days=3)
    viernes_santo = pascua - timedelta(days=2)
    fijos = [
        date(anio, 1, 1),    # Año Nuevo
        date(anio, 1, 6),    # Reyes
        date(anio, 5, 1),    # Fiesta del Trabajo
        date(anio, 7, 25),   # Santiago Apóstol (patrón de Galicia)
        date(anio, 8, 15),   # Asunción
        date(anio, 10, 12),  # Fiesta Nacional de España
        date(anio, 11, 1),   # Todos los Santos
        date(anio, 12, 6),   # Día de la Constitución
        date(anio, 12, 8),   # Inmaculada Concepción
        date(anio, 12, 25),  # Navidad
    ]
    return set(fijos + [jueves_santo, viernes_santo])


def _parsear_fecha(texto):
    """Intenta interpretar una fecha escrita a mano en varios formatos
    habituales. Devuelve un date, o None si no se pudo interpretar."""
    texto = (texto or "").strip()
    if not texto:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def calcular_fecha_retirada_optima(fecha_colocacion_texto):
    """90 días naturales desde la fecha de colocación; si ese día cae
    en sábado, domingo o festivo de Galicia, se pasa al siguiente día
    laborable. Devuelve el resultado como texto "DD/MM/AAAA", o ""
    si la fecha de colocación no se pudo interpretar."""
    fecha_col = _parsear_fecha(fecha_colocacion_texto)
    if fecha_col is None:
        return ""
    objetivo = fecha_col + timedelta(days=90)
    cache_festivos = {}
    while True:
        if objetivo.year not in cache_festivos:
            cache_festivos[objetivo.year] = _festivos_galicia(objetivo.year)
        if objetivo.weekday() >= 5 or objetivo in cache_festivos[objetivo.year]:
            objetivo += timedelta(days=1)
            continue
        break
    return objetivo.strftime("%d/%m/%Y")


def importar_centro_desde_excel(archivo_bytes):
    """Reconstruye un centro completo (datos, planos y detectores, con
    el punto exacto de cada uno sobre su plano) a partir de un Excel
    generado por generar_excel.

    Devuelve (centro_id, numero_de_detectores_creados). Lanza
    ValueError con un mensaje claro si el archivo no tiene la
    estructura esperada (no es un Excel exportado por esta app).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(archivo_bytes))
    if "Detectores" not in wb.sheetnames:
        raise ValueError(
            "El archivo no parece un Excel exportado por esta app "
            "(falta la hoja 'Detectores')."
        )
    ws = wb["Detectores"]

    nombre_centro = ws.cell(row=1, column=1).value or "Centro importado"
    area_centro = ws.cell(row=1, column=5).value or ""
    tecnico_centro = ws.cell(row=1, column=10).value or ""
    fecha_centro = ws.cell(row=1, column=15).value or ""
    direccion_centro = ws.cell(row=1, column=17).value or ""

    header_row = 2
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[val] = col

    columnas_esperadas = ["Centro", "Área / Zona", "Planta", "Sala", "Código"]
    faltan = [c for c in columnas_esperadas if c not in headers]
    if faltan:
        raise ValueError(
            "El archivo no tiene el formato esperado (faltan columnas: "
            + ", ".join(faltan) + ")."
        )

    # Imágenes incrustadas de la hoja "Detectores", indexadas por la
    # posición exacta de celda (fila0, col0) en la que se anclaron al
    # generarlas, para poder recuperar la foto de cada detector.
    imagenes_por_celda = {}
    for img in ws._images:
        try:
            fr = img.anchor._from
            imagenes_por_celda[(fr.row, fr.col)] = img._data()
        except Exception:
            continue

    centro_id = crear_centro(str(nombre_centro).strip() or "Centro importado")
    imagen_exterior_path = None

    # --- Hoja "Planos": crea TODOS los planos del centro (incluidos los
    # que ningún detector tenga asignado) y construye el mapa
    # nombre -> id para poder vincular cada detector a su plano.
    # También trae la Empresa y el CIF (guardados en esta misma hoja,
    # en sus dos primeras filas). ---
    planos_por_nombre = {}
    if "Planos" in wb.sheetnames:
        ws_p = wb["Planos"]

        empresa_importada = ws_p.cell(row=1, column=2).value
        cif_importado = ws_p.cell(row=2, column=2).value
        if empresa_importada:
            set_empresa(str(empresa_importada).strip())
        if cif_importado:
            set_cif(str(cif_importado).strip())

        imagenes_planos_por_celda = {}
        for img in ws_p._images:
            try:
                fr = img.anchor._from
                imagenes_planos_por_celda[(fr.row, fr.col)] = img._data()
            except Exception:
                continue
        orden = 0
        # Las filas 1-4 son Empresa, CIF, una fila en blanco y la
        # cabecera "Nombre"/"Imagen"; los planos empiezan en la 5.
        for fila_p in range(5, ws_p.max_row + 1):
            nombre_p = ws_p.cell(row=fila_p, column=1).value
            if not nombre_p:
                continue
            datos_img = imagenes_planos_por_celda.get((fila_p - 1, 1))
            if not datos_img:
                continue
            ruta_guardada = guardar_bytes_imagen(datos_img, f"plano_centro_{centro_id}")
            if str(nombre_p) == "(Foto exterior del centro)":
                imagen_exterior_path = ruta_guardada
            else:
                nuevo_id = insert_plano_centro(centro_id, str(nombre_p), ruta_guardada, orden)
                planos_por_nombre[str(nombre_p)] = nuevo_id
                orden += 1

    update_centro(
        centro_id, str(nombre_centro).strip() or "Centro importado",
        str(area_centro).strip(), str(fecha_centro).strip(), imagen_exterior_path,
        str(direccion_centro).strip(),
    )
    if tecnico_centro:
        set_tecnico_centro(centro_id, str(tecnico_centro).strip())

    # --- Hoja "Categorías profesionales" (si existe) ---
    if "Categorías profesionales" in wb.sheetnames:
        ws_cat = wb["Categorías profesionales"]
        for fila_cat in range(2, ws_cat.max_row + 1):
            categoria_val = ws_cat.cell(row=fila_cat, column=1).value
            num_val = ws_cat.cell(row=fila_cat, column=2).value
            if categoria_val:
                try:
                    num_final = int(num_val) if num_val is not None else 0
                except (TypeError, ValueError):
                    num_final = 0
                insert_categoria_centro(centro_id, str(categoria_val).strip(), num_final)

    def _val(fila, nombre_col, default=""):
        col = headers.get(nombre_col)
        if not col:
            return default
        v = ws.cell(row=fila, column=col).value
        return v if v is not None else default

    detectores_creados = 0
    for fila_actual in range(header_row + 1, ws.max_row + 1):
        codigo_det = _val(fila_actual, "Código")
        sala_det = _val(fila_actual, "Sala")
        if not str(codigo_det).strip() and not str(sala_det).strip():
            continue  # fila vacía, se ignora

        planta = _val(fila_actual, "Planta")
        nivel = _val(fila_actual, "Nivel")
        codigo_sala = _val(fila_actual, "Código de la sala")
        profesionales_multilinea = str(_val(fila_actual, "Profesionales en la sala"))
        profesionales = ", ".join(
            linea.strip() for linea in profesionales_multilinea.split("\n") if linea.strip()
        )
        turno = _val(fila_actual, "Turno de trabajo")
        fecha_colocacion = _val(fila_actual, "Fecha de colocación")
        hora_colocacion = _val(fila_actual, "Hora de colocación")
        fecha_retirada_real = _val(fila_actual, "Fecha de retirada real")
        hora_retirada_real = _val(fila_actual, "Hora de retirada real")
        nombre_plano_fila = _val(fila_actual, "Nombre del plano")
        punto_x_fila = _val(fila_actual, "Punto X", None)
        punto_y_fila = _val(fila_actual, "Punto Y", None)

        plano_centro_id = planos_por_nombre.get(str(nombre_plano_fila)) if nombre_plano_fila else None

        try:
            punto_x_final = float(punto_x_fila) if punto_x_fila not in (None, "") else -1
            punto_y_final = float(punto_y_fila) if punto_y_fila not in (None, "") else -1
        except (TypeError, ValueError):
            punto_x_final = punto_y_final = -1

        foto_sit_path = None
        foto_det_path = None
        col_situacion = headers.get("Foto situación")
        col_detector_foto = headers.get("Foto detector")
        if col_situacion:
            datos = imagenes_por_celda.get((fila_actual - 1, col_situacion - 1))
            if datos:
                foto_sit_path = guardar_bytes_imagen(datos, "foto_situacion")
        if col_detector_foto:
            datos = imagenes_por_celda.get((fila_actual - 1, col_detector_foto - 1))
            if datos:
                foto_det_path = guardar_bytes_imagen(datos, "foto_detector")

        data = (
            centro_id, str(planta), str(sala_det), str(fecha_colocacion), str(codigo_det),
            None, punto_x_final, punto_y_final, foto_sit_path, foto_det_path,
            _ahora_espana().strftime("%Y-%m-%d %H:%M"),
            str(codigo_sala), profesionales, str(hora_colocacion),
            str(turno) if turno else TURNOS_TRABAJO_OPCIONES[0],
            str(nivel) if nivel else NIVEL_OPCIONES[0],
            plano_centro_id, str(fecha_retirada_real), str(hora_retirada_real),
        )
        insert_detector(data)
        detectores_creados += 1

    return centro_id, detectores_creados


def generar_excel(centro_id, output_path):
    """Genera un .xlsx en formato de tabla "plana" (sin ningún bloque de
    cabecera encima), pensado para poder ir pegando debajo los datos de
    otros centros más adelante. El nombre del centro y el área figuran
    como las dos primeras columnas de cada fila, junto con todos los
    datos de cada detector, la fecha de retirada óptima calculada
    automáticamente, una columna vacía para anotar la retirada real, y
    sus tres fotos (plano con el punto marcado, situación, detector)
    incrustadas cada una en su propia celda, centradas y manteniendo su
    proporción real."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    centro = get_centro(centro_id)
    if not centro:
        raise Exception("Centro no encontrado")
    cid, nombre, zona, fecha, img_centro, tecnico, direccion = centro
    detectores = fetch_detectores(centro_id)

    FUENTE = "Arial"
    fuente_normal = Font(name=FUENTE, size=7)
    fuente_cabecera = Font(name=FUENTE, size=8, bold=True, color="FFFFFF")
    fuente_info_grande = Font(name=FUENTE, size=12, bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Estilo especial para la primera columna ("Centro"): sin fondo
    # naranja en las filas de datos, alineado a la derecha.
    alineado_derecha = Alignment(horizontal="right", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detectores"

    borde_fino = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    # --- Línea de cabecera superior con el resumen del centro (Centro,
    # Área, Día, Técnico), y debajo la fila de cabeceras de columna
    # (al doble de alto de lo normal). ---
    fila_info = 1
    header_row = 2
    headers = [
        "Centro", "Área / Zona", "ID", "Planta", "Nivel", "Sala", "Código de la sala",
        "Profesionales en la sala", "Turno de trabajo", "Código",
        "Resultado (Bq/m³/h)", "Incertidumbre",
        "Fecha de colocación", "Hora de colocación",
        "Fecha de retirada óptima", "Fecha de retirada real", "Hora de retirada real",
        "Nombre del plano", "Punto X", "Punto Y",
        "Plano", "Foto situación", "Foto detector",
    ]

    col_letra = {h: get_column_letter(i) for i, h in enumerate(headers, start=1)}
    col_idx0 = {h: i for i, h in enumerate(headers)}  # índice base-0, para los anclajes

    # Línea superior en forma de banner: el nombre del centro ocupa las
    # 4 primeras columnas fusionadas, el área las 4 siguientes, luego
    # una columna con la palabra "Técnico" y las 4 siguientes con su
    # nombre fusionadas. El resto de columnas quedan en naranja liso
    # para que la fila entera se vea como un único banner.
    ws.merge_cells(start_row=fila_info, start_column=1, end_row=fila_info, end_column=4)
    c_centro = ws.cell(row=fila_info, column=1, value=nombre or "")
    c_centro.font = fuente_info_grande
    c_centro.fill = PatternFill("solid", fgColor="F5A623")
    c_centro.alignment = centrado

    ws.merge_cells(start_row=fila_info, start_column=5, end_row=fila_info, end_column=8)
    c_area = ws.cell(row=fila_info, column=5, value=zona or "")
    c_area.font = fuente_info_grande
    c_area.fill = PatternFill("solid", fgColor="F5A623")
    c_area.alignment = centrado

    c_tecnico_label = ws.cell(row=fila_info, column=9, value="Técnico")
    c_tecnico_label.font = fuente_cabecera
    c_tecnico_label.fill = PatternFill("solid", fgColor="F5A623")
    c_tecnico_label.alignment = centrado

    ws.merge_cells(start_row=fila_info, start_column=10, end_row=fila_info, end_column=13)
    c_tecnico_val = ws.cell(row=fila_info, column=10, value=tecnico or "")
    c_tecnico_val.font = fuente_cabecera
    c_tecnico_val.fill = PatternFill("solid", fgColor="F5A623")
    c_tecnico_val.alignment = centrado

    # Fecha del centro: no se muestra a gran tamaño como Centro/Área,
    # pero se guarda en el banner para poder reconstruirla al importar.
    c_fecha_label = ws.cell(row=fila_info, column=14, value="Fecha")
    c_fecha_label.font = fuente_cabecera
    c_fecha_label.fill = PatternFill("solid", fgColor="F5A623")
    c_fecha_label.alignment = centrado
    c_fecha_val = ws.cell(row=fila_info, column=15, value=fecha or "")
    c_fecha_val.font = fuente_cabecera
    c_fecha_val.fill = PatternFill("solid", fgColor="F5A623")
    c_fecha_val.alignment = centrado

    # Dirección del centro: igual que la fecha, se guarda en el banner
    # (aunque no se muestre a gran tamaño) para poder recuperarla si se
    # reimporta este Excel más adelante.
    c_dir_label = ws.cell(row=fila_info, column=16, value="Dirección")
    c_dir_label.font = fuente_cabecera
    c_dir_label.fill = PatternFill("solid", fgColor="F5A623")
    c_dir_label.alignment = centrado
    c_dir_val = ws.cell(row=fila_info, column=17, value=direccion or "")
    c_dir_val.font = fuente_cabecera
    c_dir_val.fill = PatternFill("solid", fgColor="F5A623")
    c_dir_val.alignment = centrado

    # El resto de la fila (hasta donde llegan las columnas de la tabla
    # de abajo) también en naranja con texto centrado, aunque no haya
    # más datos que mostrar, para que quede como un único banner.
    for col in range(18, len(headers) + 1):
        c_resto = ws.cell(row=fila_info, column=col)
        c_resto.fill = PatternFill("solid", fgColor="F5A623")
        c_resto.alignment = centrado
        c_resto.font = fuente_cabecera
    ws.row_dimensions[fila_info].height = 22

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        if i == 1:
            # "Centro": mismo estilo que el resto de cabeceras (texto
            # blanco, fondo naranja), pero alineado a la derecha.
            c.font = fuente_cabecera
            c.fill = PatternFill("solid", fgColor="F5A623")
            c.alignment = alineado_derecha
        else:
            c.font = fuente_cabecera
            c.fill = PatternFill("solid", fgColor="F5A623")
            c.alignment = centrado
        c.border = borde_fino
    ws.row_dimensions[header_row].height = 30  # el doble de una fila normal (~15pt)

    ANCHO_FOTOS = 24
    for h in ("Plano", "Foto situación", "Foto detector"):
        ws.column_dimensions[col_letra[h]].width = ANCHO_FOTOS
    for h in ("Centro", "Planta", "Nivel", "Sala", "Código de la sala", "Profesionales en la sala",
              "Turno de trabajo", "Código"):
        ws.column_dimensions[col_letra[h]].width = 15
    for h in ("Fecha de colocación", "Fecha de retirada óptima", "Fecha de retirada real"):
        ws.column_dimensions[col_letra[h]].width = 16
    for h in ("Hora de colocación", "Hora de retirada real"):
        ws.column_dimensions[col_letra[h]].width = 10
    ws.column_dimensions[col_letra["Área / Zona"]].width = 16
    ws.column_dimensions[col_letra["Nombre del plano"]].width = 14
    for h in ("Punto X", "Punto Y"):
        ws.column_dimensions[col_letra[h]].width = 8

    TAM_IMG_PX = 150  # tamaño MÁXIMO (lado mayor) de cada foto dentro de su celda
    ALTO_FILA_PT = 115  # alto de fila (puntos) reservado para las fotos

    def _ancho_col_px(col_chars):
        # Aproximación estándar: píxeles ≈ caracteres * 7 + 5
        return round(col_chars * 7 + 5)

    def _alto_fila_px(alto_pt):
        # 1 punto ≈ 1.333 píxeles (96 dpi)
        return round(alto_pt * 96 / 72)

    def _anclaje_centrado(col_h, fila1, ancho_img_px, alto_img_px):
        """Ancla la imagen centrada (horizontal y verticalmente) dentro
        de su celda, calculando el hueco libre a cada lado."""
        ancho_celda_px = _ancho_col_px(ws.column_dimensions[col_letra[col_h]].width)
        alto_celda_px = _alto_fila_px(ALTO_FILA_PT)
        off_x = max(0, (ancho_celda_px - ancho_img_px) / 2)
        off_y = max(0, (alto_celda_px - alto_img_px) / 2)
        marker = AnchorMarker(
            col=col_idx0[col_h], row=fila1 - 1,
            colOff=pixels_to_EMU(off_x), rowOff=pixels_to_EMU(off_y),
        )
        size = XDRPositiveSize2D(cx=pixels_to_EMU(ancho_img_px), cy=pixels_to_EMU(alto_img_px))
        return OneCellAnchor(_from=marker, ext=size)

    fila = header_row + 1
    for d in detectores:
        (did, _, planta, sala, fecha_det, codigo, _plano_antiguo, punto_x, punto_y,
         foto_sit_p, foto_det_p, fecha_creacion, codigo_sala, profesionales_sala,
         hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real,
         resultado_bq_m3_d, incertidumbre_d) = d
        plano_p = None
        nombre_plano_d = ""
        if plano_centro_id:
            plano_info = get_plano_centro(plano_centro_id)
            if plano_info:
                plano_p = plano_info[3]
                nombre_plano_d = plano_info[2]

        fecha_retirada_optima = calcular_fecha_retirada_optima(fecha_det)

        # Los profesionales se escriben separados por comas en el
        # formulario; en la celda del Excel se muestran uno debajo de
        # otro (salto de línea + texto ajustado a la celda).
        profesionales_multilinea = "\n".join(
            p.strip() for p in (profesionales_sala or "").split(",") if p.strip()
        )

        hay_punto_valido = (
            punto_x is not None and punto_y is not None and 0 <= punto_x <= 1 and 0 <= punto_y <= 1
        )

        valores = [
            nombre or "", zona or "", did, planta or "", nivel or "", sala or "",
            codigo_sala or "", profesionales_multilinea, turno_trabajo or "", codigo or "",
            resultado_bq_m3_d if resultado_bq_m3_d is not None else "", incertidumbre_d or "",
            fecha_det or "", hora_colocacion or "",
            fecha_retirada_optima, fecha_retirada_real or "", hora_retirada_real or "",
            nombre_plano_d, round(punto_x, 4) if hay_punto_valido else "",
            round(punto_y, 4) if hay_punto_valido else "",
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            if col == 1:
                # "Centro": texto negro (normal), sin fondo, a la derecha.
                c.font = fuente_normal
                c.alignment = alineado_derecha
            else:
                c.font = fuente_normal
                c.alignment = centrado
            c.border = borde_fino
        # Las tres celdas de fotos también con borde, aunque queden vacías
        for h in ("Plano", "Foto situación", "Foto detector"):
            ws[f"{col_letra[h]}{fila}"].border = borde_fino

        ws.row_dimensions[fila].height = ALTO_FILA_PT

        hay_punto = hay_punto_valido

        for col_name, ruta in (
            ("Plano", plano_p),
            ("Foto situación", foto_sit_p),
            ("Foto detector", foto_det_p),
        ):
            # Si es URL de Supabase, descargar
            ruta_local = None
            if ruta and es_url_supabase(ruta):
                temp_path = os.path.join(get_data_dir(), f"_temp_excel_{did}_{col_name}.jpg")
                if descargar_desde_supabase(ruta, temp_path):
                    ruta_local = temp_path
            elif ruta and os.path.exists(ruta):
                ruta_local = ruta
            
            if ruta_local and os.path.exists(ruta_local):
                try:
                    with Image.open(ruta_local) as im_orig:
                        # IMPORTANTE: las fotos de móvil llevan a menudo
                        # metadatos EXIF de rotación que PIL no aplica
                        # solo; sin esto, el ancho/alto "en bruto" no
                        # coincide con cómo se ve realmente la foto y la
                        # proporción calculada sale mal.
                        im = ImageOps.exif_transpose(im_orig)
                        im = im.convert("RGB")
                        if col_name == "Plano" and hay_punto:
                            # Dibujar el punto rojo del detector sobre el
                            # plano, igual que en el informe PDF.
                            draw = ImageDraw.Draw(im)
                            w, h = im.size
                            cx, cy = punto_x * w, punto_y * h
                            r = max(6, int(min(w, h) * 0.02))
                            draw.ellipse(
                                [cx - r, cy - r, cx + r, cy + r],
                                fill=(220, 20, 20), outline=(120, 0, 0), width=2,
                            )
                        # Se sube algo la resolución respecto a antes
                        # (se veía muy borrosa) sin pasarse: esta misma
                        # imagen va dentro del propio Excel, que a
                        # veces se comparte por WhatsApp, así que
                        # engordarla demasiado podía hacer fallar ese
                        # envío por límite de tamaño.
                        im.thumbnail((600, 600))
                        buf = io.BytesIO()
                        im.save(buf, format="JPEG", quality=85)
                        buf.seek(0)
                        ancho_real, alto_real = im.size

                    # Escalar manteniendo la proporción real (sin
                    # deformar), ajustando al lado mayor = TAM_IMG_PX.
                    escala = min(TAM_IMG_PX / ancho_real, TAM_IMG_PX / alto_real)
                    ancho_final = max(1, round(ancho_real * escala))
                    alto_final = max(1, round(alto_real * escala))
                    xl_img = XLImage(buf)
                    xl_img.width = ancho_final
                    xl_img.height = alto_final
                    xl_img.anchor = _anclaje_centrado(col_name, fila, ancho_final, alto_final)
                    ws.add_image(xl_img)
                except Exception:
                    c_err = ws[f"{col_letra[col_name]}{fila}"]
                    c_err.value = "(no se pudo incrustar la imagen)"
                    c_err.font = fuente_normal
                    c_err.alignment = centrado
        fila += 1

    # --- Hoja "Planos": TODOS los planos del centro, uno por fila,
    # tengan o no algún detector que los use (los del "Detectores" solo
    # aparecen si algún detector los tiene asignado; esta hoja evita
    # que un plano "huérfano" se pierda al reimportar). También incluye
    # la foto exterior del centro, que no aparece en ningún otro sitio
    # del Excel. ---
    ws_planos = wb.create_sheet("Planos")

    # Empresa y CIF (datos globales de la empresa que realiza las
    # mediciones), en las dos primeras filas de esta hoja.
    empresa_xl = get_empresa()
    cif_xl = get_cif()
    ws_planos.cell(row=1, column=1, value="Empresa").font = fuente_cabecera
    ws_planos.cell(row=1, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=1, column=2, value=empresa_xl).font = fuente_normal
    ws_planos.cell(row=2, column=1, value="CIF").font = fuente_cabecera
    ws_planos.cell(row=2, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=2, column=2, value=cif_xl).font = fuente_normal

    # Datos del informe final (superficie, nº de plantas, comunicación
    # a los trabajadores) que se rellenan en esa pantalla: se guardan
    # aquí para que no se pierdan si se vuelve a generar el Excel.
    datos_informe_xl = get_datos_informe_centro(centro_id)
    filas_datos_informe = [
        ("Superficie construida", datos_informe_xl["superficie_construida"]),
        ("Superficie útil", datos_informe_xl["superficie_util"]),
        ("Nº de plantas", datos_informe_xl["num_plantas"]),
        ("Fecha comunicación trabajadores", datos_informe_xl["fecha_comunicacion_trab"]),
        ("Medio de comunicación", datos_informe_xl["medio_comunicacion"]),
    ]
    for i, (etiqueta_di, valor_di) in enumerate(filas_datos_informe, start=3):
        ws_planos.cell(row=i, column=1, value=etiqueta_di).font = fuente_cabecera
        ws_planos.cell(row=i, column=1).fill = PatternFill("solid", fgColor="F5A623")
        ws_planos.cell(row=i, column=2, value=valor_di).font = fuente_normal

    header_row_planos = 9
    ws_planos.cell(row=header_row_planos, column=1, value="Nombre").font = fuente_cabecera
    ws_planos.cell(row=header_row_planos, column=1).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.cell(row=header_row_planos, column=2, value="Imagen").font = fuente_cabecera
    ws_planos.cell(row=header_row_planos, column=2).fill = PatternFill("solid", fgColor="F5A623")
    ws_planos.column_dimensions["A"].width = 20
    ws_planos.column_dimensions["B"].width = 28

    fila_planos = header_row_planos + 1

    def _incrustar_en_hoja_planos(nombre_fila, ruta_imagen, fila_actual):
        ws_planos.cell(row=fila_actual, column=1, value=nombre_fila).font = fuente_normal
        ws_planos.row_dimensions[fila_actual].height = 115
        
        # Si es URL de Supabase, descargar
        ruta_local = None
        if ruta_imagen and es_url_supabase(ruta_imagen):
            temp_path = os.path.join(get_data_dir(), f"_temp_excel_plano_{fila_actual}.jpg")
            if descargar_desde_supabase(ruta_imagen, temp_path):
                ruta_local = temp_path
        elif ruta_imagen and os.path.exists(ruta_imagen):
            ruta_local = ruta_imagen
        
        if ruta_local and os.path.exists(ruta_local):
            try:
                with Image.open(ruta_local) as im_orig:
                    im = ImageOps.exif_transpose(im_orig)
                    im = im.convert("RGB")
                    # Resolución alta (aunque en la celda del Excel se
                    # vea pequeña, con este mismo dato se genera luego
                    # el plano automático del Anexo II a tamaño de
                    # página completa, así que hace falta buena
                    # definición de origen para que no se vea borroso).
                    im.thumbnail((2000, 2000))
                    buf = io.BytesIO()
                    im.save(buf, format="JPEG", quality=92)
                    buf.seek(0)
                    ancho_real, alto_real = im.size
                escala = min(180 / ancho_real, 150 / alto_real)
                xl_img = XLImage(buf)
                xl_img.width = max(1, round(ancho_real * escala))
                xl_img.height = max(1, round(alto_real * escala))
                ws_planos.add_image(xl_img, f"B{fila_actual}")
            except Exception:
                ws_planos.cell(row=fila_actual, column=2, value="(no se pudo incrustar la imagen)")

    if img_centro and os.path.exists(img_centro):
        _incrustar_en_hoja_planos("(Foto exterior del centro)", img_centro, fila_planos)
        fila_planos += 1

    for plano_c in fetch_planos_centro(centro_id):
        _, _, nombre_plano_c, ruta_plano_c, _ = plano_c
        _incrustar_en_hoja_planos(nombre_plano_c, ruta_plano_c, fila_planos)
        fila_planos += 1

    # --- Hoja "Categorías profesionales": una fila por categoría con
    # el número de personas expuestas de esa categoría en el centro. ---
    categorias_centro_xl = fetch_categorias_centro(centro_id)
    if categorias_centro_xl:
        ws_cat = wb.create_sheet("Categorías profesionales")
        ws_cat.cell(row=1, column=1, value="Categoría profesional").font = fuente_cabecera
        ws_cat.cell(row=1, column=1).fill = PatternFill("solid", fgColor="F5A623")
        ws_cat.cell(row=1, column=2, value="Nº personas expuestas").font = fuente_cabecera
        ws_cat.cell(row=1, column=2).fill = PatternFill("solid", fgColor="F5A623")
        ws_cat.cell(row=1, column=3, value="Turno").font = fuente_cabecera
        ws_cat.cell(row=1, column=3).fill = PatternFill("solid", fgColor="F5A623")
        ws_cat.column_dimensions["A"].width = 28
        ws_cat.column_dimensions["B"].width = 20
        ws_cat.column_dimensions["C"].width = 22
        for i, (_, _, categoria_xl, num_personas_xl, turno_xl) in enumerate(categorias_centro_xl, start=2):
            ws_cat.cell(row=i, column=1, value=categoria_xl).font = fuente_normal
            ws_cat.cell(row=i, column=2, value=num_personas_xl).font = fuente_normal
            ws_cat.cell(row=i, column=3, value=turno_xl or "").font = fuente_normal

    wb.save(output_path)
    return True


# ============================================================
# REGISTRO PARA LABORATORIO
# Ficha calcada del formulario "FICHA DE IDENTIFICACIÓN E INFORMACIÓN
# DE LOS DETECTORES DE TRAZAS" que exige el laboratorio de análisis,
# para que la acepten sin tener que rellenarla a mano aparte.
# ============================================================

def generar_registro_laboratorio(centro_id, output_path, tipo_firma="digital"):
    """Genera la ficha de identificación de detectores en el formato
    que exige el laboratorio, con una tabla prácticamente idéntica a
    su propio formulario en papel, más un cuadro de firma del técnico.

    tipo_firma="manual": no se añade ningún cuadro de firma (se firma
    a mano sobre el papel impreso).
    tipo_firma="digital": se añade un cuadro pequeño de firma digital
    en la esquina de TODAS las páginas del documento.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.platypus import Image as RLImage
    from reportlab.lib import colors
    from PIL import Image as PILImage

    centro = get_centro(centro_id)
    if not centro:
        raise ValueError("Centro no encontrado")
    _, nombre, zona, fecha, img_ext, tecnico, direccion = centro
    detectores = fetch_detectores(centro_id)
    empresa = get_empresa()

    def _planta_desde_nivel(nivel_valor):
        """El laboratorio no quiere el texto libre de "Nivel" tal cual
        se escribe en la app, sino el número de planta (con signo)
        que le corresponde."""
        return NIVEL_A_PLANTA_LABORATORIO.get(nivel_valor, nivel_valor or "")

    def _ubicado_en(zona_valor, nombre_centro):
        """"Ubicado en" es el área del centro (p.ej. \"Área Sanitaria da
        Coruña e Cee\"). Si esa área está vacía, o es un valor genérico
        que no identifica bien el centro (\"Atención Primaria\", \"PAC\",
        \"Atención Primaria + PAC\", \"Consultorio\"), se usa el nombre del
        propio centro en su lugar."""
        valores_genericos = (
            "atención primaria", "atencion primaria", "pac",
            "atención primaria + pac", "atencion primaria + pac",
            "consultorio",
        )
        if not zona_valor or zona_valor.strip().lower() in valores_genericos:
            return nombre_centro or ""
        return zona_valor

    # Con firma digital reservamos margen inferior de sobra para que el
    # cuadro de firma (dibujado directamente en cada página) no se
    # solape nunca con las últimas filas de la tabla. El margen superior
    # se reserva SIEMPRE para la cabecera (logo + datos + condiciones),
    # que también se dibuja directamente en cada página, para que
    # aparezca completa en todas las hojas, no solo en la primera.
    ALTURA_CABECERA = 2.3*cm
    SEPARACION_CABECERA = 0.3*cm
    margen_superior = 1.2*cm + ALTURA_CABECERA + SEPARACION_CABECERA
    margen_inferior = 3.4*cm if tipo_firma == "digital" else 1.2*cm
    doc = SimpleDocTemplate(
        output_path, pagesize=landscape(A4),
        topMargin=margen_superior, bottomMargin=margen_inferior, leftMargin=1.2*cm, rightMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    story = []

    GRIS_CLARO = colors.HexColor("#D9D9D9")

    estilo_celda = ParagraphStyle(
        "CeldaLab", parent=styles["Normal"], fontSize=7, leading=8.5, alignment=TA_CENTER,
    )
    estilo_celda_izq = ParagraphStyle(
        "CeldaLabIzq", parent=styles["Normal"], fontSize=7.5, leading=9, alignment=TA_LEFT,
    )
    estilo_cab = ParagraphStyle(
        "CabLab", parent=styles["Normal"], fontSize=7, leading=8.5, alignment=TA_CENTER,
        textColor=colors.black, fontName="Helvetica-Bold",
    )

    col_widths = [0.8*cm, 2.4*cm, 2.6*cm, 1.2*cm, 2.6*cm, 3.4*cm, 1.1*cm,
                  1.7*cm, 1.2*cm, 3.4*cm, 1.7*cm, 1.2*cm, 3.4*cm]
    ancho_hasta_mitad_habitacion = sum(col_widths[:4]) + col_widths[4] / 2
    ancho_total = sum(col_widths)

    # --- Logotipo del laboratorio, a la IZQUIERDA (hasta la mitad de
    # la columna "Habitación/Estancia"), y a su derecha, en un único
    # bloque sin líneas internas, los datos de obra/inmueble/cliente y
    # la barra de condiciones del edificio. Ambos con la misma altura.
    # Todo este bloque se dibuja directamente en el lienzo de CADA
    # página (ver _dibujar_elementos_pagina más abajo), no como
    # contenido normal de la tabla, para que aparezca completo en
    # todas las hojas del documento. ---
    ruta_logo = get_logo_laboratorio()
    ruta_logo_default = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "logo_laboratorio_default.png"
    )
    es_logo_default = bool(ruta_logo) and os.path.normpath(ruta_logo) == os.path.normpath(ruta_logo_default)

    contenido_logo = ""
    if ruta_logo and os.path.exists(ruta_logo):
        try:
            with PILImage.open(ruta_logo) as im_logo:
                w_logo, h_logo = im_logo.size
            max_w = ancho_hasta_mitad_habitacion - 0.4*cm
            max_h = ALTURA_CABECERA - 0.4*cm
            r = min(max_w / w_logo, max_h / h_logo)
            contenido_logo = RLImage(ruta_logo, width=w_logo * r, height=h_logo * r)
        except Exception:
            contenido_logo = ""

    def _construir_caja_cabecera():
        """Se reconstruye igual en cada llamada (una por página) para
        evitar reutilizar el mismo objeto Table ya "usado" por
        reportlab al dibujarlo en una página anterior."""
        caja_logo = Table(
            [[contenido_logo]], colWidths=[ancho_hasta_mitad_habitacion], rowHeights=[ALTURA_CABECERA],
        )
        caja_logo.setStyle(TableStyle([
            ("LINEAFTER", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        # El código de formulario solo tiene sentido con el logotipo
        # POR DEFECTO de la app (es el código de SU ficha); si se ha
        # cambiado por el logotipo de otro laboratorio, ese código ya
        # no pinta nada y se quita.
        if es_logo_default:
            celda_orden_trabajo = [
                Paragraph("CF-PE-CYE/39 V.3 (30/05/2025)", ParagraphStyle(
                    "CodigoDerecha", parent=styles["Normal"], fontSize=5.5, leading=7,
                    alignment=TA_RIGHT, textColor=colors.black,
                )),
                Paragraph("<b>Orden de trabajo:</b> ", estilo_celda_izq),
            ]
        else:
            celda_orden_trabajo = Paragraph("<b>Orden de trabajo:</b> ", estilo_celda_izq)

        filas_bloque_derecho = [
            [Paragraph("<b>Ref. de Obra:</b> ", estilo_celda_izq), celda_orden_trabajo],
            [Paragraph(f"<b>Inmueble:</b> {_escapar_pdf(nombre or '')}", estilo_celda_izq),
             Paragraph(f"<b>Cliente:</b> {_escapar_pdf(empresa or '')}", estilo_celda_izq)],
            [Paragraph(
                '<b><span backColor="#D9D9D9" color="black">CONDICIONES DEL EDIFICIO DURANTE LA EXPOSICIÓN:</span></b>',
                estilo_celda_izq,
            ), ""],
        ]
        ancho_bloque_derecho = ancho_total - ancho_hasta_mitad_habitacion
        caja_datos_y_condiciones = Table(
            filas_bloque_derecho, colWidths=[ancho_bloque_derecho * 0.42, ancho_bloque_derecho * 0.58],
            rowHeights=[ALTURA_CABECERA * 0.32, ALTURA_CABECERA * 0.32, ALTURA_CABECERA * 0.36],
        )
        caja_datos_y_condiciones.setStyle(TableStyle([
            ("SPAN", (0, 2), (1, 2)),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))

        cabecera = Table(
            [[caja_logo, caja_datos_y_condiciones]],
            colWidths=[ancho_hasta_mitad_habitacion, ancho_bloque_derecho],
        )
        cabecera.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return cabecera


    # --- Tabla principal de detectores. Cabecera de 3 filas: el
    # título de la ficha (integrado en la propia tabla, sin el nombre
    # de la empresa), y debajo las dos filas de cabeceras de columnas,
    # ahora en gris claro (como el modelo original) en vez de oscuro. ---
    estilo_titulo_tabla = ParagraphStyle(
        "TituloTabla", parent=styles["Normal"], fontSize=12, leading=15,
        alignment=TA_CENTER, fontName="Helvetica-Bold", textColor=colors.black,
    )
    fila_titulo = [Paragraph(
        "FICHA DE IDENTIFICACIÓN E INFORMACIÓN DE LOS DETECTORES DE TRAZAS", estilo_titulo_tabla,
    )] + [""] * 12

    headers_fila1 = [
        "Nº", "Código\nDetector", "Edificio", "Planta", "Habitación / Estancia",
        "Ubicado en", "Foto", "INSTALACIÓN", "", "", "DESINSTALACIÓN", "", "",
    ]
    headers_fila2 = [
        "", "", "", "", "", "", "", "Fecha", "Hora", "Técnico\nFirma", "Fecha", "Hora", "Técnico\nFirma",
    ]
    estilo_cab_pequeno = ParagraphStyle(
        "CabLabPequeno", parent=styles["Normal"], fontSize=5.8, leading=7,
        alignment=TA_CENTER, textColor=colors.black, fontName="Helvetica-Bold",
    )
    filas_tabla = [
        fila_titulo,
        [Paragraph(h.replace("\n", "<br/>"), estilo_cab_pequeno if h in ("Habitación / Estancia", "Planta") else estilo_cab)
         for h in headers_fila1],
        [Paragraph(h.replace("\n", "<br/>"), estilo_cab) for h in headers_fila2],
    ]

    # Suficientes filas en blanco como para llegar al final de la hoja
    # aunque haya pocos detectores (una sola página de sobra). Con
    # firma digital, menos filas para dejar sitio de sobra al cuadro
    # de firma sin que ninguna fila se le pueda solapar.
    n_filas_totales = max(len(detectores) + 3, 14 if tipo_firma == "digital" else 15)
    estilo_firma_celda = ParagraphStyle(
        "FirmaCelda", parent=styles["Normal"], fontSize=6, leading=7.5, alignment=TA_CENTER,
    )
    nombre_tecnico_pdf = _escapar_pdf(tecnico or "")
    celda_firma_tecnico = (
        Paragraph(f"{nombre_tecnico_pdf}<br/>* Ver firma digital abajo", estilo_firma_celda)
        if tipo_firma == "digital" else ""
    )

    for i in range(n_filas_totales):
        # Solo se numeran las filas que realmente tienen datos; las
        # filas en blanco de sobra no llevan número.
        numero_fila = str(i + 1) if i < len(detectores) else ""
        if i < len(detectores):
            d = detectores[i]
            did_d, _, planta_d, sala_d, fecha_d, codigo_d = d[0], d[1], d[2], d[3], d[4], d[5]
            codigo_sala_d, hora_colocacion_d, nivel_d = d[12], d[14], d[16]
            fecha_retirada_real_d, hora_retirada_real_d = d[18], d[19]
            tiene_foto = bool(d[9]) or bool(d[10])
            fila = [
                numero_fila,
                Paragraph(_escapar_pdf(codigo_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(nombre or ""), estilo_celda),
                Paragraph(_escapar_pdf(_planta_desde_nivel(nivel_d)), estilo_celda),
                Paragraph(_escapar_pdf(codigo_sala_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(_ubicado_en(zona, nombre)), estilo_celda),
                "\u2713" if tiene_foto else "",
                Paragraph(_escapar_pdf(fecha_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(hora_colocacion_d or ""), estilo_celda),
                celda_firma_tecnico,
                Paragraph(_escapar_pdf(fecha_retirada_real_d or ""), estilo_celda),
                Paragraph(_escapar_pdf(hora_retirada_real_d or ""), estilo_celda),
                celda_firma_tecnico,
            ]
        else:
            fila = [numero_fila] + [""] * 12
        filas_tabla.append(fila)

    tabla = Table(filas_tabla, colWidths=col_widths, repeatRows=3)
    estilo_tabla = [
        ("SPAN", (0, 0), (12, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), GRIS_CLARO),
        ("BACKGROUND", (0, 1), (-1, 2), GRIS_CLARO),
        ("SPAN", (0, 1), (0, 2)), ("SPAN", (1, 1), (1, 2)), ("SPAN", (2, 1), (2, 2)),
        ("SPAN", (3, 1), (3, 2)), ("SPAN", (4, 1), (4, 2)), ("SPAN", (5, 1), (5, 2)),
        ("SPAN", (6, 1), (6, 2)),
        ("SPAN", (7, 1), (9, 1)), ("SPAN", (10, 1), (12, 1)),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 3), (0, -1), "CENTER"),
        ("ALIGN", (6, 3), (6, -1), "CENTER"),
        ("FONTSIZE", (0, 3), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("ROWBACKGROUNDS", (0, 3), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]
    if tipo_firma == "manual":
        # Filas más altas para que las 15 lleguen igualmente hasta el
        # final de la hoja (aquí no hace falta reservar hueco para
        # ningún cuadro de firma al pie).
        estilo_tabla.append(("TOPPADDING", (0, 3), (-1, -1), 5))
        estilo_tabla.append(("BOTTOMPADDING", (0, 3), (-1, -1), 5))
    tabla.setStyle(TableStyle(estilo_tabla))
    story.append(tabla)
    story.append(Spacer(1, 0.01*cm))

    # --- Cabecera completa (logo + datos + condiciones) y cuadro de
    # firma digital: se dibujan en cada página con onFirstPage/
    # onLaterPages, el mecanismo normal de ReportLab. Por una
    # peculiaridad interna de ReportLab, este mecanismo deja de
    # dibujar en la ÚLTIMA página del documento cuando el contenido
    # ocupa más de una hoja; para compensarlo, después de generar el
    # PDF se comprueba y, si hace falta, se "parchea" esa última
    # página superponiéndole el cuadro que falta. ---
    def _dibujar_elementos_pagina(canvas_obj, doc_obj):
        cabecera_pagina = _construir_caja_cabecera()
        x_cabecera = doc_obj.leftMargin + 6
        y_cabecera = doc_obj.pagesize[1] - 1.2*cm - ALTURA_CABECERA
        cabecera_pagina.wrapOn(canvas_obj, ancho_total, ALTURA_CABECERA)
        cabecera_pagina.drawOn(canvas_obj, x_cabecera, y_cabecera)

        if tipo_firma == "digital":
            canvas_obj.saveState()
            ancho_caja = 4.0*cm
            alto_caja = 1.5*cm
            separacion_tabla = 0.7*cm
            x = doc_obj.leftMargin + 6 + ancho_total - ancho_caja
            y = doc_obj.bottomMargin - separacion_tabla - alto_caja
            canvas_obj.setLineWidth(0.8)
            canvas_obj.rect(x, y, ancho_caja, alto_caja)
            canvas_obj.setFont("Helvetica", 6.5)
            canvas_obj.drawCentredString(x + ancho_caja / 2, y + 5, "* Firma digital")
            canvas_obj.restoreState()

    doc.build(story, onFirstPage=_dibujar_elementos_pagina, onLaterPages=_dibujar_elementos_pagina)

    # --- Parche de la última página: por la citada peculiaridad de
    # ReportLab, si el documento tiene más de una página, la última no
    # recibe el cuadro de firma digital pese a llamarse correctamente
    # a onLaterPages para ella. Se detecta y se corrige aquí,
    # superponiendo justo ese cuadro sobre la última página. ---
    if tipo_firma == "digital":
        from pypdf import PdfReader, PdfWriter
        from reportlab.pdfgen.canvas import Canvas as _CanvasBase
        import io as _io

        buf_overlay = _io.BytesIO()
        c_overlay = _CanvasBase(buf_overlay, pagesize=landscape(A4))
        ancho_caja = 4.0*cm
        alto_caja = 1.5*cm
        separacion_tabla = 0.7*cm
        x_o = 1.2*cm + 6 + ancho_total - ancho_caja
        y_o = margen_inferior - separacion_tabla - alto_caja
        c_overlay.setLineWidth(0.8)
        c_overlay.rect(x_o, y_o, ancho_caja, alto_caja)
        c_overlay.setFont("Helvetica", 6.5)
        c_overlay.drawCentredString(x_o + ancho_caja / 2, y_o + 5, "* Firma digital")
        c_overlay.save()
        buf_overlay.seek(0)

        lector_principal = PdfReader(output_path)
        pagina_overlay = PdfReader(buf_overlay).pages[0]
        escritor = PdfWriter()
        for i, pagina in enumerate(lector_principal.pages):
            if i == len(lector_principal.pages) - 1:
                pagina.merge_page(pagina_overlay)
            escritor.add_page(pagina)
        with open(output_path, "wb") as f:
            escritor.write(f)





# ============================================================
# WIDGETS DE IMAGEN (Subir archivo / Cámara del navegador)
# Sustituyen a los botones "Seleccionar" / "Camara" de la app Tkinter.
# ============================================================

# Solo se permite un flujo de cámara abierto a la vez en TODO el
# formulario (plano, foto de situación, foto del detector, imagen del
# centro). Guarda qué campo "posee" la cámara ahora mismo; el resto no
# monta su propio st.camera_input hasta que ese campo la libere. Esto
# evita el error "no autorizado" que da el móvil cuando varios widgets
# intentan acceder a la cámara al mismo tiempo.
GLOBAL_CAM_OWNER_KEY = "_camara_activa_global"


def widget_imagen(label, state_key, key_prefix, con_camara=True, ancho_miniatura=None,
                   tab_por_defecto="subir", titulo_amarillo=False):
    """Selector de imagen con pestañas 'Subir' y 'Cámara'.

    Guarda automáticamente el archivo elegido en la carpeta de datos y
    recuerda la ruta en st.session_state[state_key]. Devuelve esa ruta
    (o None si no hay imagen). Para quitar una imagen ya subida basta
    con pulsar la "x" del propio archivo en el selector, o subir/sacar
    una nueva foto que la sustituya.

    tab_por_defecto: "subir" (normal) o "camara" (la pestaña de
    cámara aparece primero y por tanto activa por defecto al abrir
    el formulario).
    """
    if titulo_amarillo:
        st.markdown(f'<p class="subtitulo-amarillo">{label}</p>', unsafe_allow_html=True)
    else:
        st.markdown(f"**{label}**")
    file_id_key = key_prefix + "__file_id"
    cam_nonce_key = key_prefix + "__cam_nonce"
    cam_activa_key = key_prefix + "__cam_activa"
    if cam_nonce_key not in st.session_state:
        st.session_state[cam_nonce_key] = 0
    if cam_activa_key not in st.session_state:
        st.session_state[cam_activa_key] = False  # la cámara NO se activa sola

    camara_primero = con_camara and tab_por_defecto == "camara"
    if camara_primero:
        tab_labels = ["📷 Cámara", "📁 Subir"]
        idx_subir, idx_camara = 1, 0
    else:
        tab_labels = ["📁 Subir"] + (["📷 Cámara"] if con_camara else [])
        idx_subir, idx_camara = 0, 1
    tabs = st.tabs(tab_labels)
    nuevo_bytes = None
    nueva_ext = ".jpg"
    vino_de_camara = False

    with tabs[idx_subir]:
        up = st.file_uploader(
            "Selecciona una imagen", type=["png", "jpg", "jpeg"],
            key=key_prefix + "_up", label_visibility="collapsed",
        )
        if up is not None:
            fid = ("up", getattr(up, "file_id", None) or f"{up.name}_{up.size}")
            if st.session_state.get(file_id_key) != fid:
                nuevo_bytes = up.getvalue()
                nueva_ext = extension_de(up)
                st.session_state[file_id_key] = fid

    if con_camara:
        with tabs[idx_camara]:
            hay_imagen = bool(st.session_state.get(state_key))
            # Solo se permite UNA cámara activa a la vez en todo el formulario:
            # si otro campo la tiene abierta, este campo no monta su propio
            # visor (así se evita el error "no autorizado" por streams
            # de cámara simultáneos en el móvil).
            es_el_activo = st.session_state.get(GLOBAL_CAM_OWNER_KEY) == cam_activa_key
            otra_cam_en_uso = (
                st.session_state.get(GLOBAL_CAM_OWNER_KEY) is not None and not es_el_activo
            )

            if st.session_state[cam_activa_key] and es_el_activo:
                cam_key = f"{key_prefix}_cam_{st.session_state[cam_nonce_key]}"
                foto = st.camera_input(
                    "Capturar foto", key=cam_key, label_visibility="collapsed",
                )
                if foto is not None:
                    nuevo_bytes = foto.getvalue()
                    nueva_ext = ".jpg"
                    vino_de_camara = True
            else:
                if hay_imagen:
                    st.caption("Foto capturada (se muestra abajo).")
                etiqueta_btn = "📷 Tomar otra foto" if hay_imagen else "📷 Activar cámara"
                if otra_cam_en_uso:
                    st.caption("La cámara está siendo usada en otro campo. Termina esa foto primero.")
                elif st.button(etiqueta_btn, key=key_prefix + "_activar_cam"):
                    st.session_state[cam_activa_key] = True
                    st.session_state[GLOBAL_CAM_OWNER_KEY] = cam_activa_key
                    st.rerun()

    if nuevo_bytes is not None:
        path = guardar_bytes_imagen(nuevo_bytes, key_prefix, nueva_ext)
        st.session_state[state_key] = path
        # Marca para que quien llame a este widget sepa que ACABA de
        # capturarse/subirse una foto nueva en esta misma interacción
        # (se usa, por ejemplo, para fechar automáticamente la
        # colocación en el momento de sacar la foto del detector).
        st.session_state[key_prefix + "__recien_capturada"] = True
        if vino_de_camara:
            # Ocultamos el visor de la cámara y liberamos el "turno" para
            # que otro campo pueda activarla.
            st.session_state[cam_activa_key] = False
            st.session_state[cam_nonce_key] += 1
            if st.session_state.get(GLOBAL_CAM_OWNER_KEY) == cam_activa_key:
                st.session_state[GLOBAL_CAM_OWNER_KEY] = None
            st.rerun()

    path_actual = st.session_state.get(state_key)
    if path_actual:
        if es_url_supabase(path_actual):
            # Es una URL de Supabase - mostrar directamente
            if ancho_miniatura:
                st.image(path_actual, width=ancho_miniatura)
            else:
                st.image(path_actual, use_container_width=True)
            # Botón para guardar en galería (si es URL de Supabase)
            id_sufijo_galeria = "galeria_" + key_prefix
            boton_compartir_whatsapp_archivo(
                path_actual, os.path.basename(path_actual) or "imagen.jpg", "image/jpeg",
                "Foto del detector de Rn", "💾 Guardar en galería", id_sufijo_galeria,
                "Guardar foto", color_boton="#4285F4",
            )
        elif os.path.exists(path_actual):
            # Es una ruta local
            if ancho_miniatura:
                st.image(path_actual, width=ancho_miniatura)
            else:
                st.image(path_actual, use_container_width=True)
            # Un navegador no puede guardar una foto en la galería del
            # móvil de forma automática y silenciosa (por seguridad: si
            # cualquier página web pudiera escribir archivos solos en tu
            # teléfono, sería una puerta abierta a todo tipo de abusos).
            # Lo más parecido que se puede ofrecer es este botón, que abre
            # el mismo menú nativo de "Compartir" de Android/iOS -un solo
            # toque- para que elijas guardarla en Fotos/Galería.
            id_sufijo_galeria = "galeria_" + key_prefix
            boton_compartir_whatsapp_archivo(
                path_actual, os.path.basename(path_actual), "image/jpeg",
                "Foto del detector de Rn", "💾 Guardar en galería", id_sufijo_galeria,
                "Guardar foto", color_boton="#4285F4",
            )
        else:
            st.caption("Imagen no disponible")

    return st.session_state.get(state_key)


def widget_seleccionar_plano_y_punto(cid, ns):
    """Elige uno de los planos ya cargados en el centro (si hay más de
    uno) y marca con un toque la posición del detector sobre ese plano.

    El PLANO se comparte entre varios detectores del mismo centro (se
    gestiona desde la pantalla del centro), pero el PUNTO es propio de
    cada detector: cada uno marca su propia ubicación sobre el plano
    que le corresponda, aunque compartan la misma imagen de fondo.
    """
    px_key = ns + "_plano_px"
    py_key = ns + "_plano_py"
    sel_key = ns + "_plano_centro_id"

    st.markdown('<p class="subtitulo-amarillo">Plano</p>', unsafe_allow_html=True)

    planos = fetch_planos_centro(cid)
    if not planos:
        st.caption(
            "Este centro todavía no tiene ningún plano cargado. "
            "Añade uno desde la pantalla del centro (sección «🗺️ Planos del centro»)."
        )
        st.session_state[sel_key] = None
        return None, None, None

    ids = [p[0] for p in planos]
    nombres = {p[0]: p[2] for p in planos}
    rutas = {p[0]: p[3] for p in planos}

    actual = st.session_state.get(sel_key)
    if actual not in ids:
        # Ni se había elegido antes, ni el que tenía guardado ya no
        # existe (se pudo borrar): se usa el primero como valor por
        # defecto. Si solo hay un plano, este es siempre el usado.
        actual = ids[0]
        st.session_state[sel_key] = actual

    if len(planos) > 1:
        idx_actual = ids.index(actual)
        idx_sel = st.selectbox(
            "Selecciona el plano", options=list(range(len(ids))),
            format_func=lambda i: nombres[ids[i]], index=idx_actual,
            key=ns + "_plano_selector",
        )
        nuevo_id = ids[idx_sel]
        if nuevo_id != actual:
            st.session_state[sel_key] = nuevo_id
            # Al cambiar de plano, el punto anterior ya no tiene sentido
            # (era una posición relativa al plano anterior).
            st.session_state[px_key] = None
            st.session_state[py_key] = None
            st.rerun()
    else:
        st.caption(f"Plano: {nombres[actual]}")

    plano_id_actual = st.session_state[sel_key]
    plano_path = rutas[plano_id_actual]

    # Si es URL de Supabase, descargar temporalmente
    plano_local = None
    if plano_path:
        if es_url_supabase(plano_path):
            temp_path = os.path.join(get_data_dir(), f"_tmp_plano_view_{plano_id_actual}.jpg")
            if descargar_desde_supabase(plano_path, temp_path):
                plano_local = temp_path
        elif os.path.exists(plano_path):
            plano_local = plano_path
    
    if not plano_local:
        st.warning("No se pudo abrir la imagen de este plano.")
        return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)

    st.caption(
        "Toca sobre el plano para marcar la ubicación de ESTE detector "
        "(el punto rojo aparecerá en el informe)."
    )
    try:
        img_orig = Image.open(plano_local).convert("RGB")
    except Exception:
        st.warning("No se pudo abrir la imagen del plano.")
        return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)

    # Se redimensiona ANTES de mostrarla para que las coordenadas que
    # devuelve el componente coincidan exactamente con esta imagen.
    ancho_max = 680
    escala = min(1.0, ancho_max / img_orig.width)
    disp_w = max(1, int(img_orig.width * escala))
    disp_h = max(1, int(img_orig.height * escala))
    img_disp = img_orig.resize((disp_w, disp_h), Image.Resampling.LANCZOS)

    px = st.session_state.get(px_key)
    py = st.session_state.get(py_key)
    if px is not None and py is not None:
        draw = ImageDraw.Draw(img_disp)
        cx, cy = px * disp_w, py * disp_h
        r = max(6, int(min(disp_w, disp_h) * 0.02))
        draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                     fill=(220, 20, 20), outline=(120, 0, 0), width=2)

    if IMG_COORD_DISPONIBLE:
        coords = streamlit_image_coordinates(img_disp, key=ns + "_plano_coords_" + str(plano_id_actual))
        if coords is not None:
            nuevo_px = max(0.0, min(1.0, coords["x"] / disp_w))
            nuevo_py = max(0.0, min(1.0, coords["y"] / disp_h))
            if (px, py) != (nuevo_px, nuevo_py):
                st.session_state[px_key] = nuevo_px
                st.session_state[py_key] = nuevo_py
                st.rerun()
    else:
        st.image(img_disp, use_container_width=False)
        st.warning(
            "Para marcar el punto sobre el plano instala el componente:\n\n"
            "`pip install streamlit-image-coordinates`"
        )

    return plano_id_actual, st.session_state.get(px_key), st.session_state.get(py_key)


# ============================================================
# COMPARTIR PDF POR WHATSAPP (Web Share API - Android)
# ============================================================

def boton_compartir_whatsapp_archivo(ruta, nombre_archivo, mime_type, texto_mensaje,
                                      etiqueta_boton, id_sufijo, titulo_compartir,
                                      color_boton="#25D366"):
    """Botón genérico que abre el diálogo nativo de "Compartir" de
    Android con UN archivo (PDF, Excel...) ya adjunto (el usuario elige
    WhatsApp), usando la Web Share API. Requisitos del navegador/dispositivo:
      - Android + Chrome (u otro navegador compatible).
      - La app debe servirse por HTTPS (o localhost). Si usas
        `streamlit run` en tu PC y accedes desde el móvil por IP local
        (http://192.168.x.x:8501) el navegador bloqueará esta función
        por no ser un "contexto seguro": usa un túnel HTTPS (p. ej.
        Cloudflare Tunnel, ngrok) o despliega en Streamlit Community
        Cloud.
    
    Si ruta es una URL de Supabase, la descarga primero.
    """
    # Si es URL de Supabase, descargar temporalmente
    ruta_local = ruta
    temp_file = None
    if es_url_supabase(ruta):
        temp_file = os.path.join(get_data_dir(), f"_temp_share_{id_sufijo}.jpg")
        if descargar_desde_supabase(ruta, temp_file):
            ruta_local = temp_file
        else:
            st.warning("No se pudo descargar la imagen para compartir")
            return
    
    if not os.path.exists(ruta_local):
        st.warning("El archivo no está disponible para compartir")
        return
    
    with open(ruta_local, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    texto_js = texto_mensaje.replace("\\", "\\\\").replace("`", "\\`").replace("\n", "\\n")
    nombre_js = nombre_archivo.replace("\\", "\\\\").replace("`", "\\`")
    titulo_js = titulo_compartir.replace("\\", "\\\\").replace("`", "\\`")

    html = f"""
    <div id="wrap-compartir-{id_sufijo}"></div>
    <script>
      const b64_{id_sufijo} = "{b64}";
      const nombreArchivo_{id_sufijo} = `{nombre_js}`;
      const textoMensaje_{id_sufijo} = `{texto_js}`;
      const tituloCompartir_{id_sufijo} = `{titulo_js}`;
      const mimeType_{id_sufijo} = "{mime_type}";

      function b64ToBlob_{id_sufijo}(b64Data, contentType) {{
        const byteChars = atob(b64Data);
        const byteNumbers = new Array(byteChars.length);
        for (let i = 0; i < byteChars.length; i++) {{
          byteNumbers[i] = byteChars.charCodeAt(i);
        }}
        const byteArray = new Uint8Array(byteNumbers);
        return new Blob([byteArray], {{ type: contentType }});
      }}

      // IMPORTANTE: Streamlit ejecuta este código dentro de un <iframe>,
      // y los navegadores bloquean la Web Share API dentro de iframes
      // ("Permission denied") salvo que se invoque desde el documento
      // PRINCIPAL. Por eso el botón real no se pinta aquí dentro, sino
      // que se inyecta en la página principal (fuera del iframe), justo
      // en el hueco que deja este propio iframe.
      const marco_{id_sufijo} = window.frameElement;
      const docPadre_{id_sufijo} = window.parent && window.parent.document;

      function crearUI_{id_sufijo}(doc, contenedorDestino) {{
        // Si de una ejecución anterior ya quedó un botón inyectado
        // (p.ej. al marcar/desmarcar una casilla, que vuelve a
        // ejecutar este script), se quita primero para no duplicarlo.
        const anterior = doc.getElementById("inyectado-{id_sufijo}");
        if (anterior) anterior.remove();

        const cont = doc.createElement("div");
        cont.id = "inyectado-{id_sufijo}";
        cont.style.fontFamily = "'Source Sans Pro', sans-serif";
        const boton = doc.createElement("button");
        boton.textContent = "{etiqueta_boton}";
        boton.style.cssText = "background-color:{color_boton} !important; color:white !important; border:none !important; " +
          "padding:12px 20px !important; border-radius:8px !important; font-size:16px !important; " +
          "font-weight:600 !important; cursor:pointer !important; width:100% !important;";
        const msg = doc.createElement("p");
        msg.style.cssText = "margin-top:8px; font-size:13px; color:#ccc;";
        cont.appendChild(boton);
        cont.appendChild(msg);
        contenedorDestino.appendChild(cont);
        return {{ boton, msg }};
      }}

      let elementos_{id_sufijo};
      if (marco_{id_sufijo} && docPadre_{id_sufijo}) {{
        marco_{id_sufijo}.style.display = "none";
        elementos_{id_sufijo} = crearUI_{id_sufijo}(docPadre_{id_sufijo}, marco_{id_sufijo}.parentNode);
      }} else {{
        elementos_{id_sufijo} = crearUI_{id_sufijo}(document, document.getElementById("wrap-compartir-{id_sufijo}"));
      }}
      const btn_{id_sufijo} = elementos_{id_sufijo}.boton;
      const msg_{id_sufijo} = elementos_{id_sufijo}.msg;
      const navegador_{id_sufijo} = (marco_{id_sufijo} && window.parent) ? window.parent.navigator : navigator;

      btn_{id_sufijo}.addEventListener("click", async () => {{
        try {{
          const blob = b64ToBlob_{id_sufijo}(b64_{id_sufijo}, mimeType_{id_sufijo});
          const file = new File([blob], nombreArchivo_{id_sufijo}, {{ type: mimeType_{id_sufijo} }});

          if (navegador_{id_sufijo}.canShare && navegador_{id_sufijo}.canShare({{ files: [file] }})) {{
            await navegador_{id_sufijo}.share({{
              files: [file],
              title: tituloCompartir_{id_sufijo},
              text: textoMensaje_{id_sufijo},
            }});
          }} else if (navegador_{id_sufijo}.share) {{
            await navegador_{id_sufijo}.share({{ title: tituloCompartir_{id_sufijo}, text: textoMensaje_{id_sufijo} }});
            msg_{id_sufijo}.textContent = "Tu navegador no admite adjuntar el archivo directamente; descárgalo y compártelo manualmente.";
          }} else {{
            msg_{id_sufijo}.textContent = "Tu navegador no soporta compartir archivos. Descárgalo con el botón de arriba y compártelo manualmente desde WhatsApp.";
          }}
        }} catch (err) {{
          if (err.name !== "AbortError") {{
            if (err.name === "NotAllowedError") {{
              msg_{id_sufijo}.textContent = "Tu navegador/dispositivo ha bloqueado el envío directo de este archivo. Descárgalo con el botón de arriba y adjúntalo manualmente en WhatsApp.";
            }} else {{
              msg_{id_sufijo}.textContent = "No se pudo abrir el diálogo de compartir: " + err.message;
            }}
          }}
        }}
      }});
    </script>
    """
    components.html(html, height=110)


def boton_compartir_whatsapp(pdf_path, nombre_archivo, texto_mensaje, id_sufijo="pdf"):
    """Botón que comparte un PDF por WhatsApp (informe o registro para
    laboratorio). id_sufijo debe ser distinto entre botones que puedan
    coexistir en la misma pantalla, o se pisan entre sí."""
    boton_compartir_whatsapp_archivo(
        pdf_path, nombre_archivo, "application/pdf", texto_mensaje,
        "📲 Enviar PDF por WhatsApp", id_sufijo, "Informe de detectores de Rn",
    )


def boton_compartir_whatsapp_excel(xlsx_path, nombre_archivo, texto_mensaje):
    """Botón que comparte la hoja de cálculo del informe por WhatsApp."""
    boton_compartir_whatsapp_archivo(
        xlsx_path, nombre_archivo,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        texto_mensaje, "📊 Enviar Excel por WhatsApp", "xlsx",
        "Hoja de cálculo de detectores de Rn",
    )


def boton_compartir_whatsapp_fotos(imagenes, texto_mensaje, etiqueta_boton="🖼️ Enviar {n} foto(s) por WhatsApp",
                                    id_sufijo="fotos-wa", titulo_compartir="Fotos del informe de detectores de Rn",
                                    color_boton="#25D366"):
    """Botón que comparte por WhatsApp SOLO las fotos marcadas en la
    checklist (sin el PDF), usando la Web Share API con varios archivos
    a la vez. Se separa del PDF a propósito: compartir un PDF mezclado
    con imágenes en el mismo envío falla en bastantes navegadores/apps.

    Cada imagen se comprime/reduce de tamaño antes de codificarla (no
    hace falta la resolución original para verla en WhatsApp), para
    que el conjunto no se vuelva demasiado pesado cuantas más fotos se
    marquen — eso es lo que hacía fallar el envío conjunto cuando
    había muchas fotos de por medio.

    imagenes: lista de tuplas (ruta_en_disco, nombre_de_archivo).
    """
    archivos = []
    for ruta, nombre in imagenes:
        if not ruta:
            continue
        
        # Si es URL de Supabase, descargar
        ruta_local = ruta
        if es_url_supabase(ruta):
            temp_file = os.path.join(get_data_dir(), f"_temp_foto_{len(archivos)}.jpg")
            if descargar_desde_supabase(ruta, temp_file):
                ruta_local = temp_file
            else:
                continue
        
        if not os.path.exists(ruta_local):
            continue
            
        try:
            with Image.open(ruta_local) as im_orig:
                im = ImageOps.exif_transpose(im_orig)
                im = im.convert("RGB")
                im.thumbnail((1280, 1280))
                buf = io.BytesIO()
                im.save(buf, format="JPEG", quality=78)
                datos = buf.getvalue()
            nombre_final = os.path.splitext(nombre)[0] + ".jpg"
            mime_final = "image/jpeg"
        except Exception:
            # Si por lo que sea no se puede recomprimir, se manda el
            # archivo tal cual estaba en disco.
            with open(ruta_local, "rb") as f:
                datos = f.read()
            nombre_final = nombre
            ext = os.path.splitext(nombre)[1].lower()
            mime_final = "image/png" if ext == ".png" else "image/jpeg"

        archivos.append({
            "b64": base64.b64encode(datos).decode("utf-8"),
            "name": nombre_final,
            "type": mime_final,
        })

    if not archivos:
        st.caption("Marca al menos una foto arriba para poder enviarla.")
        return

    archivos_json = json.dumps(archivos)
    texto_json = json.dumps(texto_mensaje)
    n_fotos = len(archivos)
    etiqueta_boton_final = etiqueta_boton.format(n=n_fotos)
    titulo_json = json.dumps(titulo_compartir)

    html = f"""
    <div id="wrap-compartir-{id_sufijo}"></div>
    <script>
      const archivosDataFotos_{id_sufijo.replace('-', '_')} = {archivos_json};
      const textoMensajeFotos_{id_sufijo.replace('-', '_')} = {texto_json};
      const tituloFotos_{id_sufijo.replace('-', '_')} = {titulo_json};

      function b64ToBlobFotos_{id_sufijo.replace('-', '_')}(b64Data, contentType) {{
        const byteChars = atob(b64Data);
        const byteNumbers = new Array(byteChars.length);
        for (let i = 0; i < byteChars.length; i++) {{
          byteNumbers[i] = byteChars.charCodeAt(i);
        }}
        const byteArray = new Uint8Array(byteNumbers);
        return new Blob([byteArray], {{ type: contentType }});
      }}

      // IMPORTANTE: Streamlit ejecuta este código dentro de un <iframe>,
      // y los navegadores bloquean la Web Share API dentro de iframes
      // ("Permission denied") salvo que se invoque desde el documento
      // PRINCIPAL. Por eso el botón real no se pinta aquí dentro, sino
      // que se inyecta en la página principal (fuera del iframe), justo
      // en el hueco que deja este propio iframe.
      const marcoActual_{id_sufijo.replace('-', '_')} = window.frameElement;
      const docPadre_{id_sufijo.replace('-', '_')} = window.parent && window.parent.document;

      function crearUI_{id_sufijo.replace('-', '_')}(doc, contenedorDestino) {{
        // Si de una ejecución anterior ya quedó un botón inyectado
        // (p.ej. al marcar/desmarcar una casilla, que vuelve a
        // ejecutar este script), se quita primero para no duplicarlo.
        const anterior = doc.getElementById("inyectado-{id_sufijo}");
        if (anterior) anterior.remove();

        const cont = doc.createElement("div");
        cont.id = "inyectado-{id_sufijo}";
        cont.style.fontFamily = "'Source Sans Pro', sans-serif";
        const boton = doc.createElement("button");
        boton.textContent = "{etiqueta_boton_final}";
        boton.style.cssText = "background-color:{color_boton} !important; color:white !important; border:none !important; " +
          "padding:12px 20px !important; border-radius:8px !important; font-size:16px !important; " +
          "font-weight:600 !important; cursor:pointer !important; width:100% !important;";
        const msg = doc.createElement("p");
        msg.style.cssText = "margin-top:8px; font-size:13px; color:#ccc;";
        cont.appendChild(boton);
        cont.appendChild(msg);
        contenedorDestino.appendChild(cont);
        return {{ boton, msg }};
      }}

      let elementos_{id_sufijo.replace('-', '_')};
      if (marcoActual_{id_sufijo.replace('-', '_')} && docPadre_{id_sufijo.replace('-', '_')}) {{
        // Se oculta el iframe (se queda vacío) y se inserta el botón
        // real justo a continuación, en el documento principal.
        marcoActual_{id_sufijo.replace('-', '_')}.style.display = "none";
        elementos_{id_sufijo.replace('-', '_')} = crearUI_{id_sufijo.replace('-', '_')}(docPadre_{id_sufijo.replace('-', '_')}, marcoActual_{id_sufijo.replace('-', '_')}.parentNode);
      }} else {{
        // Si por algún motivo no se puede acceder al documento padre
        // (navegador muy restrictivo), se deja el botón aquí dentro
        // como respaldo, aunque pueda dar el mismo error de permiso.
        elementos_{id_sufijo.replace('-', '_')} = crearUI_{id_sufijo.replace('-', '_')}(document, document.getElementById("wrap-compartir-{id_sufijo}"));
      }}
      const btnFotos_{id_sufijo.replace('-', '_')} = elementos_{id_sufijo.replace('-', '_')}.boton;
      const msgFotos_{id_sufijo.replace('-', '_')} = elementos_{id_sufijo.replace('-', '_')}.msg;
      const navegadorParaCompartir_{id_sufijo.replace('-', '_')} = (marcoActual_{id_sufijo.replace('-', '_')} && window.parent) ? window.parent.navigator : navigator;

      btnFotos_{id_sufijo.replace('-', '_')}.addEventListener("click", async () => {{
        try {{
          const files = archivosDataFotos_{id_sufijo.replace('-', '_')}.map(
            a => new File([b64ToBlobFotos_{id_sufijo.replace('-', '_')}(a.b64, a.type)], a.name, {{ type: a.type }})
          );
          if (navegadorParaCompartir_{id_sufijo.replace('-', '_')}.canShare && navegadorParaCompartir_{id_sufijo.replace('-', '_')}.canShare({{ files }})) {{
            await navegadorParaCompartir_{id_sufijo.replace('-', '_')}.share({{
              files: files,
              title: tituloFotos_{id_sufijo.replace('-', '_')},
              text: textoMensajeFotos_{id_sufijo.replace('-', '_')},
            }});
          }} else if (navegadorParaCompartir_{id_sufijo.replace('-', '_')}.share) {{
            await navegadorParaCompartir_{id_sufijo.replace('-', '_')}.share({{ title: tituloFotos_{id_sufijo.replace('-', '_')}, text: textoMensajeFotos_{id_sufijo.replace('-', '_')} }});
            msgFotos_{id_sufijo.replace('-', '_')}.textContent = "Tu navegador no admite adjuntar varias fotos a la vez.";
          }} else {{
            msgFotos_{id_sufijo.replace('-', '_')}.textContent = "Tu navegador no soporta compartir archivos. Descarga las fotos manualmente.";
          }}
        }} catch (err) {{
          if (err.name !== "AbortError") {{
            msgFotos_{id_sufijo.replace('-', '_')}.textContent = "No se pudo abrir el diálogo de compartir: " + err.message;
          }}
        }}
      }});
    </script>
    """
    components.html(html, height=110)


CAMPOS_DETECTOR_TRACKEADOS = (
    "_planta", "_sala", "_fecha", "_codigo", "_codigo_sala", "_profesionales_sala",
    "_hora_colocacion", "_fecha_retirada_real", "_hora_retirada_real",
    "_turno_trabajo", "_nivel", "_plano_centro_id", "_plano_px", "_plano_py",
    "_foto_sit", "_foto_det",
)


def _snapshot_detector(ns):
    """Foto de los valores actuales de un detector en session_state,
    para poder compararla más tarde y saber si hay cambios sin
    guardar."""
    return {suf: st.session_state.get(ns + suf) for suf in CAMPOS_DETECTOR_TRACKEADOS}


def _detector_tiene_cambios(ns):
    """True si los campos del detector abierto en `ns` ahora mismo son
    distintos de los que había la última vez que se guardó (o se
    cargó, si no se ha guardado nunca en esta apertura)."""
    snapshot_guardado = st.session_state.get(ns + "__snapshot")
    if snapshot_guardado is None:
        return False
    return _snapshot_detector(ns) != snapshot_guardado


def _inicializar_ns_detector(cid, detector_id, ns):
    """Carga los datos del detector (o los valores por defecto de uno
    nuevo) en st.session_state, una sola vez por "apertura" (usa
    ns+"__cargado" como bandera para no repetirlo en cada rerun)."""
    init_key = ns + "__cargado"
    if st.session_state.get(init_key):
        return
    if detector_id:
        d = get_detector(detector_id)
        (did, _, planta, sala, fecha, codigo, _plano_antiguo, px, py, foto_sit, foto_det, _,
         codigo_sala, profesionales_sala, hora_colocacion, turno_trabajo, nivel, plano_centro_id, fecha_retirada_real, hora_retirada_real,
         _resultado_ignorado, _incertidumbre_ignorada) = d
    else:
        centro = get_centro(cid)
        planta = sala = codigo = codigo_sala = profesionales_sala = hora_colocacion = turno_trabajo = nivel = ""
        fecha = centro[3] if centro else _ahora_espana().strftime("%d/%m/%Y")
        foto_sit = foto_det = None
        plano_centro_id = None
        fecha_retirada_real = hora_retirada_real = ""
        px = py = None

    st.session_state[ns + "_planta"] = planta or ""
    st.session_state[ns + "_sala"] = sala or ""
    st.session_state[ns + "_fecha"] = fecha or _ahora_espana().strftime("%d/%m/%Y")
    st.session_state[ns + "_codigo"] = codigo or ""
    st.session_state[ns + "_codigo_sala"] = codigo_sala or ""
    st.session_state[ns + "_profesionales_sala"] = profesionales_sala or ""
    # Puede haber varias categorías por sala, guardadas juntas separadas
    # por comas, cada una con su propio turno (p.ej. "Enfermería (2) -
    # Mañana, Celador (1) - Mañana + tarde"), que es el mismo formato
    # que ya esperaba el resto del código (Excel, informe...); aquí se
    # descompone en una lista para poder mostrar y editar cada línea
    # por separado, como en "Categorías profesionales".
    from utils_informe.excel_parser import parse_profesionales_multiples
    _lista_prof_inicial = parse_profesionales_multiples(profesionales_sala)
    st.session_state[ns + "_profesionales_lista"] = _lista_prof_inicial
    st.session_state[ns + "_hora_colocacion"] = hora_colocacion or ""
    st.session_state[ns + "_fecha_retirada_real"] = fecha_retirada_real or ""
    st.session_state[ns + "_hora_retirada_real"] = hora_retirada_real or ""
    # "_turno_trabajo" ahora es un valor derivado (se recalcula solo a
    # partir de la lista de profesionales, no se elige suelto); aquí se
    # calcula igual que se recalculará después, para que no parezca que
    # hay cambios sin guardar nada más abrir un detector ya guardado.
    _turnos_distintos_inicial = sorted({t for _, _, t in _lista_prof_inicial if t})
    st.session_state[ns + "_turno_trabajo"] = " / ".join(_turnos_distintos_inicial)
    st.session_state[ns + "_nivel"] = (
        nivel if nivel in NIVEL_OPCIONES else None
    )
    st.session_state[ns + "_plano_centro_id"] = plano_centro_id
    st.session_state[ns + "_plano_px"] = px if (px is not None and px >= 0) else None
    st.session_state[ns + "_plano_py"] = py if (py is not None and py >= 0) else None
    st.session_state[ns + "_foto_sit"] = foto_sit
    st.session_state[ns + "_foto_det"] = foto_det
    # El código de sala solo se genera solo mientras el detector no
    # se haya guardado nunca con uno ya puesto: esta bandera se fija
    # aquí, una única vez por "apertura" (no en cada repintado, que es
    # cuando ya se ha disparado el autoguardado varias veces), así que
    # refleja fielmente si YA VENÍA con código guardado de una vez
    # anterior o no.
    st.session_state[ns + "_codigo_sala_bloqueado"] = bool((codigo_sala or "").strip())
    st.session_state[ns + "__snapshot"] = _snapshot_detector(ns)
    st.session_state[init_key] = True


def _renderizar_campos_detector(cid, detector_id, ns):
    """Dibuja todos los campos del formulario del detector (sin el
    título, sin el botón "Volver" y sin botones de guardar: los
    cambios se guardan solos). Se usa tanto en la pantalla
    independiente "Nuevo detector" como incrustado dentro de
    "Detectores colocados" al abrir uno ya existente."""
    c1, c2 = st.columns(2)
    with c1:
        st.text_input("Código del detector", key=ns + "_codigo")
        st.text_input("Planta (opcional)", key=ns + "_planta")
    with c2:
        st.text_input("Sala", key=ns + "_sala")
        st.selectbox(
            "Nivel", options=NIVEL_OPCIONES, key=ns + "_nivel",
            index=None, placeholder="Selecciona un nivel",
        )

    categorias_centro_form = fetch_categorias_centro(cid)
    opciones_categoria_form = [c[2] for c in categorias_centro_form]
    if not opciones_categoria_form:
        st.caption(
            "Todavía no has creado ninguna categoría profesional para este centro. "
            "Añádelas en «Categorías profesionales» y luego podrás elegirlas aquí."
        )
        st.session_state[ns + "_profesionales_sala"] = ""
        st.session_state[ns + "_turno_trabajo"] = ""
    else:
        st.markdown("**Profesionales en esta sala** (con su propio turno cada uno)")
        lista_key = ns + "_profesionales_lista"
        lista_actual = st.session_state.get(lista_key, [])

        # Cada línea ya añadida, con su propia casilla para poder
        # quitarla (igual que en "Categorías profesionales").
        if lista_actual:
            seleccionadas_prof = []
            for idx_prof, (cat_linea, num_linea, turno_linea) in enumerate(lista_actual):
                etiqueta_linea = f"{num_linea} {cat_linea}" + (f" — {turno_linea}" if turno_linea else "")
                marcado = st.checkbox(etiqueta_linea, key=f"{ns}_prof_chk_{idx_prof}")
                if marcado:
                    seleccionadas_prof.append(idx_prof)
            if seleccionadas_prof:
                st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                if st.button(f"❌ Eliminar seleccionadas ({len(seleccionadas_prof)})", key=f"{ns}_prof_del"):
                    st.session_state[lista_key] = [
                        entrada for i, entrada in enumerate(lista_actual) if i not in seleccionadas_prof
                    ]
                    st.rerun()
        else:
            st.caption("Todavía no has añadido ningún profesional a esta sala.")

        # Formulario para añadir una línea más: categoría + número +
        # turno, con el mismo truco de "valor pendiente" para que los
        # campos se vacíen solos después de añadir (igual que en
        # "Categorías profesionales").
        pend_add_cat_key = f"{ns}_add_categoria_pend"
        pend_add_num_key = f"{ns}_add_num_pend"
        pend_add_turno_key = f"{ns}_add_turno_pend"
        if pend_add_cat_key in st.session_state:
            st.session_state[f"{ns}_add_categoria"] = st.session_state.pop(pend_add_cat_key)
        if pend_add_num_key in st.session_state:
            st.session_state[f"{ns}_add_num"] = st.session_state.pop(pend_add_num_key)
        if pend_add_turno_key in st.session_state:
            st.session_state[f"{ns}_add_turno"] = st.session_state.pop(pend_add_turno_key)

        categoria_nueva = st.selectbox(
            "Categoría profesional", options=opciones_categoria_form, key=f"{ns}_add_categoria",
            index=None, placeholder="Selecciona una categoría",
        )
        turno_nuevo = st.selectbox(
            "Turno de trabajo", options=TURNOS_TRABAJO_OPCIONES, key=f"{ns}_add_turno",
            index=None, placeholder="Selecciona un turno",
        )
        # El número de personas ofrece hasta el total registrado para
        # esa categoría en "Categorías profesionales" (mínimo 1).
        num_maximo = 20
        if categoria_nueva:
            fila_categoria = next((c for c in categorias_centro_form if c[2] == categoria_nueva), None)
            if fila_categoria:
                num_maximo = max(int(fila_categoria[3] or 1), 1)
        if st.session_state.get(f"{ns}_add_num") not in range(1, num_maximo + 1):
            st.session_state[f"{ns}_add_num"] = 1
        etiqueta_num = "Nº de {} en esta sala en este {}".format(
            categoria_nueva if categoria_nueva else "personas",
            turno_nuevo if turno_nuevo else "turno",
        )
        num_nuevo = st.selectbox(
            etiqueta_num, options=list(range(1, num_maximo + 1)), key=f"{ns}_add_num",
        )
        st.markdown('<div class="marcador-btn-anadir-categoria"></div>', unsafe_allow_html=True)
        if st.button("➕ Añadir", key=f"{ns}_prof_add", use_container_width=True):
            if not categoria_nueva:
                st.warning("Selecciona una categoría")
            else:
                st.session_state[lista_key] = lista_actual + [(categoria_nueva, num_nuevo, turno_nuevo or "")]
                st.session_state[pend_add_cat_key] = None
                st.session_state[pend_add_num_key] = 1
                st.session_state[pend_add_turno_key] = None
                st.rerun()

        # Todas las líneas se guardan combinadas separadas por comas,
        # cada una con su turno (p.ej. "Enfermería (2) - Mañana,
        # Celador (1) - Mañana + tarde"), que es el formato que ya
        # entiende el resto de la app (Excel, informe...). El campo
        # general "Turno de trabajo" del detector (que usan la ficha
        # de laboratorio y el Excel) se calcula solo, juntando los
        # turnos distintos que haya entre todas las líneas.
        lista_final = st.session_state.get(lista_key, [])
        st.session_state[ns + "_profesionales_sala"] = ", ".join(
            f"{cat} ({num})" + (f" - {turno}" if turno else "")
            for cat, num, turno in lista_final
        )
        turnos_distintos = sorted({turno for _, _, turno in lista_final if turno})
        st.session_state[ns + "_turno_trabajo"] = " / ".join(turnos_distintos)

    st.markdown("---")
    widget_seleccionar_plano_y_punto(cid, ns)

    # El código de la sala se genera solo la PRIMERA vez (mientras el
    # detector se está creando y todavía no se ha guardado nunca con
    # un código ya puesto): se recalcula en vivo según el nivel
    # elegido arriba. En cuanto ese detector se guarda una vez con un
    # código ya puesto —al cambiar a otro en el desplegable o al
    # volver a la página del centro, que es cuando se guarda solo—,
    # deja de tocarse solo y pasa a ser un campo normal, editable a
    # mano aquí debajo del plano.
    codigo_sala_ya_fijado = st.session_state.get(ns + "_codigo_sala_bloqueado", False)
    if not codigo_sala_ya_fijado:
        centro_de_este_detector = get_centro(cid)
        zona_centro = centro_de_este_detector[2] if centro_de_este_detector else ""
        tipo_centro_actual = get_tipo_centro(cid)
        nivel_actual = st.session_state.get(ns + "_nivel", NIVEL_OPCIONES[0])
        st.session_state[ns + "_codigo_sala"] = _generar_codigo_sala(
            cid, detector_id, nivel_actual, zona_centro, tipo_centro_actual
        )
    st.text_input("Código de la sala", key=ns + "_codigo_sala")
    if not codigo_sala_ya_fijado:
        st.caption(
            "Se genera automáticamente mientras no se haya guardado todavía; "
            "después podrás corregirlo aquí a mano."
        )

    st.markdown("---")
    st.markdown('<p class="subtitulo-amarillo">Fotos detector</p>', unsafe_allow_html=True)
    fc1, fc2 = st.columns(2)
    with fc1:
        st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
        widget_imagen("Situación del detector", ns + "_foto_sit", key_prefix=ns + "_fsit",
                      tab_por_defecto="camara", titulo_amarillo=True)
    with fc2:
        st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
        widget_imagen("Detector", ns + "_foto_det", key_prefix=ns + "_fdet",
                      tab_por_defecto="camara", titulo_amarillo=True)

    # Al capturar/subir la foto del detector se fecha y hora la
    # colocación automáticamente con el momento exacto. Esta asignación
    # tiene que hacerse ANTES de crear los campos de fecha/hora de
    # aquí abajo (con un rerun inmediato) para no chocar con que ya
    # son widgets con ese mismo key.
    if st.session_state.get(ns + "_fdet__recien_capturada"):
        st.session_state[ns + "_fdet__recien_capturada"] = False
        ahora = _ahora_espana()
        st.session_state[ns + "_fecha"] = ahora.strftime("%d/%m/%Y")
        st.session_state[ns + "_hora_colocacion"] = ahora.strftime("%H:%M")
        st.rerun()

    st.markdown("---")
    st.markdown('<p class="subtitulo-amarillo">Fecha y hora de colocación</p>', unsafe_allow_html=True)
    st.caption("Se rellenan solas al hacer la foto del detector, o se pueden escribir/corregir aquí a mano.")
    fh1, fh2 = st.columns(2)
    with fh1:
        st.text_input("Fecha de colocación", key=ns + "_fecha")
    with fh2:
        st.text_input("Hora de colocación", key=ns + "_hora_colocacion")


def _guardar_detector_desde_ns(cid, detector_id, ns, mostrar_mensajes=True):
    """Guarda en la base de datos los valores que haya ahora mismo en
    st.session_state para ese detector (sirve tanto para el guardado
    manual con botón como para el guardado automático al cambiar de
    detector o salir de la pantalla). Devuelve el id del detector
    guardado (nuevo o existente), o None si no se pudo guardar por
    faltar algún campo obligatorio."""
    if not st.session_state.get(ns + "__cargado"):
        return None  # nunca se llegó a cargar/tocar este formulario

    sala_val = st.session_state.get(ns + "_sala", "").strip()
    codigo_val = st.session_state.get(ns + "_codigo", "").strip()
    if not codigo_val:
        if mostrar_mensajes:
            st.warning("El código es obligatorio")
        return None

    px = st.session_state.get(ns + "_plano_px")
    py = st.session_state.get(ns + "_plano_py")
    data = (
        cid,
        st.session_state.get(ns + "_planta", "").strip(),
        sala_val,
        st.session_state.get(ns + "_fecha", "").strip(),
        codigo_val,
        None,
        px if px is not None else -1,
        py if py is not None else -1,
        st.session_state.get(ns + "_foto_sit"),
        st.session_state.get(ns + "_foto_det"),
        _ahora_espana().strftime("%Y-%m-%d %H:%M"),
        st.session_state.get(ns + "_codigo_sala", "").strip(),
        st.session_state.get(ns + "_profesionales_sala", "").strip(),
        st.session_state.get(ns + "_hora_colocacion", "").strip(),
        st.session_state.get(ns + "_turno_trabajo") or "",
        st.session_state.get(ns + "_nivel") or "",
        st.session_state.get(ns + "_plano_centro_id"),
        st.session_state.get(ns + "_fecha_retirada_real", ""),
        st.session_state.get(ns + "_hora_retirada_real", ""),
    )
    if detector_id:
        update_detector(detector_id, data)
        nuevo_id = detector_id
    else:
        nuevo_id = insert_detector(data)

    if mostrar_mensajes:
        st.success("Detector guardado")
    return nuevo_id


def _guardar_y_actualizar_snapshot_detector(cid, abierto, ns, mostrar_mensajes=True):
    """Guarda el detector abierto (nuevo o existente) y, si sale bien,
    actualiza la "foto" de referencia usada para saber si hay cambios
    sin guardar (con esto, justo después de guardar, ya no aparecerán
    como pendientes). Gestiona también, si aplica, la transición de
    detector "nuevo" a uno ya guardado con id real: traslada el
    estado de sesión a la nueva clave y actualiza el desplegable.
    Devuelve (guardado_ok, ns_final)."""
    detector_id_abierto = None if abierto == "nuevo" else abierto
    guardado_id = _guardar_detector_desde_ns(cid, detector_id_abierto, ns, mostrar_mensajes=mostrar_mensajes)
    if not guardado_id:
        return False, ns

    ns_final = ns
    if abierto == "nuevo":
        # Se traslada el estado "de datos" de sesión de "det_nuevo_*" a
        # "det_{id}_*" (ver explicación detallada en el histórico de
        # cambios de esta función), para no perder lo ya escrito ni
        # las fotos ya subidas al pasar a tener un id real.
        ns_nuevo_real = f"det_{guardado_id}"
        sufijos_a_trasladar = (
            "_planta", "_sala", "_fecha", "_codigo", "_codigo_sala",
            "_codigo_sala_bloqueado", "_profesionales_sala", "_profesionales_lista",
            "_hora_colocacion",
            "_fecha_retirada_real", "_hora_retirada_real", "_turno_trabajo", "_nivel",
            "_plano_centro_id", "_plano_px", "_plano_py", "_foto_sit", "_foto_det",
            "__cargado",
        )
        for _sufijo in sufijos_a_trasladar:
            _k_viejo = ns + _sufijo
            if _k_viejo in st.session_state:
                st.session_state[ns_nuevo_real + _sufijo] = st.session_state.pop(_k_viejo)
        abierto_key = f"detector_abierto_{cid}"
        st.session_state[abierto_key] = guardado_id
        st.session_state["detector_form_ns"] = ns_nuevo_real
        detectores_tras_guardar = fetch_detectores(cid)
        ids_tras_guardar = [d[0] for d in detectores_tras_guardar]
        pend_sel_key = f"selector_detector_pend_{cid}"
        if guardado_id in ids_tras_guardar:
            st.session_state[pend_sel_key] = ids_tras_guardar.index(guardado_id)
        ns_final = ns_nuevo_real

    st.session_state[ns_final + "__snapshot"] = _snapshot_detector(ns_final)
    return True, ns_final


def _limpiar_namespace(ns):
    borrar = [k for k in st.session_state.keys()
              if k == ns or k.startswith(ns + "_") or k.startswith(ns + "__")]
    for k in borrar:
        del st.session_state[k]
    # Si la cámara global la tenía un campo de este formulario, la
    # liberamos para que no quede "atascada" al salir de la pantalla.
    owner = st.session_state.get(GLOBAL_CAM_OWNER_KEY)
    if owner and (owner == ns or owner.startswith(ns + "_") or owner.startswith(ns + "__")):
        st.session_state[GLOBAL_CAM_OWNER_KEY] = None


# ============================================================
# PANTALLA: INICIO (lista de centros)
# ============================================================

def pantalla_inicio():
    st.markdown(
        '<p style="color:#999999; font-size:0.85rem; font-weight:700; margin:0;">'
        'Gestión de muestreo</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p class="titulo-home">☢️ Detectores de Radón</p>',
        unsafe_allow_html=True,
    )

    if "mostrar_form_nuevo_centro" not in st.session_state:
        st.session_state.mostrar_form_nuevo_centro = False
    if "mostrar_form_importar" not in st.session_state:
        st.session_state.mostrar_form_importar = False

    with st.container(border=True):
        bc1, bc2 = st.columns(2)
        with bc1:
            marcadores_nuevo = '<div class="marcador-btn-nuevo-centro"></div>'
            if st.session_state.mostrar_form_nuevo_centro:
                marcadores_nuevo += '<div class="marcador-activo-naranja"></div>'
            st.markdown(marcadores_nuevo, unsafe_allow_html=True)
            if st.button("➕ Nuevo centro", use_container_width=True, type="secondary"):
                abrir = not st.session_state.mostrar_form_nuevo_centro
                st.session_state.mostrar_form_nuevo_centro = abrir
                if abrir:
                    st.session_state.mostrar_form_importar = False
                st.rerun()
        with bc2:
            marcadores_importar = '<div class="marcador-btn-importar"></div>'
            if st.session_state.mostrar_form_importar:
                marcadores_importar += '<div class="marcador-activo-naranja"></div>'
            st.markdown(marcadores_importar, unsafe_allow_html=True)
            if st.button("Importar centro", use_container_width=True, type="secondary",
                         icon=":material/folder_open:"):
                abrir = not st.session_state.mostrar_form_importar
                st.session_state.mostrar_form_importar = abrir
                if abrir:
                    st.session_state.mostrar_form_nuevo_centro = False
                st.rerun()

    if st.session_state.mostrar_form_nuevo_centro:
        # Fuera de un st.form: así el desplegable "Tipo de centro"
        # puede rellenar la casilla "Área / Zona :" al momento, sin
        # esperar a un envío conjunto (los campos de un st.form no
        # reaccionan entre sí hasta que se pulsa el botón).
        nombre_nuevo = st.text_input("Nombre del centro", key="nuevo_centro_nombre")
        tipo_centro_nuevo = st.selectbox(
            "Tipo de centro", options=TIPO_CENTRO_OPCIONES, key="nuevo_centro_tipo",
            index=None, placeholder="Selecciona un tipo de centro",
        )
        area_automatica = TIPO_CENTRO_A_AREA_AUTOMATICA.get(tipo_centro_nuevo, "")
        _sincronizar_valor_auto("nuevo_centro_zona", area_automatica)
        zona_nueva = st.text_input("Área / Zona :", key="nuevo_centro_zona")
        st.markdown('<div class="marcador-btn-crear-centro"></div>', unsafe_allow_html=True)
        crear = st.button("Crear centro", type="primary", key="btn_crear_centro_nuevo")
        if crear:
            if nombre_nuevo and nombre_nuevo.strip():
                cid = crear_centro(
                    nombre_nuevo.strip(),
                    zona_nueva.strip() if zona_nueva else "",
                    tipo_centro_nuevo or "",
                )
                st.session_state.centro_actual = cid
                st.session_state.mostrar_form_nuevo_centro = False
                for _k in ("nuevo_centro_nombre", "nuevo_centro_tipo", "nuevo_centro_zona",
                           "nuevo_centro_zona__ultimo_auto"):
                    st.session_state.pop(_k, None)
                st.session_state.view = "centro"
                st.rerun()
            else:
                st.warning("Escribe un nombre para el centro")

    if st.session_state.mostrar_form_importar:
        st.caption(
            "Reconstruye un centro completo (datos, planos y detectores, con "
            "el punto exacto de cada uno) a partir de un Excel generado por "
            "esta misma app."
        )
        st.markdown('<div class="marcador-uploader-importar"></div>', unsafe_allow_html=True)
        archivo_importar = st.file_uploader(
            "Selecciona el archivo .xlsx", type=["xlsx"], key="importar_centro_file",
        )
        if archivo_importar:
            st.markdown('<div class="marcador-btn-confirmar-importar"></div>', unsafe_allow_html=True)
            if st.button("Importar", key="btn_confirmar_importar", type="primary"):
                try:
                    with st.spinner("Importando centro..."):
                        nuevo_cid, n_detectores = importar_centro_desde_excel(archivo_importar.getvalue())
                    st.session_state.mostrar_form_importar = False
                    st.session_state.centro_actual = nuevo_cid
                    st.session_state.view = "centro"
                    st.success(f"Centro importado correctamente ({n_detectores} detector(es)).")
                    st.rerun()
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"No se pudo importar el archivo: {e}")

    centros = fetch_centros()
    if not centros:
        return

    st.markdown("---")
    st.markdown('<p class="subtitulo-amarillo">Centros registrados</p>', unsafe_allow_html=True)

    opciones = list(range(len(centros)))
    etiquetas = {
        i: f"{c[1] or '(sin nombre)'}"
           + (f" · {c[2]}" if c[2] else "")
        for i, c in enumerate(centros)
    }
    idx_sel = st.selectbox(
        "Centro", options=opciones, format_func=lambda i: etiquetas[i],
        key="selector_centro_home", label_visibility="collapsed",
    )
    cid_sel, nombre_sel, zona_sel, fecha_sel, img_sel = centros[idx_sel]

    b1, b2 = st.columns([3, 1])
    with b1:
        st.markdown('<div class="marcador-btn-abrir-centro"></div>', unsafe_allow_html=True)
        if st.button("Abrir centro", type="primary", use_container_width=True,
                     icon=":material/folder_open:"):
            st.session_state.centro_actual = cid_sel
            st.session_state.view = "centro"
            st.rerun()
    with b2:
        st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
        if st.button("❌ Eliminar", key="btn_eliminar_centro_home"):
            st.session_state["confirmar_borrado_centro"] = cid_sel
            st.rerun()

    if st.session_state.get("confirmar_borrado_centro") == cid_sel:
        st.warning(f"¿Eliminar el centro «{nombre_sel}» y todos sus detectores? Esta acción no se puede deshacer.")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("Sí, eliminar", key=f"confirmar_del_{cid_sel}", type="primary"):
                delete_centro(cid_sel)
                st.session_state["confirmar_borrado_centro"] = None
                st.rerun()
        with cc2:
            if st.button("Cancelar", key=f"cancelar_del_{cid_sel}"):
                st.session_state["confirmar_borrado_centro"] = None
                st.rerun()


# ============================================================
# PANTALLA: CENTRO (datos, detectores, generar PDF)
# ============================================================

def pantalla_centro_datos():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic = st.button("← Volver")
    aviso_placeholder = st.container()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">Datos del centro</p>', unsafe_allow_html=True)

    with st.container(border=True):
        # OJO: Streamlit borra el valor de un text_input de su
        # session_state en cuanto ese widget deja de renderizarse en
        # una ejecución (p.ej. al cerrar este bloque). Por eso aquí
        # se usa "value=" directamente con el dato de la base de
        # datos cada vez que se abre, en vez de precargarlo solo la
        # primera vez (ese patrón perdía los datos al volver).
        img_key = ns_centro + "_img"
        if img_key not in st.session_state:
            st.session_state[img_key] = img_path

        fecha_valor_original = fecha or _ahora_espana().strftime("%d/%m/%Y")

        c1, c2 = st.columns(2)
        with c1:
            st.text_input("Nombre", value=nombre or "", key=ns_centro + "_nombre")
            st.text_input("Área / Zona :", value=zona or "", key=ns_centro + "_zona")
            st.text_input("Dirección", value=direccion or "", key=ns_centro + "_direccion")
            st.text_input("Fecha", value=fecha_valor_original, key=ns_centro + "_fecha")
        with c2:
            # Igual de pequeña que los planos, no a tamaño completo.
            st.markdown('<div class="marcador-imagen-exterior"></div>', unsafe_allow_html=True)
            widget_imagen(
                "Imagen exterior", img_key, key_prefix=ns_centro + "_imgw",
                ancho_miniatura=140, titulo_amarillo=True,
            )

        # El botón de guardar solo aparece si hay algún cambio
        # respecto a lo que ya hay guardado; en cuanto guardas,
        # desaparece de nuevo hasta que vuelvas a tocar algo. Estos
        # cambios también se guardan solos, sin necesidad de este
        # botón, en cuanto se pulsa "Volver al centro" (más abajo).
        hay_cambios = (
            st.session_state[ns_centro + "_nombre"] != (nombre or "")
            or st.session_state[ns_centro + "_zona"] != (zona or "")
            or st.session_state[ns_centro + "_direccion"] != (direccion or "")
            or st.session_state[ns_centro + "_fecha"] != fecha_valor_original
            or st.session_state.get(img_key) != img_path
        )

        def _guardar_cambios_centro():
            nombre_in = st.session_state[ns_centro + "_nombre"].strip()
            if not nombre_in:
                return False
            update_centro(
                cid, nombre_in,
                st.session_state[ns_centro + "_zona"].strip(),
                st.session_state[ns_centro + "_fecha"].strip(),
                st.session_state.get(img_key),
                st.session_state[ns_centro + "_direccion"].strip(),
            )
            return True

        if hay_cambios:
            st.markdown('<div class="marcador-btn-guardar-centro"></div>', unsafe_allow_html=True)
            if st.button("💾 Guardar centro", type="primary"):
                if not _guardar_cambios_centro():
                    st.warning("El nombre es obligatorio")
                else:
                    st.success("Centro guardado")
                    st.rerun()

    # El propio botón "Volver al centro" se dibuja arriba del todo,
    # pero su acción se resuelve aquí abajo, una vez que ya se conoce
    # "hay_cambios" con los valores actuales de los campos: si hay
    # cambios sin guardar, se pide confirmación antes de salir.
    confirm_key = f"datos_centro_confirmar_salida_{cid}"
    if volver_clic:
        if hay_cambios:
            st.session_state[confirm_key] = True
        else:
            st.session_state.view = "centro"
        st.rerun()

    if st.session_state.get(confirm_key):
        with aviso_placeholder:
            st.warning("⚠️ Tienes cambios sin guardar en los datos del centro. ¿Qué quieres hacer?")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                if st.button("💾 Guardar y salir", key=f"cdc_guardar_{cid}", type="primary", use_container_width=True):
                    if _guardar_cambios_centro():
                        st.session_state[confirm_key] = None
                        st.session_state.view = "centro"
                    else:
                        st.warning("El nombre es obligatorio")
                    st.rerun()
            with cc2:
                if st.button("🗑️ Descartar y salir", key=f"cdc_descartar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    st.session_state.view = "centro"
                    st.rerun()
            with cc3:
                if st.button("Cancelar", key=f"cdc_cancelar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    st.rerun()



def pantalla_centro_categorias():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic = st.button("← Volver")
    aviso_placeholder = st.container()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">Categorías profesionales</p>', unsafe_allow_html=True)

    with st.container(border=True):
        # Formulario para añadir, ARRIBA del listado. Los campos se
        # vacían solos tras añadir; como no se puede tocar el valor
        # de un widget ya creado en la misma ejecución, se hace con
        # el mismo truco de "valor pendiente" + rerun de otras partes
        # de la app: se aplica ANTES de crear los propios widgets.
        pend_nombre_key = f"nueva_categoria_nombre_pend_{cid}"
        pend_num_key = f"nueva_categoria_num_pend_{cid}"
        pend_turno_key = f"nueva_categoria_turno_pend_{cid}"
        if pend_nombre_key in st.session_state:
            st.session_state[f"nueva_categoria_nombre_{cid}"] = st.session_state.pop(pend_nombre_key)
        if pend_num_key in st.session_state:
            st.session_state[f"nueva_categoria_num_{cid}"] = st.session_state.pop(pend_num_key)
        if pend_turno_key in st.session_state:
            st.session_state[f"nueva_categoria_turno_{cid}"] = st.session_state.pop(pend_turno_key)

        cnew1, cnew2 = st.columns([2.2, 1.8])
        with cnew1:
            nueva_categoria = st.text_input(
                "Categoría profesional", key=f"nueva_categoria_nombre_{cid}",
            )
        with cnew2:
            st.markdown('<div class="marcador-num-personas"></div>', unsafe_allow_html=True)
            nuevo_num_personas = st.number_input(
                "Nº de profesionales", key=f"nueva_categoria_num_{cid}",
                min_value=0, max_value=999, step=1,
            )
        nuevo_turno_cat = st.selectbox(
            "Turno", options=TURNOS_CATEGORIA_OPCIONES, key=f"nueva_categoria_turno_{cid}",
            index=None, placeholder="Selecciona un turno (opcional)",
        )
        st.caption(
            "Este turno solo se usa para el punto 3 del informe (número de trabajadores "
            "por categoría); no tiene relación con los turnos que pongas por sala en cada "
            "detector."
        )
        st.markdown('<div class="marcador-btn-anadir-categoria"></div>', unsafe_allow_html=True)
        if st.button("➕ Añadir", key=f"add_categoria_{cid}", type="primary"):
            if not nueva_categoria.strip():
                st.warning("Escribe el nombre de la categoría")
            else:
                insert_categoria_centro(cid, nueva_categoria.strip(), int(nuevo_num_personas), nuevo_turno_cat or "")
                st.session_state[pend_nombre_key] = ""
                st.session_state[pend_num_key] = 0
                st.session_state[pend_turno_key] = None
                st.rerun()

        st.markdown("---")

        categorias_centro = fetch_categorias_centro(cid)
        if categorias_centro:
            seleccionadas_cat = []
            for cat_id, _, categoria, num_personas, turno_cat in categorias_centro:
                st.markdown('<div class="marcador-checkbox-categoria"></div>', unsafe_allow_html=True)
                etiqueta_cat = f"{categoria}: {num_personas}" + (f" — {turno_cat}" if turno_cat else "")
                marcado = st.checkbox(
                    etiqueta_cat, key=f"chk_cat_{cat_id}",
                )
                if marcado:
                    seleccionadas_cat.append(cat_id)

            if seleccionadas_cat:
                st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                if st.button(f"❌ Eliminar seleccionadas ({len(seleccionadas_cat)})",
                             key=f"del_cats_sel_{cid}"):
                    for cat_id_del in seleccionadas_cat:
                        delete_categoria_centro(cat_id_del)
                    st.rerun()
        else:
            st.caption("Todavía no se ha añadido ninguna categoría profesional.")

    # Se considera "cambio sin guardar" tener escrito el nombre de una
    # categoría nueva sin haber pulsado todavía "Añadir".
    hay_cambios = bool(nueva_categoria.strip())
    confirm_key = f"categorias_confirmar_salida_{cid}"
    if volver_clic:
        if hay_cambios:
            st.session_state[confirm_key] = True
        else:
            st.session_state.view = "centro"
        st.rerun()

    if st.session_state.get(confirm_key):
        with aviso_placeholder:
            st.warning("⚠️ Tienes escrito el nombre de una categoría nueva sin añadir. ¿Qué quieres hacer?")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                if st.button("➕ Añadir y salir", key=f"cc_guardar_{cid}", type="primary", use_container_width=True):
                    insert_categoria_centro(cid, nueva_categoria.strip(), int(nuevo_num_personas), nuevo_turno_cat or "")
                    st.session_state[confirm_key] = None
                    st.session_state.view = "centro"
                    st.rerun()
            with cc2:
                if st.button("🗑️ Descartar y salir", key=f"cc_descartar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    st.session_state.view = "centro"
                    st.rerun()
            with cc3:
                if st.button("Cancelar", key=f"cc_cancelar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    st.rerun()



def pantalla_centro_planos():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic = st.button("← Volver")
    aviso_placeholder = st.container()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">Planos del centro</p>', unsafe_allow_html=True)

    detectores = fetch_detectores(cid)
    planos_centro = fetch_planos_centro(cid)

    with st.container(border=True):
        st.caption(
            "Carga aquí los planos de las plantas del centro. Cada "
            "detector elegirá luego, en su propia pantalla, sobre cuál de "
            "estos planos marcar su ubicación."
        )

        for plano_c in planos_centro:
            plano_id_c, _, nombre_plano_c, ruta_plano_c, _ = plano_c
            with st.container(border=True):
                pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
                with pcol1:
                    # Si es URL de Supabase o ruta local
                    if ruta_plano_c:
                        if es_url_supabase(ruta_plano_c):
                            # Intentar mostrar desde URL
                            st.image(ruta_plano_c, width=140)
                        elif os.path.exists(ruta_plano_c):
                            st.image(ruta_plano_c, width=140)
                        else:
                            st.caption("(imagen no encontrada)")
                    else:
                        st.caption("(imagen no encontrada)")
                with pcol2:
                    st.markdown(f"**{nombre_plano_c}**")
                with pcol3:
                    st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                    if st.button("❌ Eliminar", key=f"del_plano_{plano_id_c}"):
                        st.session_state["confirmar_borrado_plano"] = plano_id_c
                        st.rerun()

            if st.session_state.get("confirmar_borrado_plano") == plano_id_c:
                st.warning(
                    f"¿Eliminar el plano «{nombre_plano_c}»? Los detectores que lo "
                    "tuvieran asignado se quedarán sin plano ni punto marcado."
                )
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("Sí, eliminar", key=f"conf_del_plano_{plano_id_c}", type="primary"):
                        delete_plano_centro(plano_id_c)
                        st.session_state["confirmar_borrado_plano"] = None
                        st.rerun()
                with cc2:
                    if st.button("Cancelar", key=f"cancel_del_plano_{plano_id_c}"):
                        st.session_state["confirmar_borrado_plano"] = None
                        st.rerun()

        mostrar_add_plano_key = f"mostrar_add_plano_{cid}"
        if mostrar_add_plano_key not in st.session_state:
            st.session_state[mostrar_add_plano_key] = False

        st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
        if st.button("➕ Añadir plano", type="tertiary"):
            st.session_state[mostrar_add_plano_key] = not st.session_state[mostrar_add_plano_key]
            st.rerun()

        nombre_plano_nuevo = ""
        archivo_plano_nuevo = None
        if st.session_state[mostrar_add_plano_key]:
            nombre_plano_nuevo = st.text_input(
                "Nombre del plano (p.ej. «Planta baja», «Planta 1»...)",
                key=f"nuevo_plano_nombre_{cid}",
            )
            st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
            archivo_plano_nuevo = st.file_uploader(
                "Selecciona la imagen del plano", type=["png", "jpg", "jpeg"],
                key=f"nuevo_plano_file_{cid}",
            )
            if archivo_plano_nuevo:
                st.markdown('<div class="marcador-btn-plano-amarillo"></div>', unsafe_allow_html=True)
                if st.button("Guardar plano", key=f"guardar_plano_{cid}", type="primary"):
                    if not nombre_plano_nuevo.strip():
                        st.warning("Ponle un nombre al plano")
                    else:
                        ruta_nueva = guardar_bytes_imagen(
                            archivo_plano_nuevo.getvalue(), f"plano_centro_{cid}",
                            extension_de(archivo_plano_nuevo),
                        )
                        insert_plano_centro(cid, nombre_plano_nuevo.strip(), ruta_nueva, len(planos_centro))
                        st.session_state[mostrar_add_plano_key] = False
                        st.success("Plano añadido")
                    st.rerun()

    # Se considera "cambio sin guardar" tener el panel de "Añadir
    # plano" abierto con un nombre escrito y/o un archivo ya
    # seleccionado, pero sin haber pulsado todavía "Guardar plano".
    hay_cambios = st.session_state[mostrar_add_plano_key] and bool(
        nombre_plano_nuevo.strip() or archivo_plano_nuevo is not None
    )
    confirm_key = f"planos_confirmar_salida_{cid}"
    if volver_clic:
        if hay_cambios:
            st.session_state[confirm_key] = True
        else:
            st.session_state.view = "centro"
        st.rerun()

    if st.session_state.get(confirm_key):
        with aviso_placeholder:
            st.warning("⚠️ Tienes un plano nuevo sin terminar de guardar. ¿Qué quieres hacer?")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                if st.button("💾 Guardar y salir", key=f"cp_guardar_{cid}", type="primary", use_container_width=True):
                    if not nombre_plano_nuevo.strip() or archivo_plano_nuevo is None:
                        st.warning("Hace falta un nombre y una imagen para guardar el plano.")
                    else:
                        ruta_nueva = guardar_bytes_imagen(
                            archivo_plano_nuevo.getvalue(), f"plano_centro_{cid}",
                            extension_de(archivo_plano_nuevo),
                        )
                        insert_plano_centro(cid, nombre_plano_nuevo.strip(), ruta_nueva, len(planos_centro))
                        st.session_state[mostrar_add_plano_key] = False
                        st.session_state[confirm_key] = None
                        st.session_state.view = "centro"
                        st.rerun()
            with cc2:
                if st.button("🗑️ Descartar y salir", key=f"cp_descartar_{cid}", use_container_width=True):
                    st.session_state[mostrar_add_plano_key] = False
                    st.session_state[confirm_key] = None
                    st.session_state.view = "centro"
                    st.rerun()
            with cc3:
                if st.button("Cancelar", key=f"cp_cancelar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    st.rerun()



def pantalla_centro_detectores():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"
    abierto_key = f"detector_abierto_{cid}"
    confirm_key = f"detector_confirmar_salida_{cid}"
    if abierto_key not in st.session_state:
        st.session_state[abierto_key] = None

    # El desplegable de detectores ya está creado como widget más
    # abajo en esta misma función; su valor no se puede reasignar
    # después de creado en la misma ejecución (lo impide Streamlit,
    # de ahí el "pendiente" + rerun, igual que con la fecha/hora).
    pend_sel_key = f"selector_detector_pend_{cid}"
    if pend_sel_key in st.session_state:
        st.session_state[f"selector_detector_{cid}"] = st.session_state.pop(pend_sel_key)

    def _ejecutar_accion_detector(accion):
        """Lleva a cabo de verdad una acción de navegación (sin mirar
        si hay cambios sin guardar: eso ya se ha decidido antes de
        llamar a esta función)."""
        abierto_actual = st.session_state.get(abierto_key)
        if abierto_actual is not None:
            _limpiar_namespace(f"det_{abierto_actual}")
        tipo = accion["tipo"]
        if tipo == "volver":
            st.session_state[abierto_key] = None
            st.session_state.view = "centro"
        elif tipo == "nuevo":
            st.session_state[abierto_key] = "nuevo"
            st.session_state["detector_form_ns"] = "det_nuevo"
        elif tipo == "cambiar":
            st.session_state[abierto_key] = accion["detector_id"]
            st.session_state["detector_form_ns"] = f"det_{accion['detector_id']}"
        st.session_state[confirm_key] = None
        st.rerun()

    def _pedir_navegar(accion, hay_cambios):
        if hay_cambios:
            st.session_state[confirm_key] = accion
            st.rerun()
        else:
            _ejecutar_accion_detector(accion)

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic = st.button("← Volver")
    aviso_placeholder = st.container()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">Detectores colocados</p>', unsafe_allow_html=True)

    detectores = fetch_detectores(cid)
    nuevo_clic = False
    cambiar_accion = None

    with st.container(border=True):
        st.markdown('<div class="marcador-btn-nuevo-detector"></div>', unsafe_allow_html=True)
        if st.button("➕ Nuevo detector", type="tertiary"):
            nuevo_clic = True

        if not detectores:
            st.info("Todavía no se han añadido detectores para este centro.")
        else:
            opciones = list(range(len(detectores)))
            etiquetas = {}
            for i, d in enumerate(detectores):
                did, _, planta, sala, fecha_det, codigo = d[0], d[1], d[2], d[3], d[4], d[5]
                partes = [codigo or f"Detector {did}"]
                if sala:
                    partes.append(sala)
                if planta:
                    partes.append(f"Planta {planta}")
                etiquetas[i] = " · ".join(partes)

            idx_sel = st.selectbox(
                "Detector", options=opciones, format_func=lambda i: etiquetas[i],
                key=f"selector_detector_{cid}", label_visibility="collapsed",
                index=None, placeholder="Selecciona un detector",
            )
            if idx_sel is not None:
                d_sel = detectores[idx_sel]
                did_sel = d_sel[0]

                # El detector que se ve desplegado es SIEMPRE el que está
                # elegido en este desplegable (salvo que haya cambios sin
                # guardar en el que estuviera abierto, en cuyo caso se pide
                # confirmación antes). La única excepción es "Nuevo
                # detector" recién pulsado: no se debe cerrar solo por no
                # coincidir con lo que hubiera seleccionado antes en la lista.
                if (st.session_state.get(abierto_key) != did_sel
                        and st.session_state.get(abierto_key) != "nuevo"):
                    cambiar_accion = {"tipo": "cambiar", "detector_id": did_sel}

                st.markdown('<div class="marcador-btn-eliminar"></div>', unsafe_allow_html=True)
                if st.button("❌ Eliminar", key="btn_eliminar_detector_home"):
                    st.session_state["confirmar_borrado_det"] = did_sel
                    st.rerun()

                if st.session_state.get("confirmar_borrado_det") == did_sel:
                    st.warning("¿Eliminar este detector? Esta acción no se puede deshacer.")
                    cc1, cc2 = st.columns(2)
                    with cc1:
                        if st.button("Sí, eliminar", key=f"conf_del_det_{did_sel}", type="primary"):
                            delete_detector(did_sel)
                            st.session_state["confirmar_borrado_det"] = None
                            if st.session_state.get(abierto_key) == did_sel:
                                st.session_state[abierto_key] = None
                            st.rerun()
                    with cc2:
                        if st.button("Cancelar", key=f"cancel_del_det_{did_sel}"):
                            st.session_state["confirmar_borrado_det"] = None
                            st.rerun()

    # --- Datos del detector abierto (nuevo o ya existente), desplegados
    # aquí mismo debajo del selector en vez de en otra pantalla. Ya no
    # se guardan solos: hace falta pulsar "Guardar detector" (que solo
    # aparece si hay algún cambio), y si se intenta salir con cambios
    # sin guardar (cambiando de detector, pulsando "Nuevo detector" o
    # "Volver al centro"), se pide confirmación antes. ---
    abierto = st.session_state.get(abierto_key)
    hay_cambios = False
    if abierto is not None:
        detector_id_abierto = None if abierto == "nuevo" else abierto
        ns = f"det_{abierto}"
        st.session_state["detector_form_ns"] = ns
        _inicializar_ns_detector(cid, detector_id_abierto, ns)

        st.markdown("---")
        titulo_abierto = "Datos del detector" if detector_id_abierto else "Nuevo detector"
        st.markdown(f'<p class="subtitulo-amarillo">{titulo_abierto}</p>', unsafe_allow_html=True)

        with st.container(border=True):
            _renderizar_campos_detector(cid, detector_id_abierto, ns)

            hay_cambios = _detector_tiene_cambios(ns)
            if hay_cambios:
                st.markdown('<div class="marcador-btn-guardar-detector"></div>', unsafe_allow_html=True)
                if st.button("💾 Guardar detector", type="primary", use_container_width=True, key=f"guardar_{ns}"):
                    guardado_ok, ns_final = _guardar_y_actualizar_snapshot_detector(cid, abierto, ns)
                    if guardado_ok:
                        _limpiar_namespace(ns_final)
                        st.session_state[abierto_key] = None
                        # _guardar_y_actualizar_snapshot_detector deja el
                        # desplegable apuntando al detector recién
                        # guardado (para no perder la selección en el
                        # caso normal); aquí se quiere lo contrario -que
                        # la ficha se cierre del todo-, así que se
                        # sobreescribe para que el desplegable quede en
                        # blanco también.
                        st.session_state[f"selector_detector_pend_{cid}"] = None
                        st.rerun()

    # --- Resolver la acción de navegación solicitada (si la hay),
    # pidiendo confirmación primero si hace falta). Mientras haya un
    # diálogo de confirmación ya pendiente de resolver, no se procesan
    # solicitudes nuevas (si no, el propio desplegable, que sigue
    # mostrando la nueva selección hasta que se resuelve el diálogo,
    # volvería a detectarse como "solicitud de cambio" en cada
    # repintado, entrando en un bucle infinito de reruns). ---
    accion_pendiente = st.session_state.get(confirm_key)
    if not accion_pendiente:
        if volver_clic:
            _pedir_navegar({"tipo": "volver"}, hay_cambios)
        elif nuevo_clic:
            _pedir_navegar({"tipo": "nuevo"}, hay_cambios)
        elif cambiar_accion:
            _pedir_navegar(cambiar_accion, hay_cambios)

    if accion_pendiente:
        with aviso_placeholder:
            st.warning("⚠️ Tienes cambios sin guardar en este detector. ¿Qué quieres hacer?")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                if st.button("💾 Guardar y continuar", key=f"cd_guardar_{cid}", type="primary", use_container_width=True):
                    guardado_ok, _ = _guardar_y_actualizar_snapshot_detector(cid, abierto, f"det_{abierto}")
                    if guardado_ok:
                        _ejecutar_accion_detector(accion_pendiente)
            with cc2:
                if st.button("🗑️ Descartar y continuar", key=f"cd_descartar_{cid}", use_container_width=True):
                    _ejecutar_accion_detector(accion_pendiente)
            with cc3:
                if st.button("Cancelar", key=f"cd_cancelar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    if accion_pendiente["tipo"] == "cambiar" and detectores:
                        ids_actuales = [d[0] for d in detectores]
                        abierto_actual = st.session_state.get(abierto_key)
                        if abierto_actual in ids_actuales:
                            st.session_state[pend_sel_key] = ids_actuales.index(abierto_actual)
                    st.rerun()




def pantalla_centro_retirada():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"
    confirm_key = f"retirada_confirmar_salida_{cid}"
    pend_sel_key = f"selector_retirada_pend_{cid}"
    if pend_sel_key in st.session_state:
        st.session_state[f"selector_retirada_{cid}"] = st.session_state.pop(pend_sel_key)

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic = st.button("← Volver")
    aviso_placeholder = st.container()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">Retirada de detectores</p>', unsafe_allow_html=True)

    detectores = fetch_detectores(cid)
    hay_cambios = False
    did_r = None
    cambiar_sel_accion = None

    def _guardar_retirada_actual(did, fr_val, hr_val):
        actualizar_retirada_detector(did, fr_val.strip(), hr_val.strip())
        st.session_state[f"retirada_snapshot_{did}"] = (fr_val.strip(), hr_val.strip())

    with st.container(border=True):
        if not detectores:
            st.info("Todavía no se han añadido detectores para este centro.")
        else:
            opciones_r = list(range(len(detectores)))
            etiquetas_r = {}
            for i, d in enumerate(detectores):
                did_i, _, planta_i, sala_i, _, codigo_i = d[0], d[1], d[2], d[3], d[4], d[5]
                partes_i = [codigo_i or f"Detector {did_i}"]
                if sala_i:
                    partes_i.append(sala_i)
                if planta_i:
                    partes_i.append(f"Planta {planta_i}")
                etiquetas_r[i] = " · ".join(partes_i)

            idx_sel_r = st.selectbox(
                "Detector a retirar", options=opciones_r, format_func=lambda i: etiquetas_r[i],
                key=f"selector_retirada_{cid}", label_visibility="collapsed",
                index=None, placeholder="Selecciona un detector",
            )
            if idx_sel_r is None:
                st.info("Elige un detector arriba para gestionar su retirada.")
            else:
                d_r = detectores[idx_sel_r]
                did_r = d_r[0]
                codigo_r = d_r[5]
                sala_r = d_r[3]
                foto_sit_r = d_r[9]
                punto_x_r = d_r[7]
                punto_y_r = d_r[8]
                plano_centro_id_r = d_r[17]  # columna plano_centro_id (ver orden de init_db)
                fecha_retirada_real_r = d_r[18]
                hora_retirada_real_r = d_r[19]

                # Detector realmente abierto en esta pantalla (puede no
                # coincidir con lo que muestra el desplegable si hay un
                # cambio de selección pendiente de confirmar).
                abierto_retirada_key = f"retirada_abierto_{cid}"
                if abierto_retirada_key not in st.session_state:
                    st.session_state[abierto_retirada_key] = did_r
                if st.session_state[abierto_retirada_key] != did_r:
                    cambiar_sel_accion = {"tipo": "cambiar", "detector_id": did_r, "idx": idx_sel_r}
                    # mientras se resuelve, se sigue mostrando el detector
                    # que estaba realmente abierto, no el recién elegido.
                    did_mostrar = st.session_state[abierto_retirada_key]
                    d_mostrar = next((d for d in detectores if d[0] == did_mostrar), d_r)
                    did_r = d_mostrar[0]
                    codigo_r = d_mostrar[5]
                    sala_r = d_mostrar[3]
                    foto_sit_r = d_mostrar[9]
                    punto_x_r = d_mostrar[7]
                    punto_y_r = d_mostrar[8]
                    plano_centro_id_r = d_mostrar[17]
                    fecha_retirada_real_r = d_mostrar[18]
                    hora_retirada_real_r = d_mostrar[19]

                st.markdown(f"**Código:** {codigo_r or '-'}")
                st.markdown(f"**Sala:** {sala_r or '-'}")

                fr_key = f"retirada_fecha_{did_r}"
                hr_key = f"retirada_hora_{did_r}"
                pend_fr_key = f"retirada_fecha_pend_{did_r}"
                pend_hr_key = f"retirada_hora_pend_{did_r}"
                snapshot_key = f"retirada_snapshot_{did_r}"

                # Igual que con la fecha/hora de colocación: al ser
                # widgets ya creados, su valor solo se puede actualizar
                # mediante un valor "pendiente" aplicado justo ANTES de
                # crearlos (+ rerun), nunca después.
                if pend_fr_key in st.session_state:
                    st.session_state[fr_key] = st.session_state.pop(pend_fr_key)
                if pend_hr_key in st.session_state:
                    st.session_state[hr_key] = st.session_state.pop(pend_hr_key)

                if fr_key not in st.session_state:
                    st.session_state[fr_key] = fecha_retirada_real_r or ""
                if hr_key not in st.session_state:
                    st.session_state[hr_key] = hora_retirada_real_r or ""
                if snapshot_key not in st.session_state:
                    st.session_state[snapshot_key] = (fecha_retirada_real_r or "", hora_retirada_real_r or "")

                st.text_input("Fecha de retirada", key=fr_key)
                st.text_input("Hora de retirada", key=hr_key)

                hay_cambios = (
                    st.session_state[fr_key].strip(), st.session_state[hr_key].strip()
                ) != st.session_state[snapshot_key]

                st.markdown('<div class="marcador-btn-retirada-amarillo"></div>', unsafe_allow_html=True)
                if st.button("🕒 Capturar fecha y hora", key=f"capturar_retirada_{did_r}", use_container_width=True):
                    ahora = _ahora_espana()
                    st.session_state[pend_fr_key] = ahora.strftime("%d/%m/%Y")
                    st.session_state[pend_hr_key] = ahora.strftime("%H:%M")
                    st.rerun()

                if hay_cambios:
                    st.markdown('<div class="marcador-btn-guardar-detector"></div>', unsafe_allow_html=True)
                    if st.button("💾 Guardar retirada", type="primary", use_container_width=True, key=f"guardar_retirada_{did_r}"):
                        _guardar_retirada_actual(did_r, st.session_state[fr_key], st.session_state[hr_key])
                        st.session_state[f"retirada_abierto_{cid}"] = None
                        st.session_state[f"selector_retirada_pend_{cid}"] = None
                        st.rerun()

                # Plano (con el punto de este detector) y foto de situación,
                # al final del todo y uno al lado del otro (se reutiliza el
                # mismo marcador que fuerza a st.columns a quedarse en fila
                # también en el móvil, en vez de apilarse como hace
                # Streamlit por defecto).
                st.markdown("---")
                st.markdown('<div class="marcador-tabla-resultado-fila"></div>', unsafe_allow_html=True)
                fcol1, fcol2 = st.columns(2)
                with fcol1:
                    st.caption("Plano")
                    if (plano_centro_id_r and punto_x_r is not None and punto_y_r is not None
                            and punto_x_r >= 0 and punto_y_r >= 0):
                        plano_info_r = get_plano_centro(plano_centro_id_r)
                        if plano_info_r:
                            plano_path_r = plano_info_r[3]
                            if es_url_supabase(plano_path_r):
                                # Descargar temporalmente
                                temp_plano = os.path.join(get_data_dir(), f"_tmp_plano_retirada_{did_r}.jpg")
                                if descargar_desde_supabase(plano_path_r, temp_plano):
                                    plano_path_r = temp_plano
                            if os.path.exists(plano_path_r):
                                ruta_tmp_punto_r = os.path.join(
                                    get_data_dir(), f"_tmp_plano_punto_retirada_{did_r}.jpg"
                                )
                                if generar_plano_con_punto(plano_path_r, punto_x_r, punto_y_r, ruta_tmp_punto_r):
                                    st.image(ruta_tmp_punto_r, use_container_width=True)
                                else:
                                    st.caption("No se pudo generar el plano con el punto")
                            else:
                                st.caption("Sin plano asignado")
                        else:
                            st.caption("Sin plano asignado")
                    else:
                        st.caption("Este detector no tiene plano ni punto marcado")
                with fcol2:
                    st.caption("Foto de situación")
                    if foto_sit_r:
                        if es_url_supabase(foto_sit_r):
                            st.image(foto_sit_r, use_container_width=True)
                        elif os.path.exists(foto_sit_r):
                            st.image(foto_sit_r, use_container_width=True)
                        else:
                            st.caption("Sin foto de situación")
                    else:
                        st.caption("Sin foto de situación")

    # --- Confirmación antes de perder cambios sin guardar, al volver
    # al centro o al cambiar de detector en el desplegable ---
    accion_pendiente = st.session_state.get(confirm_key)
    if not accion_pendiente:
        if volver_clic:
            if hay_cambios:
                st.session_state[confirm_key] = {"tipo": "volver"}
                st.rerun()
            else:
                st.session_state.view = "centro"
                st.rerun()
        elif cambiar_sel_accion:
            if hay_cambios:
                st.session_state[confirm_key] = cambiar_sel_accion
                st.rerun()
            else:
                st.session_state[f"retirada_abierto_{cid}"] = cambiar_sel_accion["detector_id"]
                st.rerun()

    accion_pendiente = st.session_state.get(confirm_key)
    if accion_pendiente:
        with aviso_placeholder:
            st.warning("⚠️ Tienes cambios sin guardar en la retirada de este detector. ¿Qué quieres hacer?")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                if st.button("💾 Guardar y continuar", key=f"cr_guardar_{cid}", type="primary", use_container_width=True):
                    _guardar_retirada_actual(did_r, st.session_state[f"retirada_fecha_{did_r}"], st.session_state[f"retirada_hora_{did_r}"])
                    st.session_state[confirm_key] = None
                    if accion_pendiente["tipo"] == "volver":
                        st.session_state.view = "centro"
                    else:
                        st.session_state[f"retirada_abierto_{cid}"] = accion_pendiente["detector_id"]
                    st.rerun()
            with cc2:
                if st.button("🗑️ Descartar y continuar", key=f"cr_descartar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    if accion_pendiente["tipo"] == "volver":
                        st.session_state.view = "centro"
                    else:
                        st.session_state[f"retirada_abierto_{cid}"] = accion_pendiente["detector_id"]
                    st.rerun()
            with cc3:
                if st.button("Cancelar", key=f"cr_cancelar_{cid}", use_container_width=True):
                    st.session_state[confirm_key] = None
                    if accion_pendiente["tipo"] == "cambiar" and "idx" in accion_pendiente:
                        st.session_state[pend_sel_key] = None  # se recalcula abajo
                        detectores_actuales = fetch_detectores(cid)
                        ids_actuales = [d[0] for d in detectores_actuales]
                        abierto_actual = st.session_state.get(f"retirada_abierto_{cid}")
                        if abierto_actual in ids_actuales:
                            st.session_state[pend_sel_key] = ids_actuales.index(abierto_actual)
                    st.rerun()




def _sincronizar_valor_auto(key, valor_calculado):
    """Para los cuadros de texto editables cuyo contenido se calcula
    automáticamente a partir de otros datos (punto 3, conclusiones):
    Streamlit solo tiene en cuenta el "value=" la primera vez que se
    crea el widget; en las siguientes ejecuciones, aunque cambien los
    datos de los que depende, seguiría mostrando lo mismo si no se
    hace esto. Aquí se actualiza st.session_state[key] con el nuevo
    valor calculado SOLO si el usuario no lo ha modificado a mano
    (es decir, si el texto seguía siendo igual al último valor
    calculado); si el usuario ya escribió lo suyo, no se toca."""
    prev_key = key + "__ultimo_auto"
    if key not in st.session_state:
        st.session_state[key] = valor_calculado
        st.session_state[prev_key] = valor_calculado
    elif st.session_state.get(prev_key) == st.session_state[key]:
        st.session_state[key] = valor_calculado
        st.session_state[prev_key] = valor_calculado
    else:
        st.session_state[prev_key] = valor_calculado


def _miniatura_pdf(pdf_bytes, ancho_px=140):
    """Genera una miniatura (PNG en bytes) de la primera página de un PDF,
    para previsualizarlo en pequeño en la propia app. Devuelve None si no
    se puede generar (p.ej. si el paquete pymupdf no estuviera disponible,
    o el archivo no fuera un PDF válido), sin que eso rompa nada más."""
    try:
        import pymupdf
        with pymupdf.open(stream=pdf_bytes, filetype="pdf") as doc:
            pagina = doc[0]
            escala = ancho_px / pagina.rect.width
            pix = pagina.get_pixmap(matrix=pymupdf.Matrix(escala, escala))
            return pix.tobytes("png")
    except Exception:
        return None


def pantalla_centro_informe_completo():
    """Genera el informe oficial completo (Word, con anexos opcionales) a
    partir de los datos ya cargados en la app para este centro: genera el
    Excel automáticamente (como si se hubiera pulsado "Generar" en Informes
    y descargas) y lo carga sin necesidad de subirlo a mano."""
    import sys as _sys
    import os as _os
    import pandas as pd
    _dir_app = _os.path.dirname(_os.path.abspath(__file__))
    if _dir_app not in _sys.path:
        _sys.path.insert(0, _dir_app)

    from utils_informe.excel_parser import (
        COL_CODIGO_DETECTOR, ExcelFormatError, areas_muestreadas, categorias_resumen,
        categorias_turnos_bullets, extraer_resultados_pdf_laboratorio, filter_group,
        group_options, load_workbook, merge_resultados, postos_traballo_bullets,
        salas_medidas, traducir_es_gl,
    )
    from utils_informe.docx_generator import (
        ReportContext, _area_grupo, _quitar_turno_de_puestos, generar_conclusion_automatica,
        generar_texto_objeto_automatico, generar_texto_punto3_automatico,
        generar_texto4_automatico, generar_texto5_automatico, generate_report,
    )
    from utils_informe.pdf_tools import PdfToolsError, construir_pdf_completo, libreoffice_disponible
    from utils_informe.anexo2 import Anexo2Error, extraer_datos_planos, generar_documento_anexo2
    from utils_informe.assets import anexo3_por_defecto, anexo4_por_defecto

    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    if st.button("← Volver"):
        if st.session_state.get(f"ic_hay_cambios_{cid}"):
            with st.spinner("Actualizando la hoja Excel con los últimos cambios..."):
                nombre_xlsx_auto = _nombre_documento(nombre, "HOJA-DATOS") + ".xlsx"
                ruta_xlsx_auto = os.path.join(get_data_dir(), nombre_xlsx_auto)
                generar_excel(cid, ruta_xlsx_auto)
                st.session_state["ultimo_excel"] = ruta_xlsx_auto
                st.session_state["ultimo_excel_nombre"] = nombre_xlsx_auto
                st.session_state["ultimo_excel_centro"] = cid
            st.session_state[f"ic_hay_cambios_{cid}"] = False
        st.session_state.view = "centro_informes"
        st.rerun()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p class="subtitulo-amarillo">INFORME DE RESULTADOS DE MEDICIONES DE Rn</p>',
        unsafe_allow_html=True,
    )

    detectores = fetch_detectores(cid)
    if not detectores:
        st.info("Añade al menos un detector a este centro para poder generar el informe.")
        return

    # --- Generar y cargar el Excel automáticamente, sin pedirlo a mano ---
    ruta_excel = os.path.join(get_data_dir(), f"_tmp_informe_completo_{cid}.xlsx")
    try:
        generar_excel(cid, ruta_excel)
        wb = load_workbook(ruta_excel)
    except ExcelFormatError as e:
        st.error(f"El Excel generado no tiene el formato esperado: {e}")
        return
    except Exception as e:
        st.error(f"No se ha podido generar/cargar el Excel de este centro: {e}")
        return

    df = wb["detectores"]
    det_meta = wb["detectores_meta"]
    planos_meta = wb["planos_meta"]
    categorias_df = wb["categorias"]

    opciones = group_options(df, "Centro")
    if not opciones:
        st.error("El Excel generado no tiene ningún valor en la columna 'Centro'.")
        return
    selected_value = opciones[0]  # un solo centro: el actual
    df_center = filter_group(df, selected_value, "Centro")

    areas = areas_muestreadas(df_center)
    salas = salas_medidas(df_center)
    total_personas, categorias_texto = categorias_resumen(categorias_df)
    tipo_centro_actual = (get_tipo_centro(cid) or "").strip().lower()
    # Solo cuenta como "Atención Primaria" para el horario de 14 a 21
    # h si el tipo de centro incluye ese texto exactamente; "PAC" a
    # secas (sin "Atención Primaria" delante) no cuenta.
    es_ap_centro = any(p in tipo_centro_actual for p in ("atención primaria", "atencion primaria"))
    postos_bullets_default = postos_traballo_bullets(df_center, es_atencion_primaria=es_ap_centro)
    total_traballadores, categorias_bullets_default = categorias_turnos_bullets(
        categorias_df, es_atencion_primaria=es_ap_centro,
    )

    st.success(f"Excel generado y cargado automáticamente: {len(df_center)} detector(es) en «{selected_value}».")

    logo_bytes, _logo_nombre = get_logo_informe()

    # =========================================================
    # 1. IDENTIFICACIÓN DEL CENTRO DE TRABAJO
    # =========================================================
    xerencia_previa = str(planos_meta.get("Empresa", ""))
    cif_previa = str(planos_meta.get("CIF", ""))
    enderezo_previo = str(det_meta.get("Dirección", ""))
    completo_1 = bool(
        st.session_state.get(f"ic_xerencia_{cid}", xerencia_previa).strip()
        and st.session_state.get(f"ic_cif_{cid}", cif_previa).strip()
        and st.session_state.get(f"ic_enderezo_{cid}", enderezo_previo).strip()
    )
    with _acordeon_informe("1", "IDENTIFICACIÓN DEL CENTRO DE TRABAJO", completo_1):
        st.caption("Autorrellenados a partir de los datos ya guardados en la app. Revísalos y complétalos.")
        col1, col2 = st.columns(2)
        with col1:
            xerencia = st.text_input("Xerencia (Empresa)", value=xerencia_previa, key=f"ic_xerencia_{cid}")
            cif = st.text_input("CIF", value=cif_previa, key=f"ic_cif_{cid}")
            centro_nombre = st.text_input("Nombre completo del centro", value=selected_value, key=f"ic_centro_nombre_{cid}")
            servizo_unidade = st.text_input("Servizo / Unidade mostrexada (Área)", value=", ".join(areas), key=f"ic_servizo_{cid}")
            tipo_zona_elegido = ""
            # Si el Área/Zona ya empieza por "Unidad"/"Servicio" (o su
            # forma en galego, "unidade"/"Servizo"), no hace falta
            # preguntar con el checklist: se entiende directamente de
            # lo que ya se ha escrito, y se separa la palabra
            # ("unidade"/"servizo", para las plantillas) del resto del
            # nombre del área (que es lo que luego se traduce y se
            # muestra en el informe).
            _match_zona_directa = re.match(
                r"^(unidad|unidade|servicio|servizo)\b\s*(?:de\s+)?(.*)$",
                servizo_unidade.strip(), re.IGNORECASE,
            ) if servizo_unidade.strip() else None
            if _match_zona_directa:
                _tipo_palabra = _match_zona_directa.group(1).lower()
                tipo_zona_elegido = "unidade" if _tipo_palabra in ("unidad", "unidade") else "servizo"
                servizo_unidade_efectivo = _match_zona_directa.group(2).strip()
            elif _area_grupo(servizo_unidade) == "B":
                st.caption(
                    "El Área/Zona no es Atención Primaria, PAC, Atención Primaria + PAC ni "
                    "está en blanco: indica si es una Unidad, un Servicio, u otro tipo de zona."
                )
                opcion_tipo_zona = st.radio(
                    "¿Qué es?", options=["Unidad", "Servicio", "Otro tipo de zona"],
                    key=f"ic_tipo_zona_{cid}", horizontal=True, label_visibility="collapsed",
                )
                tipo_zona_elegido = {"Unidad": "unidade", "Servicio": "servizo", "Otro tipo de zona": "outro"}[opcion_tipo_zona]
                if tipo_zona_elegido == "outro":
                    preview_zona = ""
                else:
                    preview_zona = f"{opcion_tipo_zona} de {servizo_unidade}" if servizo_unidade else ""
                st.session_state[f"ic_preview_zona_{cid}"] = preview_zona
                st.text_input(
                    "Vista previa (así aparecerá en el informe; en blanco si es \"Otro tipo de zona\", para completarlo tú a mano)",
                    disabled=True, key=f"ic_preview_zona_{cid}",
                )
                servizo_unidade_efectivo = servizo_unidade
            else:
                servizo_unidade_efectivo = servizo_unidade
            enderezo = st.text_input("Dirección (Enderezo)", value=enderezo_previo, key=f"ic_enderezo_{cid}")
        with col2:
            _datos_informe_guardados = get_datos_informe_centro(cid)

            def _guardar_dato_informe(campo, key_widget):
                set_datos_informe_centro(cid, **{campo: st.session_state.get(key_widget, "")})
                st.session_state[f"ic_hay_cambios_{cid}"] = True

            superficie_construida = st.text_input(
                "Superficie construida (m²)", value=_datos_informe_guardados["superficie_construida"],
                key=f"ic_sup_constr_{cid}",
                on_change=_guardar_dato_informe, args=("superficie_construida", f"ic_sup_constr_{cid}"),
            )
            superficie_util = st.text_input(
                "Superficie útil (m²)", value=_datos_informe_guardados["superficie_util"],
                key=f"ic_sup_util_{cid}",
                on_change=_guardar_dato_informe, args=("superficie_util", f"ic_sup_util_{cid}"),
            )
            num_plantas = st.text_input(
                "N.º de plantas", value=_datos_informe_guardados["num_plantas"],
                key=f"ic_num_plantas_{cid}",
                on_change=_guardar_dato_informe, args=("num_plantas", f"ic_num_plantas_{cid}"),
            )
            # Por defecto, la fecha del informe es la de hoy (el día en
            # que se genera), aunque se puede cambiar a mano; se
            # muestra en formato día/mes/año.
            fecha_default = _ahora_espana().date()
            data_informe = st.date_input(
                "Fecha del informe", value=fecha_default, key=f"ic_fecha_informe_{cid}",
                format="DD/MM/YYYY",
            )

    # =========================================================
    # 2. OBJETO DEL INFORME (siempre gris: solo necesita el nombre
    # del centro, que siempre está disponible)
    # =========================================================
    with _acordeon_informe("2", "OBJETO DEL INFORME", True):
        _ctx_previa_objeto = ReportContext(
            centro=traducir_es_gl(centro_nombre), servizo_unidade=traducir_es_gl(servizo_unidade_efectivo),
            tipo_zona=tipo_zona_elegido,
        )
        _texto_objeto_auto = generar_texto_objeto_automatico(_ctx_previa_objeto)
        _sincronizar_valor_auto(f"ic_objeto_{cid}", _texto_objeto_auto)
        objeto_manual = st.text_area(
            "texto_objeto", label_visibility="collapsed", height=140, key=f"ic_objeto_{cid}",
        )
        objeto_es_auto = st.session_state.get(f"ic_objeto_{cid}__ultimo_auto") == objeto_manual

    # =========================================================
    # 3. INFORMACIÓN SOBRE LOS TRABAJADORES
    # =========================================================
    completo_3 = bool(str(total_traballadores or "0").strip() not in ("", "0"))
    with _acordeon_informe("3", "INFORMACIÓN SOBRE LOS TRABAJADORES", completo_3):
        st.caption("Puestos por sala y quendas por categoría, calculados desde los datos guardados. Editables.")
        col3, col4 = st.columns(2)
        with col3:
            st.markdown("**Puestos de trabajo por sala** (una línea por puesto)")
            _sincronizar_valor_auto(f"ic_postos_{cid}", "\n".join(postos_bullets_default))
            postos_text = st.text_area(
                "postos_text", label_visibility="collapsed", height=120, key=f"ic_postos_{cid}",
            )
            postos_bullets = [line.strip() for line in postos_text.splitlines() if line.strip()]
            num_traballadores_total = st.text_input(
                "N.º total de trabajadores adscritos", value=str(total_traballadores or ""), key=f"ic_num_trab_{cid}",
            )
        with col4:
            st.markdown("**Categorías profesionales y quendas** (una línea por categoría)")
            _sincronizar_valor_auto(f"ic_categorias_{cid}", "\n".join(categorias_bullets_default))
            categorias_text = st.text_area(
                "categorias_text", label_visibility="collapsed", height=120, key=f"ic_categorias_{cid}",
            )
            categorias_bullets = [line.strip() for line in categorias_text.splitlines() if line.strip()]
            data_informacion_traballadores = st.text_input(
                "Fecha de comunicación a los trabajadores", value=_datos_informe_guardados["fecha_comunicacion_trab"],
                key=f"ic_data_info_trab_{cid}",
                on_change=_guardar_dato_informe, args=("fecha_comunicacion_trab", f"ic_data_info_trab_{cid}"),
            )
            medio_informacion_traballadores = st.text_input(
                "Medio de comunicación", value=_datos_informe_guardados["medio_comunicacion"] or "correo electrónico",
                key=f"ic_medio_info_{cid}",
                on_change=_guardar_dato_informe, args=("medio_comunicacion", f"ic_medio_info_{cid}"),
            )

        st.markdown("**Punto 3 completo del informe** (tal como quedaría; puedes modificarlo directamente)")
        _ctx_previa_punto3 = ReportContext(
            centro=traducir_es_gl(centro_nombre), servizo_unidade=traducir_es_gl(servizo_unidade_efectivo), postos_bullets=postos_bullets,
            num_traballadores_total=num_traballadores_total, categorias_bullets=categorias_bullets,
            medio_informacion_traballadores=medio_informacion_traballadores,
            data_informacion_traballadores=data_informacion_traballadores,
            tipo_zona=tipo_zona_elegido,
        )
        _texto_punto3_auto = generar_texto_punto3_automatico(_ctx_previa_punto3)
        _sincronizar_valor_auto(f"ic_punto3_{cid}", _texto_punto3_auto)
        texto_punto3_manual = st.text_area(
            "texto_punto3", label_visibility="collapsed", height=220, key=f"ic_punto3_{cid}",
        )
        # Si no se ha tocado (sigue igual que el último texto
        # calculado automáticamente, en gallego), se puede volver a
        # generar en el idioma que haga falta a la hora de generar el
        # informe; si el usuario ya escribió lo suyo a mano, se
        # mantiene igual en los dos idiomas (solo hay un cuadro de
        # texto, no uno por idioma).
        punto3_es_auto = st.session_state.get(f"ic_punto3_{cid}__ultimo_auto") == texto_punto3_manual
        st.caption("Las líneas que empiecen por '-' se muestran como viñetas en el informe.")

    # =========================================================
    # 4. CONDICIONES DE LA EXPOSICIÓN (siempre gris: texto fijo)
    # =========================================================
    with _acordeon_informe("4", "CONDICIONES DE LA EXPOSICIÓN", True):
        _texto4_auto = generar_texto4_automatico()
        _sincronizar_valor_auto(f"ic_texto4_{cid}", _texto4_auto)
        texto4_manual = st.text_area(
            "texto4", label_visibility="collapsed", height=160, key=f"ic_texto4_{cid}",
        )
        texto4_es_auto = st.session_state.get(f"ic_texto4_{cid}__ultimo_auto") == texto4_manual

    # =========================================================
    # 5. PLANOS (siempre gris: texto fijo)
    # =========================================================
    with _acordeon_informe("5", "PLANOS", True):
        _texto5_auto = generar_texto5_automatico()
        _sincronizar_valor_auto(f"ic_texto5_{cid}", _texto5_auto)
        texto5_manual = st.text_area(
            "texto5", label_visibility="collapsed", height=120, key=f"ic_texto5_{cid}",
        )
        texto5_es_auto = st.session_state.get(f"ic_texto5_{cid}__ultimo_auto") == texto5_manual

    # =========================================================
    # 6. RESULTADO DE LAS MEDICIONES REALIZADAS
    # =========================================================
    df_working = df_center.copy()

    resultados_pdf_file_key = f"ic_resultados_pdf_{cid}"
    pendientes_previo = int(df_working["Resultado Bq/m3"].isna().sum()) if "Resultado Bq/m3" in df_working else len(df_working)
    completo_6 = pendientes_previo == 0 and len(df_working) > 0

    with _acordeon_informe("6", "RESULTADO DE LAS MEDICIONES REALIZADAS", completo_6):
        st.caption(
            "Ya vienen los datos disponibles (código de zona, código de detector, fechas, "
            "puestos). Rellena el resultado y la incertidumbre a mano en la tabla, o sube el "
            "PDF de resultados del laboratorio para completarlos automáticamente."
        )

        resultados_pdf_file = st.file_uploader(
            "PDF de resultados del laboratorio (opcional) — se buscan los códigos de detector "
            "que coincidan y se rellenan solos la concentración y la incertidumbre",
            type=["pdf"], key=resultados_pdf_file_key,
        )
        if resultados_pdf_file:
            try:
                resultados_pdf_df = extraer_resultados_pdf_laboratorio(resultados_pdf_file)
                if resultados_pdf_df.empty:
                    st.warning(
                        "No se ha encontrado ninguna fila de resultados con el formato esperado "
                        "en este PDF."
                    )
                else:
                    coincidencias = set(resultados_pdf_df["Código"]) & set(
                        df_working[COL_CODIGO_DETECTOR].astype(str).str.strip()
                    )
                    df_working = merge_resultados(df_working, resultados_pdf_df)

                    # Se incluyen los códigos de detector actuales en
                    # el identificador, para que corregir un código y
                    # volver a subir el mismo PDF sí se note.
                    firma_codigos_pdf = ",".join(sorted(df_working[COL_CODIGO_DETECTOR].astype(str)))
                    fid_resultados_pdf = f"{resultados_pdf_file.name}_{resultados_pdf_file.size}_{firma_codigos_pdf}"
                    if st.session_state.get(f"ic_resultados_pdf_fid_{cid}") != fid_resultados_pdf:
                        st.session_state[f"ic_resultados_pdf_fid_{cid}"] = fid_resultados_pdf
                        for idx_r, fila_r in df_working.reset_index(drop=True).iterrows():
                            resultado_r = fila_r.get("Resultado Bq/m3")
                            try:
                                resultado_r_val = float(resultado_r) if pd.notna(resultado_r) else None
                            except (TypeError, ValueError):
                                resultado_r_val = None
                            st.session_state[f"ic_resultado_{cid}_{idx_r}"] = resultado_r_val
                            incert_r = fila_r.get("Incerteza expandida e K")
                            incert_r_val = (
                                "" if incert_r is None or (isinstance(incert_r, float) and pd.isna(incert_r))
                                else str(incert_r)
                            )
                            st.session_state[f"ic_incerteza_{cid}_{idx_r}"] = incert_r_val
                            detector_id_r = fila_r.get("ID")
                            try:
                                detector_id_r = int(detector_id_r) if pd.notna(detector_id_r) else None
                            except (TypeError, ValueError):
                                detector_id_r = None
                            if detector_id_r:
                                actualizar_resultado_detector(detector_id_r, resultado_r_val, incert_r_val)
                        st.session_state[f"ic_hay_cambios_{cid}"] = True

                    if coincidencias:
                        st.success(
                            f"Encontrados {len(coincidencias)} detector(es) de este centro en el PDF: "
                            + ", ".join(sorted(coincidencias))
                        )
                    else:
                        st.warning(
                            "El PDF se ha leído correctamente, pero ninguno de sus códigos de "
                            "detector coincide con los de este centro."
                        )
            except Exception as e:
                st.error(f"No se ha podido leer el PDF de resultados: {e}")

        # Tabla propia (no st.data_editor: ese componente pinta la
        # cabecera y las celdas ya confirmadas dentro de un <canvas>,
        # con los colores fijados por el tema oscuro de la app -letra
        # blanca sobre fondo oscuro-, sin ninguna forma de forzarlo
        # por CSS). Aquí se dibuja una tabla real, lo más parecida
        # posible a la que sale en el informe final: una fila de
        # cabecera con las mismas columnas del punto 6, y debajo, una
        # fila por detector con los datos ya rellenados y dos campos
        # editables (Resultado e Incertidumbre).
        df_working_reset = df_working.reset_index(drop=True)
        resultado_valores = []
        incerteza_valores = []

        def _texto_seguro(valor):
            """Convierte a texto tratando NaN/None como cadena vacía
            (evita que aparezca literalmente "nan" cuando el dato falta)."""
            if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                return ""
            return str(valor)

        with st.container(border=True):
            st.markdown('<div class="marcador-tabla-resultado-fila"></div>', unsafe_allow_html=True)
            cabA, cabB, cabC = st.columns([1, 1, 1])
            with cabA:
                st.markdown("**Código zona / detector**")
            with cabB:
                st.markdown("**Fechas**")
            with cabC:
                st.markdown("**Sala / puestos**")
            st.markdown("---")

            for idx, fila in df_working_reset.iterrows():
                codigo_zona_v = _texto_seguro(fila.get("Código de la sala", ""))
                codigo_v = _texto_seguro(fila.get("Código", ""))
                sala_v = _texto_seguro(fila.get("Sala", ""))
                fecha_ini_v = _texto_seguro(fila.get("Fecha de colocación fmt", ""))
                fecha_fin_v = _texto_seguro(fila.get("Fecha de retirada real fmt", ""))
                puestos_v = _quitar_turno_de_puestos(_texto_seguro(fila.get("Profesionales en la sala", "")))
                detector_id_fila = fila.get("ID")
                try:
                    detector_id_fila = int(detector_id_fila) if pd.notna(detector_id_fila) else None
                except (TypeError, ValueError):
                    detector_id_fila = None

                resultado_previo = fila.get("Resultado Bq/m3")
                try:
                    resultado_previo_val = float(resultado_previo) if pd.notna(resultado_previo) else None
                except (TypeError, ValueError):
                    resultado_previo_val = None
                incerteza_previa = _texto_seguro(fila.get("Incerteza expandida e K", ""))

                filaA, filaB, filaC = st.columns([1, 1, 1])
                with filaA:
                    st.markdown(f"{html.escape(codigo_zona_v)}<br/>**{html.escape(codigo_v)}**", unsafe_allow_html=True)
                with filaB:
                    st.markdown(f"{html.escape(fecha_ini_v)}<br/>{html.escape(fecha_fin_v)}", unsafe_allow_html=True)
                with filaC:
                    texto_sala_puestos = html.escape(sala_v)
                    if puestos_v:
                        texto_sala_puestos += f"<br/><span style='font-size:0.85em;'>{html.escape(puestos_v)}</span>"
                    st.markdown(texto_sala_puestos, unsafe_allow_html=True)

                key_resultado = f"ic_resultado_{cid}_{idx}"
                key_incerteza = f"ic_incerteza_{cid}_{idx}"
                # El valor inicial ("value=") solo se tiene en cuenta
                # la primera vez que se crea el campo; en repintados
                # posteriores Streamlit usa lo que haya en
                # session_state, así que aquí solo se rellena si
                # todavía no existe (la actualización al subir un PDF
                # nuevo ya se hace aparte, más arriba, escribiendo
                # directamente en session_state).
                if key_resultado not in st.session_state:
                    st.session_state[key_resultado] = resultado_previo_val
                if key_incerteza not in st.session_state:
                    st.session_state[key_incerteza] = incerteza_previa

                def _guardar_resultado_cambio(did=detector_id_fila, k_res=key_resultado, k_inc=key_incerteza):
                    # Se guarda en cuanto se cambia el valor (no hace
                    # falta ningún botón aparte), para que quede en la
                    # base de datos y aparezca también en el Informe
                    # PDF de colocación y en el Excel, no solo en este
                    # informe.
                    if did:
                        actualizar_resultado_detector(
                            did, st.session_state.get(k_res), st.session_state.get(k_inc, "").strip(),
                        )
                        st.session_state[f"ic_hay_cambios_{cid}"] = True

                f3, f4 = st.columns(2)
                with f3:
                    valor_resultado = st.number_input(
                        "Resultado (Bq/m³)", step=1.0, key=key_resultado,
                        on_change=_guardar_resultado_cambio,
                    )
                with f4:
                    valor_incerteza = st.text_input(
                        "Incertidumbre", placeholder="p.ej. ±15% (k=2)", key=key_incerteza,
                        on_change=_guardar_resultado_cambio,
                    )
                st.markdown("---")

                resultado_valores.append(valor_resultado)
                incerteza_valores.append(valor_incerteza)

        df_final = df_working_reset.copy()
        df_final["Resultado Bq/m3"] = resultado_valores
        df_final["Incerteza expandida e K"] = incerteza_valores

        pendientes = int(df_final["Resultado Bq/m3"].isna().sum()) if "Resultado Bq/m3" in df_final else 0
        if pendientes:
            st.info(f"ℹ️ Quedan {pendientes} detector(es) sin resultado.")

        exceeded = int((df_final["Resultado Bq/m3"] > 300).sum()) if "Resultado Bq/m3" in df_final else 0
        if exceeded:
            st.warning(f"⚠️ {exceeded} medición(es) superan el nivel de referencia de 300 Bq/m³.")
        elif not pendientes:
            st.success("Ninguna medición supera el nivel de referencia de 300 Bq/m³.")

    # =========================================================
    # 7. CONCLUSIONES
    # =========================================================
    completo_7 = pendientes == 0
    with _acordeon_informe("7", "CONCLUSIONES", completo_7):
        conclusion_default = generar_conclusion_automatica(df_final)
        _sincronizar_valor_auto(f"ic_conclusion_{cid}", conclusion_default)
        conclusion_manual = st.text_area(
            "Texto de conclusiones (editable)", height=200, key=f"ic_conclusion_{cid}",
        )
        conclusion_es_auto = st.session_state.get(f"ic_conclusion_{cid}__ultimo_auto") == conclusion_manual
        if not (punto3_es_auto and conclusion_es_auto and objeto_es_auto and texto4_es_auto and texto5_es_auto):
            st.caption(
                "✏️ Como has escrito a mano alguno de los textos (objeto, punto 3, "
                "conclusiones...), ese texto se usará igual en gallego y en castellano si "
                "generas también la versión en castellano (no se traduce automáticamente)."
            )

    # =========================================================
    # 8. FIRMA
    # =========================================================
    tecnico_previo = str(tecnico_centro or det_meta.get("Técnico", "") or "")
    completo_8 = bool(st.session_state.get(f"ic_tecnico_{cid}", tecnico_previo).strip())
    with _acordeon_informe("8", "FIRMA", completo_8):
        tecnico_nome = st.text_input(
            "Nombre del/de la técnico/a que firma el informe",
            value=tecnico_previo, key=f"ic_tecnico_{cid}",
        )

    # =========================================================
    # 9. ANEXOS
    # =========================================================
    with _acordeon_informe("9", "ANEXOS", True):
        st.caption(
            "Documentos de cada anexo (PDF, Word o imagen). Si subes alguno, se genera también "
            "un ZIP con el informe y los anexos juntos."
        )

        ANEXOS_INFO = [
            ("anexo1", "ANEXO I: FORMULARIOS TOMA DE DATOS"),
            ("anexo2", "ANEXO II: ESQUEMA GRÁFICO DO EDIFICIO E PLANOS DE CADA PLANTA"),
            ("anexo3", "ANEXO III: INFORME DE ENSAIO DO LABORATORIO ACREDITADO"),
            ("anexo4", "ANEXO IV: CERTIFICADO ENAC DO LABORATORIO ACREDITADO"),
        ]
        ANEXOS_POR_DEFECTO = {"anexo3": anexo3_por_defecto, "anexo4": anexo4_por_defecto}

        anexos_datos = {}

        def _titulo_anexo(texto):
            """Título de cada apartado de anexo (I, II, III, IV): mismo
            tamaño y en amarillo para los cuatro por igual.

            OJO: no vale con un "style" en línea con !important aquí (el
            propio st.markdown sanea el HTML y elimina los "!important" de
            los estilos en línea); por eso se usa el mismo patrón de
            "marcador" + regla CSS en el bloque <style> global de toda la
            app, que sí conserva el !important."""
            st.markdown('<div class="marcador-titulo-anexo"></div>', unsafe_allow_html=True)
            st.markdown(f'<p style="font-weight:700;">{html.escape(texto)}</p>', unsafe_allow_html=True)

        for key, label in ANEXOS_INFO:
            key_cid = f"{key}_{cid}"
            if key == "anexo2":
                _titulo_anexo(label)
                modo_anexo2 = st.radio(
                    "¿Cómo quieres aportar este anexo?",
                    options=["Generar automáticamente a partir del Excel", "Subir otro plano"],
                    key=f"ic_modo_anexo2_{cid}", horizontal=True, label_visibility="collapsed",
                )
                if modo_anexo2 == "Generar automáticamente a partir del Excel":
                    if st.button("🗺️ Generar plano automáticamente", key=f"ic_generar_anexo2_{cid}"):
                        try:
                            datos_planos = extraer_datos_planos(ruta_excel)
                            anexo2_docx = generar_documento_anexo2(datos_planos, logo_bytes=logo_bytes, centro=centro_nombre)
                            nombre_generado = f"ANEXO_II_planos_{selected_value}.docx"
                            st.session_state[f"ic_anexo2_guardado_{cid}"] = (nombre_generado, anexo2_docx)
                            # Miniaturas de cada plano YA con los puntos
                            # dibujados encima (la misma composición que se
                            # mete dentro del Anexo II), para poder verlas
                            # aquí mismo sin tener que abrir el .docx.
                            from utils_informe.anexo2 import componer_plano
                            miniaturas_planos = []
                            for clave_plano, info_plano in datos_planos.get("planos", {}).items():
                                puntos_plano = datos_planos.get("puntos", {}).get(clave_plano, [])
                                try:
                                    miniaturas_planos.append((
                                        info_plano.get("nombre", clave_plano),
                                        componer_plano(info_plano["imagen"], puntos_plano),
                                    ))
                                except Exception:
                                    pass
                            st.session_state[f"ic_anexo2_miniaturas_{cid}"] = miniaturas_planos
                            st.success("Plano generado correctamente.")
                        except Anexo2Error as e:
                            st.error(f"No se ha podido generar el plano automáticamente: {e}")
                    if st.session_state.get(f"ic_anexo2_guardado_{cid}"):
                        nombre_guardado, _ = st.session_state[f"ic_anexo2_guardado_{cid}"]
                        anexos_datos["anexo2"] = st.session_state[f"ic_anexo2_guardado_{cid}"]
                        st.caption(f"📎 Plano disponible: **{nombre_guardado}**")
                        for nombre_plano_mini, imagen_mini in st.session_state.get(f"ic_anexo2_miniaturas_{cid}", []):
                            st.image(imagen_mini, width=200, caption=nombre_plano_mini)
                    else:
                        anexos_datos["anexo2"] = None
                    continue
                else:
                    st.session_state.pop(f"ic_anexo2_guardado_{cid}", None)
                    st.session_state.pop(f"ic_anexo2_miniaturas_{cid}", None)

            default_fn = ANEXOS_POR_DEFECTO.get(key)
            valor_por_defecto_bytes = default_fn() if default_fn else None
            _titulo_anexo(label)
            anexos_datos[key] = _widget_archivo_con_eliminar(
                f"ic_{key_cid}", "Sube el archivo",
                ["pdf", "doc", "docx", "jpg", "jpeg", "png"],
                valor_por_defecto=valor_por_defecto_bytes,
            )

    # =========================================================
    # GENERACIÓN DEL INFORME (sin número)
    # =========================================================
    with _acordeon_informe("", "GENERACIÓN DEL INFORME", True):
        col_opt1, col_opt2 = st.columns(2)
        with col_opt1:
            generar_castellano = st.checkbox("Generar también en castellano (.docx)", key=f"ic_castellano_{cid}")
        with col_opt2:
            pdf_disponible = libreoffice_disponible()
            generar_pdf_completo = st.checkbox(
                "Generar también un PDF completo (informe + anexos)",
                disabled=not pdf_disponible, key=f"ic_pdf_completo_{cid}",
            )
            if not pdf_disponible:
                st.caption("⚠️ LibreOffice no está disponible en este entorno; no se puede generar el PDF completo.")

        st.markdown('<div class="marcador-btn-guardar-detector"></div>', unsafe_allow_html=True)
        if st.button("📄 Generar informe", type="primary", use_container_width=True, key=f"ic_generar_{cid}"):
            try:
                ctx_gl = ReportContext(
                    xerencia=xerencia, cif=cif, centro=traducir_es_gl(centro_nombre), servizo_unidade=traducir_es_gl(servizo_unidade_efectivo),
                    enderezo=enderezo, superficie_construida=superficie_construida, superficie_util=superficie_util,
                    num_plantas=num_plantas, postos_bullets=postos_bullets, num_traballadores_total=num_traballadores_total,
                    categorias_bullets=categorias_bullets, texto_punto3_manual=texto_punto3_manual,
                    data_informacion_traballadores=data_informacion_traballadores,
                    medio_informacion_traballadores=medio_informacion_traballadores,
                    data_informe=data_informe.strftime("%d/%m/%Y"), tecnico_nome=tecnico_nome,
                    incertezas_por_defecto="", conclusion_manual=conclusion_manual, logo=logo_bytes,
                    tipo_zona=tipo_zona_elegido, objeto_manual=objeto_manual,
                    texto4_manual=texto4_manual, texto5_manual=texto5_manual,
                )
                buffer_gl = generate_report(ctx_gl, df_final, idioma="gl")
                report_name_gl = _nombre_documento(selected_value, "INFORME-FINAL", "-gl") + ".docx"
                st.session_state[f"ic_report_gl_{cid}"] = (buffer_gl.getvalue(), report_name_gl)

                docx_buffers_para_zip = [(report_name_gl, buffer_gl.getvalue())]
                docx_buffers_para_pdf = [("gl", report_name_gl, buffer_gl.getvalue())]

                if generar_castellano:
                    postos_bullets_es = postos_traballo_bullets(
                        df_center, traducir_galego=False, es_atencion_primaria=es_ap_centro,
                    )
                    _, categorias_bullets_es = categorias_turnos_bullets(
                        categorias_df, traducir_galego=False, es_atencion_primaria=es_ap_centro,
                    )

                    # El objeto, el punto 3 y las conclusiones son
                    # cuadros de texto editables (uno solo, no uno por
                    # idioma): si el usuario no los ha tocado (siguen
                    # en modo automático), se recalculan aquí en
                    # castellano en vez de reutilizar el texto en
                    # gallego tal cual, que es lo que causaba que
                    # saliera parte del informe en el idioma que no
                    # tocaba. Si el usuario SÍ escribió algo a mano, se
                    # deja igual (no hay forma de saber en qué idioma
                    # lo escribió).
                    objeto_es = objeto_manual
                    if objeto_es_auto:
                        _ctx_previa_objeto_es = ReportContext(
                            centro=centro_nombre, servizo_unidade=servizo_unidade_efectivo, tipo_zona=tipo_zona_elegido,
                        )
                        objeto_es = generar_texto_objeto_automatico(_ctx_previa_objeto_es, idioma="es")

                    texto_punto3_es = texto_punto3_manual
                    if punto3_es_auto:
                        _ctx_previa_punto3_es = ReportContext(
                            centro=centro_nombre, servizo_unidade=servizo_unidade_efectivo, postos_bullets=postos_bullets_es,
                            num_traballadores_total=num_traballadores_total, categorias_bullets=categorias_bullets_es,
                            medio_informacion_traballadores=medio_informacion_traballadores,
                            data_informacion_traballadores=data_informacion_traballadores,
                            tipo_zona=tipo_zona_elegido,
                        )
                        texto_punto3_es = generar_texto_punto3_automatico(_ctx_previa_punto3_es, idioma="es")

                    conclusion_es = conclusion_manual
                    if conclusion_es_auto:
                        conclusion_es = generar_conclusion_automatica(df_final, idioma="es")

                    texto4_es = generar_texto4_automatico("es") if texto4_es_auto else texto4_manual
                    texto5_es = generar_texto5_automatico("es") if texto5_es_auto else texto5_manual

                    ctx_es = ReportContext(
                        xerencia=xerencia, cif=cif, centro=centro_nombre, servizo_unidade=servizo_unidade_efectivo,
                        enderezo=enderezo, superficie_construida=superficie_construida, superficie_util=superficie_util,
                        num_plantas=num_plantas, postos_bullets=postos_bullets_es, num_traballadores_total=num_traballadores_total,
                        categorias_bullets=categorias_bullets_es, texto_punto3_manual=texto_punto3_es,
                        data_informacion_traballadores=data_informacion_traballadores,
                        medio_informacion_traballadores=medio_informacion_traballadores,
                        data_informe=data_informe.strftime("%d/%m/%Y"), tecnico_nome=tecnico_nome,
                        incertezas_por_defecto="", conclusion_manual=conclusion_es, logo=logo_bytes,
                        tipo_zona=tipo_zona_elegido, objeto_manual=objeto_es,
                        texto4_manual=texto4_es, texto5_manual=texto5_es,
                    )
                    buffer_es = generate_report(ctx_es, df_final, idioma="es")
                    report_name_es = _nombre_documento(selected_value, "INFORME-FINAL", "-es") + ".docx"
                    st.session_state[f"ic_report_es_{cid}"] = (buffer_es.getvalue(), report_name_es)
                    docx_buffers_para_zip.append((report_name_es, buffer_es.getvalue()))
                    docx_buffers_para_pdf.append(("es", report_name_es, buffer_es.getvalue()))
                else:
                    st.session_state.pop(f"ic_report_es_{cid}", None)

                anexos_subidos = {k: v for k, v in anexos_datos.items() if v is not None}
                numeros = {"anexo1": "I", "anexo2": "II", "anexo3": "III", "anexo4": "IV"}

                if anexos_subidos:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for nombre_docx, contenido_docx in docx_buffers_para_zip:
                            zf.writestr(nombre_docx, contenido_docx)
                        for key, (nombre_original, contenido) in anexos_subidos.items():
                            ext = nombre_original.rsplit(".", 1)[-1] if "." in nombre_original else "pdf"
                            zf.writestr(f"ANEXO_{numeros[key]}.{ext}", contenido)
                    nombre_zip_final = _nombre_documento(selected_value, "INFORME-FINAL", "-con-anexos") + ".zip"
                    st.session_state[f"ic_zip_{cid}"] = (zip_buffer.getvalue(), nombre_zip_final)
                else:
                    st.session_state.pop(f"ic_zip_{cid}", None)

                if generar_pdf_completo and pdf_disponible:
                    anexos_para_pdf = [(n, c) for (n, c) in anexos_subidos.values()]
                    for idioma_pdf, nombre_docx, contenido_docx in docx_buffers_para_pdf:
                        try:
                            pdf_bytes = construir_pdf_completo(contenido_docx, anexos_para_pdf)
                            pdf_name = nombre_docx.rsplit(".", 1)[0] + "_completo.pdf"
                            st.session_state[f"ic_pdf_{idioma_pdf}_{cid}"] = (pdf_bytes, pdf_name)
                        except PdfToolsError as e:
                            st.error(f"No se ha podido generar el PDF completo ({idioma_pdf}): {e}")
                else:
                    st.session_state.pop(f"ic_pdf_gl_{cid}", None)
                    st.session_state.pop(f"ic_pdf_es_{cid}", None)

                st.success("Informe generado correctamente")
            except Exception as e:
                st.error(f"Error al generar el informe: {e}")

        for key_ss, etiqueta, mime in [
            (f"ic_report_gl_{cid}", "⬇️ Descargar informe en gallego (.docx)",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            (f"ic_report_es_{cid}", "⬇️ Descargar informe en castellano (.docx)",
             "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            (f"ic_pdf_gl_{cid}", "⬇️ Descargar PDF completo en gallego", "application/pdf"),
            (f"ic_pdf_es_{cid}", "⬇️ Descargar PDF completo en castellano", "application/pdf"),
            (f"ic_zip_{cid}", "⬇️ Descargar informe(s) + anexos (.zip)", "application/zip"),
        ]:
            if key_ss in st.session_state:
                contenido, nombre_archivo = st.session_state[key_ss]
                st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                st.download_button(etiqueta, data=contenido, file_name=nombre_archivo, mime=mime, key=f"dl_{key_ss}")


def pantalla_centro_informes():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.session_state.view = "inicio"
        st.rerun()
        return
    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro
    ns_centro = f"centro_{cid}"

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    if st.button("← Volver"):
        st.session_state.view = "centro"
        st.rerun()

    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="subtitulo-amarillo">📄 Informes y descargas</p>', unsafe_allow_html=True)

    detectores = fetch_detectores(cid)
    planos_centro = fetch_planos_centro(cid)

    with st.container(border=True):
      if not detectores:
          st.caption("Añade al menos un detector para poder generar los archivos.")
      else:
          # Recopilar todas las fotos del informe (imagen exterior del
          # centro y plano-con-punto/situación/detector de cada uno).
          # Se usa tanto para el ZIP de descarga como para la checklist
          # de WhatsApp.
          fotos_disponibles = []
          if img_path and os.path.exists(img_path):
              ext = os.path.splitext(img_path)[1] or ".jpg"
              fotos_disponibles.append({
                  "ruta": img_path,
                  "nombre_archivo": f"exterior_{_slug(nombre)}{ext}",
                  "etiqueta": "Foto exterior del centro",
              })
          # Los planos se listan una sola vez cada uno (se comparten
          # entre varios detectores, así que no se repiten).
          for plano_c in planos_centro:
              _, _, nombre_plano_c, ruta_plano_c, _ = plano_c
              if ruta_plano_c and os.path.exists(ruta_plano_c):
                  ext = os.path.splitext(ruta_plano_c)[1] or ".jpg"
                  fotos_disponibles.append({
                      "ruta": ruta_plano_c,
                      "nombre_archivo": f"plano_{_slug(nombre_plano_c)}{ext}",
                      "etiqueta": f"Plano — {nombre_plano_c}",
                  })
          for d in detectores:
              did_d, codigo_d = d[0], d[5]
              foto_sit_d, foto_det_d = d[9], d[10]
              punto_x_d, punto_y_d = d[7], d[8]
              plano_centro_id_d = d[17]  # columna plano_centro_id (ver orden de init_db)
              base_id = codigo_d or f"Detector {did_d}"

              # Plano de ESTE detector, con su propio punto rojo dibujado
              # encima (no la imagen del plano "en blanco").
              if (plano_centro_id_d and punto_x_d is not None and punto_y_d is not None
                      and punto_x_d >= 0 and punto_y_d >= 0):
                  plano_info_d = get_plano_centro(plano_centro_id_d)
                  if plano_info_d:
                      plano_path_d = plano_info_d[3]
                      if es_url_supabase(plano_path_d):
                          temp_path = os.path.join(get_data_dir(), f"_tmp_plano_punto_det{did_d}.jpg")
                          if descargar_desde_supabase(plano_path_d, temp_path):
                              plano_path_d = temp_path
                      if os.path.exists(plano_path_d):
                          ruta_tmp_punto = os.path.join(
                              get_data_dir(), f"_tmp_plano_punto_det{did_d}.jpg"
                          )
                          if generar_plano_con_punto(plano_path_d, punto_x_d, punto_y_d, ruta_tmp_punto):
                              fotos_disponibles.append({
                                  "ruta": ruta_tmp_punto,
                                  "nombre_archivo": _nombre_foto_plano(codigo_d, nombre, zona) + ".jpg",
                                  "etiqueta": f"Plano con punto — {base_id}",
                              })

              if foto_sit_d:
                  if es_url_supabase(foto_sit_d):
                      # Se mostrará desde la URL
                      fotos_disponibles.append({
                          "ruta": foto_sit_d,
                          "nombre_archivo": _nombre_foto_situacion(codigo_d, nombre, zona) + ".jpg",
                          "etiqueta": f"Foto situación — {base_id}",
                      })
                  elif os.path.exists(foto_sit_d):
                      ext = os.path.splitext(foto_sit_d)[1] or ".jpg"
                      fotos_disponibles.append({
                          "ruta": foto_sit_d,
                          "nombre_archivo": _nombre_foto_situacion(codigo_d, nombre, zona) + ext,
                          "etiqueta": f"Foto situación — {base_id}",
                      })
              
              if foto_det_d:
                  if es_url_supabase(foto_det_d):
                      fotos_disponibles.append({
                          "ruta": foto_det_d,
                          "nombre_archivo": _nombre_foto_detector(codigo_d, nombre, zona) + ".jpg",
                          "etiqueta": f"Foto detector — {base_id}",
                      })
                  elif os.path.exists(foto_det_d):
                      ext = os.path.splitext(foto_det_d)[1] or ".jpg"
                      fotos_disponibles.append({
                          "ruta": foto_det_d,
                          "nombre_archivo": _nombre_foto_detector(codigo_d, nombre, zona) + ext,
                          "etiqueta": f"Foto detector — {base_id}",
                      })

          # --- Acceso al generador de informe completo (Word/PDF con el
          # modelo oficial UPRL/SERGAS), en una pantalla aparte. ---
          st.markdown('<div class="marcador-btn-informe-completo"></div>', unsafe_allow_html=True)
          if st.button("📝 Informe de resultados completo (Word)", use_container_width=True):
              st.session_state.view = "centro_informe_completo"
              st.rerun()
          st.caption(
              "Genera el informe oficial completo (identificación del centro, "
              "trabajadores expuestos, resultados y anexos) a partir de los datos "
              "de este centro, en formato Word."
          )
          st.markdown("---")

          # --- Checklist de qué documentos generar ---
          st.markdown(
              '<p class="subtitulo-amarillo">¿Qué quieres generar?</p>',
              unsafe_allow_html=True,
          )
          gen_pdf = st.checkbox("Informe PDF", value=False, key=f"gen_chk_pdf_{cid}")
          gen_excel = st.checkbox("Hoja Excel", value=False, key=f"gen_chk_excel_{cid}")
          gen_fotos = st.checkbox("Fotos", value=False, key=f"gen_chk_fotos_{cid}")
          gen_lab = st.checkbox("Registro para laboratorio", value=False, key=f"gen_chk_lab_{cid}")
          tipo_firma_lab = "digital"
          if gen_lab:
              tipo_firma_lab = st.radio(
                  "Tipo de firma del técnico",
                  options=["digital", "manual"],
                  format_func=lambda v: "Firma digital" if v == "digital" else "Firma a mano",
                  key=f"gen_firma_tipo_{cid}", horizontal=True,
              )
              logo_lab_actual = get_logo_laboratorio()
              if logo_lab_actual and os.path.exists(logo_lab_actual):
                  st.image(logo_lab_actual, width=180)
              nuevo_logo_lab = st.file_uploader(
                  "Logotipo del laboratorio (opcional, sustituye al que trae la app)",
                  type=["png", "jpg", "jpeg"], key=f"logo_lab_up_{cid}",
              )
              if nuevo_logo_lab is not None:
                  fid_logo = getattr(nuevo_logo_lab, "file_id", None) or \
                      f"{nuevo_logo_lab.name}_{nuevo_logo_lab.size}"
                  if st.session_state.get(f"logo_lab_last_fid_{cid}") != fid_logo:
                      ruta_logo_guardada = guardar_bytes_imagen(
                          nuevo_logo_lab.getvalue(), "logo_laboratorio", extension_de(nuevo_logo_lab),
                      )
                      set_logo_laboratorio(ruta_logo_guardada)
                      st.session_state[f"logo_lab_last_fid_{cid}"] = fid_logo
                      st.success("Logotipo del laboratorio actualizado")
                      st.rerun()

          # --- Un único botón que genera lo marcado, pequeño y
          # alineado a la izquierda ---
          col_gen, _col_resto = st.columns([1, 3])
          with col_gen:
              st.markdown('<div class="marcador-btn-generar"></div>', unsafe_allow_html=True)
              generar_clic = st.button(
                  "📦 Generar", type="tertiary", use_container_width=True,
              )

          if generar_clic:
              if not (gen_pdf or gen_excel or gen_fotos or gen_lab):
                  st.warning("Marca al menos un documento para generar.")
              else:
                  try:
                      marca_tiempo = _ahora_espana().strftime('%Y%m%d_%H%M%S')
                      with st.spinner("Generando documentos..."):
                          if gen_pdf:
                              nombre_pdf = _nombre_documento(nombre, "INFORME-COLOCACIÓN") + ".pdf"
                              ruta_pdf = os.path.join(get_data_dir(), nombre_pdf)
                              generar_pdf(cid, ruta_pdf)
                              st.session_state["ultimo_pdf"] = ruta_pdf
                              st.session_state["ultimo_pdf_nombre"] = nombre_pdf
                              st.session_state["ultimo_pdf_centro"] = cid

                          if gen_excel:
                              nombre_xlsx = _nombre_documento(nombre, "HOJA-DATOS") + ".xlsx"
                              ruta_xlsx = os.path.join(get_data_dir(), nombre_xlsx)
                              generar_excel(cid, ruta_xlsx)
                              st.session_state["ultimo_excel"] = ruta_xlsx
                              st.session_state["ultimo_excel_nombre"] = nombre_xlsx
                              st.session_state["ultimo_excel_centro"] = cid

                          if gen_fotos:
                              nombre_zip = _nombre_documento(nombre, "FOTOS") + ".zip"
                              ruta_zip = os.path.join(get_data_dir(), nombre_zip)
                              modo_fotos_gen = st.session_state.get(f"modo_fotos_{cid}", "todas")
                              if modo_fotos_gen == "individual":
                                  fotos_a_incluir = [
                                      foto for i, foto in enumerate(fotos_disponibles)
                                      if st.session_state.get(
                                          f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}")
                                  ]
                              else:
                                  fotos_a_incluir = fotos_disponibles
                              with zipfile.ZipFile(ruta_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                                  for foto in fotos_a_incluir:
                                      # Si es URL de Supabase, descargar primero
                                      ruta_local = foto["ruta"]
                                      if es_url_supabase(ruta_local):
                                          temp_file = os.path.join(get_data_dir(), f"_temp_zip_{len(fotos_a_incluir)}.jpg")
                                          if descargar_desde_supabase(ruta_local, temp_file):
                                              ruta_local = temp_file
                                          else:
                                              continue
                                      if os.path.exists(ruta_local):
                                          zf.write(ruta_local, arcname=foto["nombre_archivo"])
                              st.session_state["ultimo_zip_fotos"] = ruta_zip
                              st.session_state["ultimo_zip_fotos_nombre"] = nombre_zip
                              st.session_state["ultimo_zip_fotos_centro"] = cid

                          if gen_lab:
                              nombre_lab = _nombre_documento(nombre, "REGISTRO-LABORATORIO") + ".pdf"
                              ruta_lab = os.path.join(get_data_dir(), nombre_lab)
                              generar_registro_laboratorio(cid, ruta_lab, tipo_firma=tipo_firma_lab)
                              st.session_state["ultimo_lab"] = ruta_lab
                              st.session_state["ultimo_lab_nombre"] = nombre_lab
                              st.session_state["ultimo_lab_centro"] = cid

                      st.success("Documentos generados correctamente")
                  except Exception as e:
                      st.error(f"Error al generar los documentos: {e}")

          hay_pdf = st.session_state.get("ultimo_pdf_centro") == cid
          hay_excel = st.session_state.get("ultimo_excel_centro") == cid
          hay_zip = st.session_state.get("ultimo_zip_fotos_centro") == cid
          hay_lab = st.session_state.get("ultimo_lab_centro") == cid

          if hay_pdf or hay_excel or hay_zip or hay_lab:
              ultimo_pdf = st.session_state.get("ultimo_pdf")
              ultimo_excel = st.session_state.get("ultimo_excel")
              ultimo_zip = st.session_state.get("ultimo_zip_fotos")
              ultimo_lab = st.session_state.get("ultimo_lab")

              # 1) Descargar PDF
              if hay_pdf and ultimo_pdf and os.path.exists(ultimo_pdf):
                  with open(ultimo_pdf, "rb") as f:
                      pdf_bytes = f.read()
                  st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                  st.download_button(
                      "Descargar informe PDF", data=pdf_bytes,
                      file_name=st.session_state.get("ultimo_pdf_nombre", "informe.pdf"),
                      mime="application/pdf", use_container_width=True,
                      icon=":material/picture_as_pdf:",
                  )

              # 2) Enviar PDF por WhatsApp
              if hay_pdf and ultimo_pdf and os.path.exists(ultimo_pdf):
                  texto_wa = f"Informe de colocación de detectores de Rn – {nombre or ''}"
                  boton_compartir_whatsapp(
                      ultimo_pdf,
                      st.session_state.get("ultimo_pdf_nombre", "informe.pdf"),
                      texto_wa,
                  )

              # 3) Descargar Excel
              if hay_excel and ultimo_excel and os.path.exists(ultimo_excel):
                  with open(ultimo_excel, "rb") as f:
                      excel_bytes = f.read()
                  st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                  st.download_button(
                      "Descargar Excel", data=excel_bytes,
                      file_name=st.session_state.get("ultimo_excel_nombre", "detectores.xlsx"),
                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      use_container_width=True,
                      icon=":material/grid_on:",
                  )

              # 4) Enviar Excel por WhatsApp
              if hay_excel and ultimo_excel and os.path.exists(ultimo_excel):
                  texto_wa_excel = f"Hoja de cálculo de detectores de Rn – {nombre or ''}"
                  boton_compartir_whatsapp_excel(
                      ultimo_excel,
                      st.session_state.get("ultimo_excel_nombre", "detectores.xlsx"),
                      texto_wa_excel,
                  )

              # 5) Descargar Registro para laboratorio
              if hay_lab and ultimo_lab and os.path.exists(ultimo_lab):
                  with open(ultimo_lab, "rb") as f:
                      lab_bytes = f.read()
                  st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                  st.download_button(
                      "Descargar registro para laboratorio", data=lab_bytes,
                      file_name=st.session_state.get("ultimo_lab_nombre", "registro_laboratorio.pdf"),
                      mime="application/pdf", use_container_width=True,
                      icon=":material/picture_as_pdf:",
                  )

              # 6) Enviar Registro para laboratorio por WhatsApp
              if hay_lab and ultimo_lab and os.path.exists(ultimo_lab):
                  texto_wa_lab = f"Registro para laboratorio – {nombre or ''}"
                  boton_compartir_whatsapp(
                      ultimo_lab,
                      st.session_state.get("ultimo_lab_nombre", "registro_laboratorio.pdf"),
                      texto_wa_lab,
                      id_sufijo="lab",
                  )

              # 7) Fotos: por defecto se incluyen todas; el botón
              # "Selección individual" muestra la checklist para
              # elegir solo algunas (y así no alargar la pantalla
              # con el listado completo salvo que se pida).
              seleccionadas = []
              if fotos_disponibles:
                  st.markdown(
                      '<p class="subtitulo-amarillo">📷 Fotos del informe</p>',
                      unsafe_allow_html=True,
                  )

                  modo_key = f"modo_fotos_{cid}"
                  if modo_key not in st.session_state:
                      st.session_state[modo_key] = "todas"
                  modo_actual = st.session_state[modo_key]

                  st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                  etiqueta_toggle = (
                      "🔎 Selección individual" if modo_actual == "todas" else "☑️ Seleccionar todos"
                  )
                  if st.button(etiqueta_toggle, key=f"toggle_modo_fotos_{cid}", use_container_width=True):
                      if modo_actual == "todas":
                          st.session_state[modo_key] = "individual"
                      else:
                          st.session_state[modo_key] = "todas"
                          for i, foto in enumerate(fotos_disponibles):
                              st.session_state[
                                  f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}"
                              ] = True
                      st.rerun()

                  if st.session_state[modo_key] == "individual":
                      for i, foto in enumerate(fotos_disponibles):
                          marcado = st.checkbox(
                              foto["etiqueta"],
                              key=f"wa_chk_{cid}_{i}_{_slug(foto['nombre_archivo'])}",
                          )
                          if marcado:
                              seleccionadas.append((foto["ruta"], foto["nombre_archivo"]))
                  else:
                      st.caption(
                          f"Se incluirán las {len(fotos_disponibles)} fotos disponibles. "
                          "Pulsa «Selección individual» para elegir solo algunas."
                      )
                      seleccionadas = [(f["ruta"], f["nombre_archivo"]) for f in fotos_disponibles]

              # 8) Descargar fotos
              if hay_zip and ultimo_zip and os.path.exists(ultimo_zip):
                  with open(ultimo_zip, "rb") as f:
                      zip_bytes = f.read()
                  st.markdown('<div class="marcador-btn-descarga-amarillo"></div>', unsafe_allow_html=True)
                  st.download_button(
                      "⬇️ Descargar fotos", data=zip_bytes,
                      file_name=st.session_state.get("ultimo_zip_fotos_nombre", "fotos.zip"),
                      mime="application/zip", use_container_width=True,
                  )

              # 9) Enviar fotos seleccionadas por WhatsApp
              if fotos_disponibles:
                  texto_wa_fotos = f"Fotos del informe – {nombre or ''}"
                  boton_compartir_whatsapp_fotos(seleccionadas, texto_wa_fotos)



def pantalla_centro():
    cid = st.session_state.centro_actual
    centro = get_centro(cid) if cid else None
    if not centro:
        st.error("Centro no encontrado.")
        st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
        if st.button("← Inicio"):
            st.session_state.view = "inicio"
            st.rerun()
        return

    cid, nombre, zona, fecha, img_path, tecnico_centro, direccion = centro

    top1, top2 = st.columns([3, 1])
    with top1:
        st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
        if st.button("← Inicio"):
            st.session_state[f"pedir_descarga_salida_{cid}"] = True
            st.rerun()
    with top2:
        if st.button("Datos de la empresa", use_container_width=True):
            st.session_state.view = "ajustes"
            st.rerun()

    if st.session_state.get(f"pedir_descarga_salida_{cid}"):
        with st.container(border=True):
            st.warning("¿Quieres descargar los datos de este centro antes de salir?")
            nombre_xlsx_salida = _nombre_documento(nombre, "HOJA-DATOS") + ".xlsx"
            ruta_xlsx_salida = os.path.join(get_data_dir(), "_tmp_" + nombre_xlsx_salida)
            try:
                generar_excel(cid, ruta_xlsx_salida)
                with open(ruta_xlsx_salida, "rb") as _f_xlsx_salida:
                    xlsx_bytes_salida = _f_xlsx_salida.read()
                cd1, cd2 = st.columns(2)
                with cd1:
                    descargado = st.download_button(
                        "⬇️ Descargar y salir", data=xlsx_bytes_salida, file_name=nombre_xlsx_salida,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary",
                    )
                    if descargado:
                        st.session_state[f"pedir_descarga_salida_{cid}"] = False
                        st.session_state.view = "inicio"
                        st.rerun()
                with cd2:
                    if st.button("Salir sin descargar", use_container_width=True):
                        st.session_state[f"pedir_descarga_salida_{cid}"] = False
                        st.session_state.view = "inicio"
                        st.rerun()
            except Exception as e:
                st.error(f"No se ha podido preparar la descarga: {e}")
                if st.button("Salir de todos modos"):
                    st.session_state[f"pedir_descarga_salida_{cid}"] = False
                    st.session_state.view = "inicio"
                    st.rerun()
            if st.button("Cancelar (seguir aquí)"):
                st.session_state[f"pedir_descarga_salida_{cid}"] = False
                st.rerun()
        return

    st.markdown(
        '<p style="color:#999999; font-size:0.85rem; font-weight:700; margin:0;">'
        'Gestión de Detectores</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<p class="titulo-centro">🏢 {html.escape(nombre or "")}'
        f'{" · " + html.escape(zona) if zona and zona.strip() else ""}</p>',
        unsafe_allow_html=True,
    )

    # Los seis apartados del centro, cada uno en su propia pantalla
    # (con su botón de "← Volver" dentro), en vez del acordeón
    # de antes: con muchos detectores/planos la pantalla se hacía muy
    # larga y costaba encontrar cada bloque.
    opciones_menu = [
        ("Datos del centro", "centro_datos"),
        ("Categorías profesionales", "centro_categorias"),
        ("Planos del centro", "centro_planos"),
        ("Detectores colocados", "centro_detectores"),
        ("Retirada de detectores", "centro_retirada"),
        ("📄 Informes y descargas", "centro_informes"),
    ]
    for etiqueta, vista in opciones_menu:
        if st.button(etiqueta, key=f"menu_{vista}_{cid}", use_container_width=True):
            st.session_state.view = vista
            st.rerun()




# ============================================================
# PANTALLA: DETECTOR (nuevo / editar)
# ============================================================

TURNOS_TRABAJO_OPCIONES = ["Mañana", "Tarde", "Noche", "PAC", "Rotatorio", "Rotatorio complejo"]
# Turno a nivel de categoría profesional (independiente del turno por
# sala/detector): solo se usa para la segunda parte del punto 3 del
# informe ("O número de traballadores... divididos nas seguintes
# categorías"), sin relación con los turnos de los detectores.
TURNOS_CATEGORIA_OPCIONES = [
    "Mañana", "Tarde", "Noche", "Mañana/Tarde", "Mañana/Tarde/Noche",
    "Horario PAC", "Rotatorio", "Rotatorio complejo",
]
NIVEL_OPCIONES = [
    "3 niveles bajo rasante (Sótano -3)",
    "2 niveles bajo rasante (Sótano -2)",
    "1 nivel bajo rasante (Sótano -1)",
    "Nivel de la rasante (Planta Baja)",
    "1 nivel sobre rasante",
    "2 niveles sobre rasante",
    "3 niveles sobre rasante",
]
# Segundo bloque del código de sala (p.ej. "S-1", "PB", "02"): según
# la opción elegida en "Nivel".
NIVEL_A_CODIGO = {
    "3 niveles bajo rasante (Sótano -3)": "S-3",
    "2 niveles bajo rasante (Sótano -2)": "S-2",
    "1 nivel bajo rasante (Sótano -1)": "S-1",
    "Nivel de la rasante (Planta Baja)": "PB",
    "1 nivel sobre rasante": "01",
    "2 niveles sobre rasante": "02",
    "3 niveles sobre rasante": "03",
}
# Columna "Planta" de la ficha de registro para laboratorio: número
# de planta (con signo) que corresponde a cada opción de "Nivel".
NIVEL_A_PLANTA_LABORATORIO = {
    "3 niveles bajo rasante (Sótano -3)": "-3",
    "2 niveles bajo rasante (Sótano -2)": "-2",
    "1 nivel bajo rasante (Sótano -1)": "-1",
    "Nivel de la rasante (Planta Baja)": "0",
    "1 nivel sobre rasante": "1",
    "2 niveles sobre rasante": "2",
    "3 niveles sobre rasante": "3",
}
TIPO_CENTRO_OPCIONES = [
    "Atención Primaria", "PAC", "Atención Primaria + PAC", "Consultorio",
    "Centro de especialidades", "Hospital", "Otro",
]
# Para estos cuatro, la casilla "Área / Zona :" se rellena sola; para el
# resto (incluido "Otro") se deja en blanco para que se escriba a mano.
# "Consultorio" es un caso especial: su Área/Zona es "Atención
# Primaria" (no "Consultorio", que solo se usa como tipo de centro
# para el prefijo del código de sala).
TIPO_CENTRO_A_AREA_AUTOMATICA = {
    "Atención Primaria": "Atención Primaria",
    "PAC": "PAC",
    "Atención Primaria + PAC": "Atención Primaria + PAC",
    "Consultorio": "Atención Primaria",
}


def pantalla_detector():
    """Solo se usa ya para el flujo de "➕ Nuevo detector" (pantalla
    aparte). Para editar un detector YA EXISTENTE, ver
    pantalla_centro_detectores(): ahora se hace en línea, debajo del
    selector, con guardado automático al cambiar de detector o salir."""
    cid = st.session_state.centro_actual
    detector_id = st.session_state.detector_actual
    ns = f"det_{detector_id or 'nuevo'}"
    st.session_state["detector_form_ns"] = ns

    _inicializar_ns_detector(cid, detector_id, ns)

    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    if st.button("← Volver"):
        _limpiar_namespace(ns)
        st.session_state.view = "centro_detectores"
        st.rerun()

    texto_titulo_detector = "Datos del detector" if detector_id else "Nuevo detector"
    st.markdown(
        f'<p class="titulo-home">{texto_titulo_detector}</p>',
        unsafe_allow_html=True,
    )

    _renderizar_campos_detector(cid, detector_id, ns)

    st.markdown("---")
    st.markdown('<div class="marcador-btn-guardar-detector"></div>', unsafe_allow_html=True)
    if st.button("💾 Guardar detector", type="primary", use_container_width=True):
        guardado_id = _guardar_detector_desde_ns(cid, detector_id, ns)
        if guardado_id:
            _limpiar_namespace(ns)
            st.session_state.view = "centro_detectores"
            st.rerun()


# ============================================================

def pantalla_ajustes():
    cid = st.session_state.get("centro_actual")
    centro = get_centro(cid) if cid else None

    st.markdown('<p class="titulo-home">Datos de la empresa</p>', unsafe_allow_html=True)
    st.markdown('<div class="marcador-btn-azul-claro"></div>', unsafe_allow_html=True)
    volver_clic_ajustes = st.button("← Volver")

    if not centro:
        if volver_clic_ajustes:
            st.session_state.view = "inicio"
            st.rerun()
        st.error("No hay ningún centro seleccionado.")
        return

    _, nombre_centro, _, _, _, tecnico_actual, _ = centro
    st.caption(f"Centro: {nombre_centro}")

    empresa_actual = get_empresa()
    cif_actual = get_cif()

    st.text_input(
        "Empresa",
        value=empresa_actual,
        key="ajustes_empresa",
    )
    st.text_input(
        "CIF",
        value=cif_actual,
        key="ajustes_cif",
    )
    st.text_input(
        "Técnico (aparece en el PDF de este centro)",
        value=tecnico_actual or "",
        key="ajustes_tecnico",
    )

    st.markdown("**Logotipo** (aparece en el informe final y en el Informe PDF de colocación de detectores)")
    logo_bytes_actual, logo_nombre_actual = get_logo_informe()
    _widget_archivo_con_eliminar(
        "ajustes_logo", "Logo (si no subes uno, se usa el de UPRL / SERGAS por defecto)",
        ["png", "jpg", "jpeg"],
        valor_por_defecto=(logo_nombre_actual, logo_bytes_actual) if logo_bytes_actual else None,
    )
    # Se guarda en la base de datos en cuanto cambia de verdad (se sube
    # uno nuevo o se elimina), comparando contra lo último que se
    # sincronizó -para no reescribir el archivo en cada repintado-.
    _logo_guardado_sesion = st.session_state.get("ajustes_logo_guardado")
    _ultimo_logo_sincronizado = st.session_state.get("_ultimo_logo_sincronizado")
    if _logo_guardado_sesion and _logo_guardado_sesion != _ultimo_logo_sincronizado:
        set_logo_informe(*_logo_guardado_sesion)
        st.session_state["_ultimo_logo_sincronizado"] = _logo_guardado_sesion
    elif st.session_state.get("ajustes_logo_eliminado") and _ultimo_logo_sincronizado is not None:
        set_logo_informe(None, None)
        st.session_state["_ultimo_logo_sincronizado"] = None

    hay_cambios_ajustes = (
        st.session_state["ajustes_empresa"] != (empresa_actual or "")
        or st.session_state["ajustes_cif"] != (cif_actual or "")
        or st.session_state["ajustes_tecnico"] != (tecnico_actual or "")
    )

    def _guardar_ajustes():
        set_empresa(st.session_state["ajustes_empresa"].strip())
        set_cif(st.session_state["ajustes_cif"].strip())
        set_tecnico_centro(cid, st.session_state["ajustes_tecnico"].strip())

    # Al pulsar "Volver", si hay cambios sin guardar se guardan solos
    # (sin pedir confirmación), en vez de perderlos sin más.
    if volver_clic_ajustes:
        if hay_cambios_ajustes:
            _guardar_ajustes()
        st.session_state.view = "centro"
        st.rerun()

    if hay_cambios_ajustes:
        st.markdown('<div class="marcador-btn-guardar-ajustes"></div>', unsafe_allow_html=True)
        if st.button("Guardar", type="primary"):
            _guardar_ajustes()
            st.session_state.view = "centro"
            st.rerun()

    # =========================================================
    # SECCIÓN SUPABASE - VERSIÓN MEJORADA CON DIAGNÓSTICO
    # =========================================================
    st.markdown("---")
    st.markdown("### ☁️ Supabase")

    # Botón para forzar recarga de secrets
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔄 Recargar configuración de Supabase", key="reload_supabase", use_container_width=True):
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()
    
    with col2:
        # Botón para mostrar diagnóstico detallado
        if st.button("🔍 Diagnóstico completo", key="diagnostico_supabase", use_container_width=True):
            st.session_state["mostrar_diagnostico"] = not st.session_state.get("mostrar_diagnostico", False)
            st.rerun()

    # ---- DIAGNÓSTICO DETALLADO ----
    if st.session_state.get("mostrar_diagnostico", False):
        with st.container(border=True):
            st.markdown("### 🔍 DIAGNÓSTICO DE SUPABASE")
            
            try:
                # 1. Verificar que el módulo esté disponible
                st.markdown("**1. Disponibilidad de Supabase:**")
                if SUPABASE_AVAILABLE:
                    st.success("✅ Supabase está disponible (librerías instaladas)")
                else:
                    st.error("❌ Supabase NO está disponible (faltan librerías)")
                    st.info("Ejecuta: pip install st-supabase-connection supabase httpx")
                
                # 2. Verificar secrets
                st.markdown("**2. Secretos disponibles:**")
                st.write("Claves en st.secrets:", list(st.secrets.keys()))
                
                if "connections" in st.secrets:
                    st.write("Claves en connections:", list(st.secrets["connections"].keys()))
                    
                    if "supabase" in st.secrets["connections"]:
                        st.success("✅ 'connections.supabase' encontrado")
                        
                        # Mostrar URL (parcial) y Key (parcial)
                        url = st.secrets["connections"]["supabase"]["SUPABASE_URL"]
                        key = st.secrets["connections"]["supabase"]["SUPABASE_KEY"]
                        
                        st.markdown("**3. Valores configurados:**")
                        st.write(f"📡 URL: `{url}`")
                        st.write(f"🔑 Key: `{key[:25]}...{key[-15:] if len(key) > 40 else ''}`")
                        
                        # Validar URL
                        st.markdown("**4. Validación de URL:**")
                        if url.startswith("https://"):
                            st.success("✅ URL comienza con https://")
                        else:
                            st.error("❌ URL debe comenzar con https://")
                        
                        if url.endswith(".supabase.co"):
                            st.success("✅ URL termina con .supabase.co")
                        else:
                            st.error("❌ URL debe terminar con .supabase.co")
                        
                        if url.endswith("/"):
                            st.error("❌ URL tiene barra al final - debe eliminarla")
                        else:
                            st.success("✅ URL no tiene barra al final")
                        
                        # Validar Key
                        st.markdown("**5. Validación de Key:**")
                        if len(key) > 50:
                            st.success("✅ Key tiene longitud adecuada")
                        else:
                            st.error("❌ Key parece demasiado corta")
                        
                        if key.startswith("eyJ"):
                            st.success("✅ Key comienza con 'eyJ' (formato JWT correcto)")
                        else:
                            st.error("❌ Key no comienza con 'eyJ' - puede ser incorrecta")
                        
                        # 6. Probar conexión
                        st.markdown("**6. Prueba de conexión:**")
                        try:
                            conn = get_supabase_connection()
                            if conn:
                                try:
                                    result = conn.table("centros").select("*").limit(1).execute()
                                    st.success(f"✅ Conexión exitosa! ({len(result.data)} centros encontrados)")
                                    if result.data:
                                        st.write("📋 Primer centro:", result.data[0])
                                    else:
                                        st.info("ℹ️ No hay centros en la base de datos")
                                except Exception as e:
                                    st.error(f"❌ Error en consulta: {e}")
                            else:
                                st.error("❌ No se pudo crear la conexión")
                        except Exception as e:
                            st.error(f"❌ Error general: {e}")
                    else:
                        st.error("❌ 'supabase' no está en connections")
                else:
                    st.error("❌ 'connections' no está en secrets")
                    st.info("💡 Asegúrate de que el archivo secrets.toml tenga la sección [connections.supabase]")
                    
            except Exception as e:
                st.error(f"❌ Error en diagnóstico: {e}")
                st.exception(e)

    # ---- CONEXIÓN NORMAL ----
    if SUPABASE_AVAILABLE:
        verificar_conexion_supabase()
    else:
        st.warning("⚠️ Supabase no está disponible. Instala las dependencias.")
        st.info("Asegúrate de tener instalado: st-supabase-connection>=0.2.0 y supabase>=2.0.0")
    
    # Botones de migración
    migrar_datos_a_supabase()


# ============================================================
# TRADUCCIÓN DE TEXTOS FIJOS EN INGLÉS (cámara, subida de archivos...)
# ============================================================
# Streamlit no permite traducir directamente los textos internos de
# algunos widgets nativos (p.ej. "Take Photo" del selector de cámara,
# o "Drag and drop file here" del subidor de archivos). Se inyecta un
# pequeño script que busca esos textos en la página y los sustituye por
# su equivalente en castellano, y los vuelve a aplicar cada vez que
# Streamlit repinta la interfaz. Si el navegador bloquea el acceso al
# documento padre (restricción de origen cruzado), simplemente no se
# traduce y el texto se queda en inglés, sin romper nada más.

def inyectar_traduccion_widgets():
    html = """
    <script>
    (function() {
      const traducciones = [
        ["Take Photo", "Sacar foto"],
        ["Clear photo", "Quitar foto"],
        ["Switch camera", "Cambiar cámara"],
        ["Drag and drop file here", "Arrastra el archivo aquí"],
        ["Drag and drop files here", "Arrastra los archivos aquí"],
        ["Browse files", "Buscar archivo"],
        ["Upload", "Subir archivo"],
        ["200MB per file", "200MB max."],
      ];

      function traducirNodo(nodo) {
        if (nodo.nodeType === Node.TEXT_NODE) {
          const original = nodo.nodeValue;
          for (const [en, es] of traducciones) {
            if (original.includes(en)) {
              nodo.nodeValue = original.split(en).join(es);
            }
          }
        } else if (nodo.nodeType === Node.ELEMENT_NODE) {
          nodo.childNodes.forEach(traducirNodo);
        }
      }

      // Campos de texto/número con algo escrito: fondo gris claro.
      // Vacíos: blanco (como siempre). Como escribir no cambia el
      // árbol del DOM (solo la propiedad "value"), hace falta mirar
      // cada campo con sus propios listeners de "input"/"change",
      // además de revisarlos todos cada vez que Streamlit vuelve a
      // dibujar la pantalla (rerun), que es cuando aparecen campos
      // nuevos todavía sin su listener enganchado.
      const GRIS_CAMPO_RELLENO = "#C9C9C9";
      const ROSA_CAMPO_VACIO = "#FBE1E6";

      function aplicarFondoSegunValor(campo) {
        const tieneValor = campo.value !== undefined && campo.value !== null
          && String(campo.value).trim() !== "";
        // Hay una regla CSS con !important que fuerza el fondo blanco
        // en estas casillas; para poder pisarla hace falta marcar
        // este estilo en línea TAMBIÉN como "important" (si no, no
        // tiene efecto por más que se aplique correctamente).
        campo.style.setProperty(
          "background-color",
          tieneValor ? GRIS_CAMPO_RELLENO : ROSA_CAMPO_VACIO,
          "important",
        );
      }

      function engancharYAplicarCampos(raiz) {
        const campos = raiz.querySelectorAll(
          'input[type="text"], input[type="number"], input:not([type]), textarea'
        );
        campos.forEach(function(campo) {
          // Las celdas de las tablas editables (st.data_editor) NO
          // llevan este fondo gris/rosa: son de "glide-data-grid"
          // (clase gdg-input) y se pintan aparte, en negro sobre
          // blanco, para que el texto se pueda leer bien siempre.
          if (campo.classList.contains("gdg-input")) return;
          if (!campo.dataset.fondoSegunValor) {
            campo.dataset.fondoSegunValor = "1";
            campo.addEventListener("input", function() { aplicarFondoSegunValor(campo); });
            campo.addEventListener("change", function() { aplicarFondoSegunValor(campo); });
          }
          aplicarFondoSegunValor(campo);
        });
      }

      try {
        const doc = window.parent.document;
        traducirNodo(doc.body);
        engancharYAplicarCampos(doc.body);
        const observer = new MutationObserver(function() {
          traducirNodo(doc.body);
          engancharYAplicarCampos(doc.body);
        });
        observer.observe(doc.body, { childList: true, subtree: true, characterData: true });
      } catch (e) {
        // Si el navegador no permite acceder al documento padre, se
        // deja tal cual (seguirá en inglés en ese caso concreto).
      }
    })();
    </script>
    """
    components.html(html, height=0)


# ============================================================
# MAIN
# ============================================================

def _leer_clave_acceso():
    """Lee la clave de acceso desde un archivo de texto llamado
    "Asacec" (sin extensión), en la misma carpeta que este script. Se
    puede cambiar en cualquier momento editando ese archivo
    directamente desde GitHub (sin tocar nada del código), subiendo el
    cambio y esperando a que Streamlit Cloud vuelva a desplegar (o
    haciendo un git pull + reinicio si se usa en local/Termux).
    Devuelve None si el archivo no existe o está vacío."""
    ruta_clave = os.path.join(_carpeta_script, "Asacec")
    try:
        with open(ruta_clave, "r", encoding="utf-8") as f:
            clave = f.read().strip()
        return clave or None
    except FileNotFoundError:
        return None


def pantalla_login():
    st.markdown('<p class="titulo-home">🔒 Acceso restringido</p>', unsafe_allow_html=True)
    clave_correcta = _leer_clave_acceso()
    if clave_correcta is None:
        st.error(
            "No se ha configurado ninguna clave de acceso: falta el archivo «Asacec» "
            "en la carpeta de la app (o está vacío). Créalo con la clave que quieras "
            "usar como único contenido."
        )
        return
    st.markdown('<div class="marcador-btn-guardar-ajustes"></div>', unsafe_allow_html=True)
    clave_introducida = st.text_input("Clave de acceso", type="password", key="login_clave")
    if st.button("Entrar", type="primary", use_container_width=True):
        if clave_introducida == clave_correcta:
            st.session_state["autenticado"] = True
            st.rerun()
        else:
            st.error("Clave incorrecta.")


def main():
    if not st.session_state.get("autenticado"):
        pantalla_login()
        return

    init_db()
    inyectar_traduccion_widgets()

    if "view" not in st.session_state:
        st.session_state.view = "inicio"
    if "centro_actual" not in st.session_state:
        st.session_state.centro_actual = None
    if "detector_actual" not in st.session_state:
        st.session_state.detector_actual = None

    # Para que el botón "atrás" del móvil (o del navegador) vuelva a la
    # pantalla anterior DENTRO de la app, en vez de salir directamente
    # a la de inicio: se guarda la vista y el centro actuales en la
    # propia URL (con st.query_params). El navegador trata cada
    # cambio de la URL como un paso más en su historial, así que su
    # botón de atrás nativo puede recuperarlo.
    #
    # OJO: no basta con comparar la URL contra el session_state.view
    # ACTUAL para saber si el cambio viene de fuera (botón atrás) o de
    # dentro (un botón que acaba de cambiar de pantalla y ha llamado a
    # st.rerun() a mitad de guión, sin llegar a la sincronización de
    # más abajo en ESTA vuelta): en ese caso la URL se queda "atrasada"
    # un instante, y si se comparara contra la vista nueva parecería
    # (por error) que el usuario ha pulsado atrás, deshaciendo la
    # navegación que se acababa de pedir. Por eso se guarda aparte cuál
    # fue la ÚLTIMA URL que esta misma app escribió con éxito, y solo
    # se considera "el usuario ha pulsado atrás" cuando la URL cambia
    # respecto a ESO, no respecto al session_state.view más reciente.
    qp_view = st.query_params.get("view")
    qp_centro = st.query_params.get("centro")
    url_actual = (qp_view, qp_centro)
    ultima_url_propia = st.session_state.get("_ultima_url_sincronizada")
    if qp_view and ultima_url_propia is not None and url_actual != ultima_url_propia:
        st.session_state.view = qp_view
        st.session_state.centro_actual = int(qp_centro) if qp_centro and qp_centro.isdigit() else None

    view = st.session_state.view
    if view == "inicio":
        pantalla_inicio()
    elif view == "centro":
        pantalla_centro()
    elif view == "centro_datos":
        pantalla_centro_datos()
    elif view == "centro_categorias":
        pantalla_centro_categorias()
    elif view == "centro_planos":
        pantalla_centro_planos()
    elif view == "centro_detectores":
        pantalla_centro_detectores()
    elif view == "centro_retirada":
        pantalla_centro_retirada()
    elif view == "centro_informes":
        pantalla_centro_informes()
    elif view == "centro_informe_completo":
        pantalla_centro_informe_completo()
    elif view == "detector":
        pantalla_detector()
    elif view == "ajustes":
        pantalla_ajustes()
    else:
        st.session_state.view = "inicio"
        st.rerun()

    # Se sincroniza la URL con la vista que se acaba de mostrar (por si
    # ha cambiado durante esta ejecución), para que el botón de atrás
    # pueda volver a ella más adelante; y se recuerda como "la última
    # URL propia", para la comprobación de arriba en la próxima vuelta.
    nuevo_qp_view = st.session_state.view
    nuevo_qp_centro = str(st.session_state.centro_actual) if st.session_state.centro_actual else None
    if (st.query_params.get("view"), st.query_params.get("centro")) != (nuevo_qp_view, nuevo_qp_centro):
        st.query_params["view"] = nuevo_qp_view
        if nuevo_qp_centro:
            st.query_params["centro"] = nuevo_qp_centro
        elif "centro" in st.query_params:
            del st.query_params["centro"]
    st.session_state["_ultima_url_sincronizada"] = (nuevo_qp_view, nuevo_qp_centro)


if __name__ == "__main__":
    main()
